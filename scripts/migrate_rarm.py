"""
Migrate existing GRC_Tool.xlsm to new RARM structure:
  1. Insert 4 rows after row 3 (rows 4-7: CIA C/I/A objectief + aangevinkt teller)
  2. Insert 1 column after column D (column E: # Aangevinkt per control row)
  3. Update VBA constants + guards via win32com
  4. Call SyncRARMKolommen to populate CIA rows from DA sheet
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import time
import sys

XLSM = Path("data/Import/GRC_Tool.xlsm").resolve()

if not XLSM.exists():
    print(f"ERROR: file not found: {XLSM}")
    sys.exit(1)

# ── Stap 1: structurele wijzigingen via openpyxl ─────────────────────────────
print("Stap 1: RARM structuur aanpassen via openpyxl...")
wb = openpyxl.load_workbook(str(XLSM), keep_vba=True)

if "RARM" not in wb.sheetnames:
    print("ERROR: geen RARM sheet gevonden")
    sys.exit(1)

ws = wb["RARM"]

# Controleer of de migratie al is uitgevoerd (rij 4 zou "C Objectief" moeten hebben
# of rij 3 is de kwetsbaarheden rij en rij 8+ zijn controls)
row4_a = ws.cell(row=4, column=1).value
row8_a = ws.cell(row=8, column=1).value
if row4_a == "C Objectief":
    print("SKIP: RARM sheet is al gemigreerd (rij 4 = 'C Objectief')")
else:
    print(f"  Huidige rij 4, col A: {row4_a!r}")
    print(f"  Huidige rij 8, col A: {row8_a!r}")

    # Voeg 4 rijen in na rij 3 (nieuwe rijen 4-7)
    print("  Invoegen 4 rijen na rij 3...")
    ws.insert_rows(4, amount=4)

    # Voeg 1 kolom in na kolom D (nieuwe kolom E = # Aangevinkt)
    print("  Invoegen kolom E (# Aangevinkt)...")
    ws.insert_cols(5, amount=1)

    # Schrijf labels voor nieuwe rijen in kolom A
    labels = {4: "C Objectief", 5: "I Objectief", 6: "A Objectief", 7: "Aangevinkt"}
    for row_idx, label in labels.items():
        ws.cell(row=row_idx, column=1).value = label

    # Schrijf kolomkop voor nieuwe kolom E in rij 2
    ws.cell(row=2, column=5).value = "# Aangevinkt"

    # Schrijf COUNTIF-formule in kolom E voor elke controlrij (rij 8+)
    last_da_col = get_column_letter(6 + 100)
    for row_idx in range(8, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value:
            col_start = get_column_letter(6)
            ws.cell(row=row_idx, column=5).value = (
                f"=COUNTIF({col_start}{row_idx}:{last_da_col}{row_idx},CHAR(10004))"
            )

    # Update freeze panes
    ws.freeze_panes = "G8"

    print("  Opslaan...")
    wb.save(str(XLSM))
    print("  openpyxl wijzigingen opgeslagen.")

wb.close()

# ── Stap 2: VBA volledig vervangen via win32com ───────────────────────────────
# Lees de volledige VBA_CODE en RARM_SHEET_CODE strings uit build_grc.py
# en injecteer ze in de bestaande xlsm (vervangt de volledige module-code).
print("\nStap 2: VBA volledig vervangen via win32com...")
try:
    import win32com.client
except ImportError:
    print("win32com niet beschikbaar — VBA moet handmatig worden bijgewerkt.")
    sys.exit(0)

import re

# Extract VBA code strings from build_grc.py
build_src_path = Path("scripts/build_grc.py")
build_src = build_src_path.read_text(encoding="utf-8")

def extract_triple_quoted(src: str, var_name: str) -> str:
    """Extract content of triple-quoted string assigned to var_name.
    Handles both r'''content (VBA_CODE) and '''\\ (RARM_SHEET_CODE) formats.
    """
    # Format 1: VAR = r'''<content_starts_same_line>
    m = re.search(rf"{re.escape(var_name)}\s*=\s*r'''(.*?)\n'''", src, re.DOTALL)
    if m:
        return m.group(1)
    # Format 2: VAR = '''\<newline><content>
    m = re.search(rf"{re.escape(var_name)}\s*=\s*'''\\\n(.*?)\n'''", src, re.DOTALL)
    if m:
        return m.group(1)
    raise ValueError(f"Could not find {var_name} in build_grc.py")

vba_macros_code  = extract_triple_quoted(build_src, "VBA_CODE")
rarm_sheet_code  = extract_triple_quoted(build_src, "RARM_SHEET_CODE")

# Strip "Attribute VB_Name" lines (VBE metadata, not valid in AddFromString)
def strip_attr(code: str) -> str:
    return "\n".join(l for l in code.splitlines() if not l.startswith("Attribute VB_Name"))

vba_macros_code = strip_attr(vba_macros_code)
rarm_sheet_code = strip_attr(rarm_sheet_code)

xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = False
xl.DisplayAlerts = False

try:
    wb_com = xl.Workbooks.Open(str(XLSM))
    time.sleep(2)

    proj = wb_com.VBProject
    rarm_code_name = wb_com.Sheets("RARM").CodeName

    for comp in proj.VBComponents:
        if comp.Name == "GRC_Macros":
            print(f"  Vervangen: GRC_Macros ({comp.CodeModule.CountOfLines} regels -> nieuw)")
            comp.CodeModule.DeleteLines(1, comp.CodeModule.CountOfLines)
            comp.CodeModule.AddFromString(vba_macros_code)
        elif comp.Name == rarm_code_name:
            print(f"  Vervangen: RARM sheet module ({comp.Name})")
            comp.CodeModule.DeleteLines(1, comp.CodeModule.CountOfLines)
            comp.CodeModule.AddFromString(rarm_sheet_code)

    # ── Stap 3: SyncRARMKolommen uitvoeren om CIA-rijen te vullen ──────────────
    print("\nStap 3: SyncRARMKolommen uitvoeren...")
    try:
        xl.Run("SyncRARMKolommen")
        time.sleep(3)
        print("  SyncRARMKolommen uitgevoerd.")
    except Exception as e:
        print(f"  WAARSCHUWING: SyncRARMKolommen mislukt: {e}")
        print("  Voer de macro handmatig uit in Excel.")

    wb_com.Save()
    print("\nMigratie opgeslagen.")

except Exception as e:
    print(f"ERROR: {e}")
    raise
finally:
    try:
        wb_com.Close(SaveChanges=False)
    except Exception:
        pass
    xl.Quit()

print("\nMigratie voltooid!")
print("Controleer de RARM sheet: rijen 4-7 (CIA + aangevinkt), kolom E (# Aangevinkt)")
