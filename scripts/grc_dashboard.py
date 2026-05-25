"""
GRC Tool v0.3 — Management Dashboard & Werktool
Leest én schrijft GRC_Tool.xlsm vanuit Streamlit.

Starten:  streamlit run scripts/grc_dashboard.py
Bronbestand: data/Import/GRC_Tool.xlsm  (gevulde werkversie)
"""

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTEN
# ══════════════════════════════════════════════════════════════════════════════
_CONFIG_FILE  = Path(__file__).parent.parent / "data" / "dashboard_config.json"
_DEFAULT_XLSM = Path(__file__).parent.parent / "data" / "Import" / "GRC_Tool.xlsm"

# Assurance-niveaus en welke controls getoond worden per entiteitsniveau
ASSURANCE_FILTER = {
    "Basic":     ["Basic"],
    "Important": ["Basic", "Important"],
    "Essential": ["Basic", "Important", "Essential"],
}

LEVEL_LABEL = {1: "Laag", 2: "Gemiddeld", 3: "Hoog", 4: "Zeer Hoog", 5: "Kritiek"}
LEVEL_COLOR = {
    "Laag":      "#22c55e",
    "Gemiddeld": "#eab308",
    "Hoog":      "#f97316",
    "Zeer Hoog": "#ef4444",
    "Kritiek":   "#7c3aed",
    "—":         "#94a3b8",
}
CYFUN_COLORS = {
    "GV": "#6366f1", "ID": "#0ea5e9", "PR": "#22c55e",
    "DE": "#f59e0b", "RS": "#ef4444", "RC": "#8b5cf6",
}
STATUS_COLORS = {
    "Uitgevoerd":     "#22c55e",
    "Gepland":        "#eab308",
    "Niet uitvoeren": "#94a3b8",
    "":               "#e2e8f0",
}

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG-BESTAND (JSON) — pad-persistentie
# ══════════════════════════════════════════════════════════════════════════════
def _load_config() -> dict:
    if _CONFIG_FILE.exists():
        try:
            return json.loads(_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"xlsm_path": str(_DEFAULT_XLSM), "recent": []}

def _save_config(cfg: dict) -> None:
    _CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    _CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")

def _add_recent(cfg: dict, path: str) -> dict:
    recent = [p for p in cfg.get("recent", []) if p != path]
    cfg["recent"] = ([path] + recent)[:5]
    return cfg

# Initialiseer sessiestatus
if "xlsm_path" not in st.session_state:
    st.session_state["xlsm_path"] = _load_config().get("xlsm_path", str(_DEFAULT_XLSM))

# ══════════════════════════════════════════════════════════════════════════════
# SCHRIJF-FUNCTIE
# ══════════════════════════════════════════════════════════════════════════════
def _safe_write(ws, row: int, col: int, value) -> None:
    """
    Schrijf een waarde naar een cel, ook als die cel deel is van een merged range.
    Merged slave-cellen zijn read-only in openpyxl; unmerge eerst, schrijf dan.
    """
    # Zoek of deze cel in een merged range zit
    to_unmerge = [
        str(mr) for mr in list(ws.merged_cells.ranges)
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col
    ]
    for mr_str in to_unmerge:
        ws.unmerge_cells(mr_str)
    ws.cell(row=row, column=col).value = value


def write_to_xlsm(xlsm_path: str, updates: list) -> None:
    """
    Schrijft een lijst van cel-updates naar de xlsm en wist de cache.

    updates = [{"sheet": str, "row": int, "col": int, "value": any}, ...]
    Gebruikt _safe_write om merged-cell fouten te vermijden.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    for u in updates:
        ws = wb[u["sheet"]]
        _safe_write(ws, u["row"], u["col"], u["value"])
    wb.save(xlsm_path)
    wb.close()
    load_all.clear()

# ══════════════════════════════════════════════════════════════════════════════
# DATA LADEN (gecached per bestandspad)
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=60)
def load_all(xlsm_path: str) -> dict:
    """Laad alle relevante sheets uit de xlsm en geef terug als dict van DataFrames + scalars."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=True, data_only=True)
    result = {}

    # ── Config ───────────────────────────────────────────────────────────────
    ws = wb["Config"]
    result["cfg_org"]        = ws.cell(row=5,  column=4).value or ""
    result["cfg_dept"]       = ws.cell(row=6,  column=4).value or ""
    result["cfg_lang"]       = ws.cell(row=9,  column=4).value or "NL"
    result["cfg_assurance"]  = ws.cell(row=19, column=4).value or ""
    result["cfg_version"]    = ws.cell(row=14, column=4).value or ""
    result["cfg_updated"]    = ws.cell(row=15, column=4).value or ""
    result["cfg_updated_by"] = ws.cell(row=16, column=4).value or ""

    # ── Processen ─────────────────────────────────────────────────────────────
    # Kolom A (ID) is leeg in gevulde xlsm — filter op naam in kolom B
    ws = wb["Processes"]
    proc_rows = []
    proc_idx = 1
    for row in ws.iter_rows(min_row=6, values_only=True):
        naam = row[1] if len(row) > 1 else None
        if not naam:
            continue
        proc_rows.append({
            "ID":               row[0] if row[0] else f"P{proc_idx:03d}",
            "Naam":             naam,
            "Eigenaar":         row[3] if len(row) > 3 else None,
            "Integriteit":      row[5] if len(row) > 5  and isinstance(row[5],  (int, float)) else None,
            "Beschikbaarheid":  row[7] if len(row) > 7  and isinstance(row[7],  (int, float)) else None,
            "IA":               row[10] if len(row) > 10 else None,
            "DA":               row[11] if len(row) > 11 else None,
        })
        proc_idx += 1
    result["df_proc"] = pd.DataFrame(proc_rows)

    # ── Informatieassets ──────────────────────────────────────────────────────
    # Kolom A (ID) is leeg in gevulde xlsm — filter op naam in kolom B
    ws = wb["Information Assets"]
    ia_rows = []
    ia_idx = 1
    for row in ws.iter_rows(min_row=6, values_only=True):
        naam = row[1] if len(row) > 1 else None
        if not naam:
            continue
        ia_rows.append({
            "ID":                   row[0] if row[0] else f"A{ia_idx:03d}",
            "Naam":                 naam,
            "Eigenaar":             row[3] if len(row) > 3 else None,
            "Confidentialiteit":    row[5] if len(row) > 5 and isinstance(row[5], (int, float)) else None,
        })
        ia_idx += 1
    result["df_ia"] = pd.DataFrame(ia_rows)

    # ── Dependent Assets ─────────────────────────────────────────────────────
    # Kolom A (ID) is leeg in gevulde xlsm — filter op naam in kolom B
    # CIA-kolommen (1-based): vereist = 10/12/14 (idx 9/11/13)
    #                          objectief = 16/18/20 (idx 15/17/19)
    ws = wb["Dependent Assets"]
    da_rows = []
    da_name_to_row = {}   # DA-naam → Excel-rijnummer (voor terugschrijven objectieven)
    da_idx = 1
    xl_row = 6
    for row in ws.iter_rows(min_row=6, values_only=True):
        naam = row[1] if len(row) > 1 else None
        if naam:
            def _int(r, i): return r[i] if len(r) > i and isinstance(r[i], (int, float)) else None
            da_rows.append({
                "ID":           row[0] if row[0] else f"D{da_idx:03d}",
                "Naam":         naam,
                "Omschrijving": row[2] if len(row) > 2 else None,
                "Eigenaar":     row[3] if len(row) > 3 else None,
                # Security Requirements (VBA-ingevuld)
                "C_req": _int(row, 9),
                "I_req": _int(row, 11),
                "A_req": _int(row, 13),
                # Security Objectives (editeerbaar)
                "C_obj": _int(row, 15),
                "I_obj": _int(row, 17),
                "A_obj": _int(row, 19),
                # Commentaarvelden (editeerbaar)
                "Opmerkingen": row[6]  if len(row) > 6  and row[6]  else "",
                "Commentaar":  row[24] if len(row) > 24 and row[24] else "",
                "_row":  xl_row,
            })
            da_name_to_row[naam] = xl_row
            da_idx += 1
        xl_row += 1
    result["df_da"]            = pd.DataFrame(da_rows)
    result["da_name_to_row"]   = da_name_to_row

    # ── Verantwoordelijken ────────────────────────────────────────────────────
    ws = wb["Responsible Persons"]
    verant_rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if any(v for v in row):
            verant_rows.append(row)
    _vcols = ["ID","Naam","Voornaam","Functietitel","Dienst","Email","Telefoon","Rol","Opmerkingen"]
    if verant_rows:
        n = max(len(r) for r in verant_rows)
        _vcols_ext = (_vcols + [None]*n)[:n]
        verant_rows = [tuple(r) + (None,)*(n - len(r)) for r in verant_rows]
        result["df_verant"] = pd.DataFrame(verant_rows, columns=_vcols_ext).dropna(how="all")
    else:
        result["df_verant"] = pd.DataFrame(columns=_vcols)

    # ── RARM: controls + DA-kolommen ─────────────────────────────────────────
    ws = wb["RARM"]
    rarm_rows = list(ws.iter_rows(min_row=1, values_only=True))
    RARM_COL_DA = 4  # kolom E = index 4 (0-based)

    da_header_row  = rarm_rows[1] if len(rarm_rows) > 1 else []
    da_names_rarm  = [str(v) for v in da_header_row[RARM_COL_DA:] if v]
    n_da = len(da_names_rarm)

    # Controls (rij 4+): Control ID | Richtlijn | Assurance | Sleutelmaatregel | DA-checks
    controls = []
    total_controls = 0
    for row in rarm_rows[3:]:
        if not row[0]:
            continue
        total_controls += 1
        da_checks = {}
        for j, da_name in enumerate(da_names_rarm):
            col_idx = RARM_COL_DA + j
            da_checks[da_name] = (row[col_idx] == "✔") if col_idx < len(row) else False
        controls.append({
            "Control ID": row[0],
            "Richtlijn":  row[1],
            "Assurance":  row[2],
            "Sleutel":    row[3] == "✔" if row[3] else False,
            **{f"da_{da}": v for da, v in da_checks.items()},
        })
    result["df_rarm"]         = pd.DataFrame(controls)
    result["da_names"]        = da_names_rarm
    result["total_controls"]  = total_controls

    # ── Kwetsbaarheden ────────────────────────────────────────────────────────
    ws = wb["Kwetsbaarheden"]
    all_kw_rows = list(ws.iter_rows(min_row=1, values_only=True))
    vuln_names = [v for v in (all_kw_rows[1] if len(all_kw_rows) > 1 else [])
                  if v and v not in ("Control ID", "Richtlijn")]
    cia_c = [v == "✔" for v in (all_kw_rows[2] if len(all_kw_rows) > 2 else [])[2:]]
    cia_i = [v == "✔" for v in (all_kw_rows[3] if len(all_kw_rows) > 3 else [])[2:]]
    cia_a = [v == "✔" for v in (all_kw_rows[4] if len(all_kw_rows) > 4 else [])[2:]]
    vuln_check_counts = [0] * len(vuln_names)
    for row in all_kw_rows[5:]:
        if row[0]:
            for j, v in enumerate(row[2:2 + len(vuln_names)]):
                if v == "✔":
                    vuln_check_counts[j] += 1
    result["df_vuln"] = pd.DataFrame({
        "Kwetsbaarheid": vuln_names[:len(vuln_check_counts)],
        "C": cia_c[:len(vuln_check_counts)],
        "I": cia_i[:len(vuln_check_counts)],
        "A": cia_a[:len(vuln_check_counts)],
        "Aantal controls": vuln_check_counts,
    })

    # ── Risicobeheer ──────────────────────────────────────────────────────────
    if "Risicobeheer" in wb.sheetnames:
        ws = wb["Risicobeheer"]
        rb_rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and not str(row[0]).startswith("(leeg"):
                rb_rows.append({
                    "Control ID":           row[0],
                    "Richtlijn":            row[1],
                    "Assurance":            row[2],
                    "Dependent Asset":      row[3],
                    "Status":               row[4] or "",
                    "Datum":                row[5],
                    "Verantwoordelijke":    row[6] or "",
                    "Opmerkingen":          row[7] or "",
                    "_row":                 ws.min_row,  # placeholder, see below
                })
        # Hercompute werkelijke rijnummers
        rb_rows_with_idx = []
        r_idx = 3
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and not str(row[0]).startswith("(leeg"):
                rb_rows_with_idx.append({
                    "Control ID":           row[0],
                    "Richtlijn":            row[1],
                    "Assurance":            row[2],
                    "Dependent Asset":      row[3],
                    "Status":               row[4] or "",
                    "Datum":                row[5],
                    "Verantwoordelijke":    row[6] or "",
                    "Opmerkingen":          row[7] or "",
                    "_row":                 r_idx,
                })
            r_idx += 1
        result["df_risico"] = pd.DataFrame(rb_rows_with_idx)
    else:
        result["df_risico"] = pd.DataFrame(columns=[
            "Control ID","Richtlijn","Assurance","Dependent Asset",
            "Status","Datum","Verantwoordelijke","Opmerkingen","_row"])

    # ── Controles ─────────────────────────────────────────────────────────────
    if "Controles" in wb.sheetnames:
        ws = wb["Controles"]
        ctrl_rows = []
        r_idx = 3
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and not str(row[0]).startswith("(leeg"):
                ctrl_rows.append({
                    "DA":                   row[0],
                    "Control ID":           row[1],
                    "Maatregel":            row[2],
                    "Frequentie":           row[3] or "",
                    "Methode":              row[4] or "",
                    "Artefacten locatie":   row[5] or "",
                    "Laatste controle":     row[6],
                    "Volgende controle":    row[7],
                    "Status":               row[8] or "",
                    "_row":                 r_idx,
                })
            r_idx += 1
        result["df_controles"] = pd.DataFrame(ctrl_rows)
    else:
        result["df_controles"] = pd.DataFrame(columns=[
            "DA","Control ID","Maatregel","Frequentie","Methode",
            "Artefacten locatie","Laatste controle","Volgende controle","Status","_row"])

    # ── Acties ────────────────────────────────────────────────────────────────
    if "Acties" in wb.sheetnames:
        ws = wb["Acties"]
        act_rows = []
        r_idx = 3
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and not str(row[0]).startswith("(leeg"):
                act_rows.append({
                    "Actie ID":             row[0],
                    "Type":                 row[1] or "",
                    "DA":                   row[2] or "",
                    "Control ID":           row[3] or "",
                    "Omschrijving":         row[4] or "",
                    "Vervaldatum":          row[5],
                    "Verantwoordelijke":    row[6] or "",
                    "Status":               row[7] or "Open",
                    "Aangemaakt op":        row[8],
                    "_row":                 r_idx,
                })
            r_idx += 1
        result["df_acties"] = pd.DataFrame(act_rows)
    else:
        result["df_acties"] = pd.DataFrame(columns=[
            "Actie ID","Type","DA","Control ID","Omschrijving",
            "Vervaldatum","Verantwoordelijke","Status","Aangemaakt op","_row"])

    # ── Afwijkingen (optioneel) ────────────────────────────────────────────────
    afw_rows = []
    if "Afwijkingen" in wb.sheetnames:
        ws = wb["Afwijkingen"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:
                afw_rows.append({"Kwetsbaarheid": row[0], "CIA": row[1],
                                 "Control ID": row[2], "Reden": row[3]})
    result["df_afw"] = pd.DataFrame(afw_rows)

    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT OPMAAK
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="GRC Tool",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .kpi-box {
    background: #f8fafc; border-radius: 10px; padding: 18px 20px 12px;
    border-left: 5px solid #2563eb; text-align: center;
  }
  .kpi-val  { font-size: 2.4rem; font-weight: 700; color: #0f172a; line-height: 1.1; }
  .kpi-lbl  { font-size: 0.82rem; color: #64748b; margin-top: 4px; }
  .warn-box {
    background: #fef9c3; border-left: 5px solid #eab308;
    padding: 10px 14px; border-radius: 6px; margin-bottom: 8px;
    font-size: 0.88rem; color: #713f12;
  }
  .ok-box {
    background: #dcfce7; border-left: 5px solid #16a34a;
    padding: 10px 14px; border-radius: 6px; margin-bottom: 8px;
    font-size: 0.88rem; color: #14532d;
  }
  .err-box {
    background: #fee2e2; border-left: 5px solid #dc2626;
    padding: 10px 14px; border-radius: 6px; margin-bottom: 8px;
    font-size: 0.88rem; color: #7f1d1d;
  }
  .section-hdr {
    font-size: 1.05rem; font-weight: 700; color: #1e3a6e;
    border-bottom: 2px solid #2563eb; padding-bottom: 4px;
    margin: 18px 0 10px;
  }
  div[data-testid="metric-container"] { background:#f8fafc; border-radius:8px; padding:10px; }
  .stDataFrame thead tr th { background-color: #1e3a6e !important; color: white !important; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def kpi(col, val, label, color="#2563eb"):
    col.markdown(
        f'<div class="kpi-box" style="border-left-color:{color}">'
        f'<div class="kpi-val">{val}</div>'
        f'<div class="kpi-lbl">{label}</div></div>',
        unsafe_allow_html=True,
    )

def level_lbl(n):
    return LEVEL_LABEL.get(n, "—")

def completeness(df, cols):
    total = len(df) * len(cols)
    filled = sum(df[c].notna().sum() for c in cols if c in df.columns)
    return round(filled / total * 100) if total else 0

def next_actie_id(df_acties: pd.DataFrame) -> str:
    """Genereer volgend Actie ID: A-001, A-002, ..."""
    if df_acties.empty:
        return "A-001"
    ids = df_acties["Actie ID"].dropna().astype(str)
    nums = []
    for i in ids:
        m = re.search(r"(\d+)$", i)
        if m:
            nums.append(int(m.group(1)))
    return f"A-{(max(nums) + 1 if nums else 1):03d}"

def calc_volgende(laatste, frequentie: str):
    """Bereken de volgende controledatum o.b.v. frequentie."""
    if not laatste:
        return None
    if isinstance(laatste, str):
        try:
            laatste = datetime.strptime(laatste, "%d/%m/%Y").date()
        except Exception:
            return None
    if isinstance(laatste, datetime):
        laatste = laatste.date()
    freq_map = {
        "Maandelijks":    30,
        "Kwartaal":       91,
        "Halfjaarlijks":  183,
        "Jaarlijks":      365,
        "Ad hoc":         None,
    }
    delta = freq_map.get(frequentie)
    if delta is None:
        return None
    return laatste + timedelta(days=delta)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LADEN
# ══════════════════════════════════════════════════════════════════════════════
_active_path = st.session_state["xlsm_path"]
try:
    data = load_all(_active_path)
    load_ok = True
except Exception as e:
    load_ok = False
    load_err = str(e)
    data = {}


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.image(
        "https://upload.wikimedia.org/wikipedia/commons/thumb/6/65/Shield-security.svg/200px-Shield-security.svg.png",
        width=60,
    )
    st.title("GRC Tool")
    _assurance = data.get("cfg_assurance", "") if load_ok else ""
    _org       = data.get("cfg_org", "")       if load_ok else ""
    if _org:
        st.caption(_org)
    if _assurance:
        st.caption(f"Assurance niveau: **{_assurance}**")
    else:
        st.warning("⚠️ Assurance niveau niet ingesteld")
    st.divider()

    pagina = st.radio("Navigatie", [
        "🏠 Dashboard",
        "⚙️ Configuratie",
        "📋 Processen & Assets",
        "🎯 Maatregelen",
        "🔍 Periodieke Controles",
        "📋 Acties",
        "🛡️ Kwetsbaarheden",
        "🔧 Instellingen",
    ])
    st.divider()
    _active_file = Path(_active_path)
    if load_ok:
        st.success("✅ Bestand geladen")
        st.caption(f"📄 {_active_file.name}")
    else:
        st.error("❌ Laadfouten")
        st.caption(f"📄 {_active_file.name}")

if not load_ok and pagina != "🔧 Instellingen":
    st.error(f"Kan bestand niet laden: {load_err}")
    st.info("Ga naar **🔧 Instellingen** om een geldig .xlsm-bestand te selecteren.")
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if pagina == "🏠 Dashboard":
    org = data.get("cfg_org", "")
    st.title(f"🛡️ GRC Dashboard — {org}" if org else "🛡️ GRC Dashboard")
    st.caption("Overzicht stand van zaken · automatisch gegenereerd uit xlsm")
    st.divider()

    df_proc      = data["df_proc"]
    df_ia        = data["df_ia"]
    df_da        = data["df_da"]
    df_vuln      = data["df_vuln"]
    df_rarm      = data["df_rarm"]
    df_risico    = data["df_risico"]
    df_acties    = data["df_acties"]
    da_names     = data.get("da_names", [])
    assurance    = data.get("cfg_assurance", "")

    # Filter RARM en risicobeheer op assurance level
    allowed_levels = ASSURANCE_FILTER.get(assurance, []) if assurance else []
    if allowed_levels and not df_rarm.empty and "Assurance" in df_rarm.columns:
        df_rarm_filtered = df_rarm[df_rarm["Assurance"].isin(allowed_levels)]
    else:
        df_rarm_filtered = df_rarm

    if allowed_levels and not df_risico.empty and "Assurance" in df_risico.columns:
        df_risico_filtered = df_risico[df_risico["Assurance"].isin(allowed_levels)]
    else:
        df_risico_filtered = df_risico

    n_relevant = len(df_rarm_filtered)   # controls relevant voor dit assurance niveau

    # KPI-rij
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    kpi(c1, len(df_proc),  "Processen",           "#2563eb")
    kpi(c2, len(df_ia),    "Informatieassets",    "#0ea5e9")
    kpi(c3, len(df_da),    "Afhankelijke assets", "#8b5cf6")
    kpi(c4, len(df_vuln),  "Kwetsbaarheden",      "#f97316")
    kpi(c5, n_relevant,
        f"Controls ({assurance})" if assurance else "CyFun controls", "#22c55e")

    n_open_act = len(df_acties[df_acties["Status"] == "Open"]) if not df_acties.empty else 0
    kpi(c6, n_open_act, "Open acties", "#ef4444" if n_open_act > 0 else "#22c55e")

    if assurance:
        lvls = " + ".join(allowed_levels)
        st.caption(f"🎯 Assurance niveau: **{assurance}** — controls getoond: {lvls}")
    else:
        st.warning("⚠️ Assurance niveau niet ingesteld — ga naar ⚙️ Configuratie")

    st.markdown("<br>", unsafe_allow_html=True)

    # Voortgangsgrafieken
    col_l, col_m, col_r = st.columns([1, 1, 1.2])

    with col_l:
        st.markdown('<div class="section-hdr">Maatregelen-status</div>', unsafe_allow_html=True)
        if not df_risico_filtered.empty:
            status_counts = df_risico_filtered["Status"].value_counts().reset_index()
            status_counts.columns = ["Status", "Aantal"]
            fig = px.pie(status_counts, names="Status", values="Aantal",
                         color="Status",
                         color_discrete_map={
                             "Uitgevoerd":     "#22c55e",
                             "Gepland":        "#eab308",
                             "Niet uitvoeren": "#94a3b8",
                             "":               "#e2e8f0",
                         },
                         title="Status risicobeheerplan")
            fig.update_layout(height=280, margin=dict(t=40,b=0,l=0,r=0),
                              legend=dict(orientation="h", y=-0.2))
            fig.update_traces(textinfo="label+percent")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Nog geen maatregelen ingevuld")

    with col_m:
        st.markdown('<div class="section-hdr">Controls per DA</div>', unsafe_allow_html=True)
        if da_names and not df_rarm_filtered.empty:
            da_totals = []
            for da in da_names:
                col_key = f"da_{da}"
                if col_key in df_rarm_filtered.columns:
                    da_totals.append({"DA": da, "Controls": df_rarm_filtered[col_key].sum()})
            if da_totals:
                df_da_bar = pd.DataFrame(da_totals)
                level_label = assurance if assurance else "alle"
                fig2 = px.bar(df_da_bar, x="Controls", y="DA", orientation="h",
                              color="Controls",
                              color_continuous_scale=["#dbeafe", "#2563eb"],
                              title=f"Controls per DA ({level_label})")
                fig2.update_layout(height=280, margin=dict(t=40,b=0,l=0,r=0),
                                   showlegend=False, coloraxis_showscale=False)
                st.plotly_chart(fig2, use_container_width=True)

    with col_r:
        st.markdown('<div class="section-hdr">Voortgang per DA</div>', unsafe_allow_html=True)
        if not df_risico_filtered.empty and "Dependent Asset" in df_risico_filtered.columns:
            prog = df_risico_filtered.groupby("Dependent Asset")["Status"].apply(
                lambda s: round(100 * (s == "Uitgevoerd").sum() / len(s)) if len(s) else 0
            ).reset_index()
            prog.columns = ["DA", "% Uitgevoerd"]
            prog = prog.sort_values("% Uitgevoerd")
            fig3 = px.bar(prog, x="% Uitgevoerd", y="DA", orientation="h",
                          color="% Uitgevoerd",
                          color_continuous_scale=["#fee2e2","#eab308","#22c55e"],
                          range_color=[0, 100],
                          title="Uitgevoerde maatregelen per DA")
            fig3.update_layout(height=280, margin=dict(t=40,b=0,l=0,r=0),
                               coloraxis_showscale=False)
            fig3.update_traces(text=prog["% Uitgevoerd"].astype(str) + "%",
                               textposition="outside")
            st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("Nog geen maatregelen ingevuld")

    # Acties-tabel
    st.divider()
    st.markdown('<div class="section-hdr">🔴 Open acties (vervalt binnenkort)</div>',
                unsafe_allow_html=True)
    if not df_acties.empty:
        open_acts = df_acties[df_acties["Status"] == "Open"].copy()
        if not open_acts.empty:
            open_acts["Vervaldatum"] = pd.to_datetime(open_acts["Vervaldatum"], errors="coerce")
            open_acts = open_acts.sort_values("Vervaldatum")
            st.dataframe(
                open_acts[["Actie ID","Type","DA","Omschrijving","Vervaldatum","Verantwoordelijke"]]
                .head(10),
                use_container_width=True, hide_index=True,
            )
        else:
            st.markdown('<div class="ok-box">✅ Geen open acties</div>', unsafe_allow_html=True)
    else:
        st.info("Nog geen acties aangemaakt")

    # Assurance-waarschuwing
    if not assurance:
        st.divider()
        st.markdown(
            '<div class="err-box">⚠️ <strong>Assurance niveau niet ingesteld.</strong> '
            'Ga naar ⚙️ Configuratie om het assurance niveau in te stellen.</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 2 — CONFIGURATIE
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "⚙️ Configuratie":
    st.title("⚙️ Configuratie")
    st.caption("Instellingen van de organisatie en het assurance niveau.")
    st.divider()

    cfg_assurance   = data.get("cfg_assurance",  "")
    cfg_org         = data.get("cfg_org",         "")
    cfg_dept        = data.get("cfg_dept",        "")
    cfg_lang        = data.get("cfg_lang",        "NL")
    cfg_version     = data.get("cfg_version",     "")
    cfg_updated     = data.get("cfg_updated",     "")
    cfg_updated_by  = data.get("cfg_updated_by",  "")

    col_form, col_meta = st.columns([1.5, 1])

    with col_form:
        st.markdown('<div class="section-hdr">Organisatiegegevens</div>', unsafe_allow_html=True)
        new_org  = st.text_input("Naam instelling",         value=cfg_org,  key="cfg_org_input")
        new_dept = st.text_input("Dienst / Entiteit",       value=cfg_dept, key="cfg_dept_input")
        new_lang = st.selectbox("Taal",                     options=["NL","FR","EN"],
                                index=["NL","FR","EN"].index(cfg_lang) if cfg_lang in ["NL","FR","EN"] else 0,
                                key="cfg_lang_input")

        st.markdown('<div class="section-hdr">Assurance niveau entiteit</div>', unsafe_allow_html=True)
        if not cfg_assurance:
            st.markdown(
                '<div class="err-box">⚠️ Verplicht in te stellen — bepaalt welke CyFun-controls '
                'getoond worden in het Maatregelenplan.</div>',
                unsafe_allow_html=True,
            )

        new_assurance = st.selectbox(
            "Assurance niveau",
            options=["", "Basic", "Important", "Essential"],
            index=["", "Basic", "Important", "Essential"].index(cfg_assurance)
                  if cfg_assurance in ["", "Basic", "Important", "Essential"] else 0,
            key="cfg_assurance_input",
        )

        # Uitleg per niveau
        if new_assurance == "Basic":
            st.info("**Basic**: Basismaatregelen — minimumvereisten voor cyberveiligheid.")
        elif new_assurance == "Important":
            st.info("**Important**: Toont **Basic + Important** controls — verhoogde beveiliging.")
        elif new_assurance == "Essential":
            st.info("**Essential**: Toont **Basic + Important + Essential** controls — hoogste beveiligingsniveau.")

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 Opslaan naar xlsm", type="primary"):
            updates = [
                {"sheet": "Config", "row": 5,  "col": 4, "value": new_org},
                {"sheet": "Config", "row": 6,  "col": 4, "value": new_dept},
                {"sheet": "Config", "row": 9,  "col": 4, "value": new_lang},
                {"sheet": "Config", "row": 19, "col": 4, "value": new_assurance},
            ]
            try:
                write_to_xlsm(_active_path, updates)
                st.success("✅ Configuratie opgeslagen in xlsm!")
                st.rerun()
            except Exception as e:
                st.error(f"Fout bij opslaan: {e}")

    with col_meta:
        st.markdown('<div class="section-hdr">Metadata</div>', unsafe_allow_html=True)
        st.metric("Versie",          cfg_version)
        st.metric("Bijgewerkt op",   str(cfg_updated))
        st.metric("Bijgewerkt door", str(cfg_updated_by))
        if cfg_assurance:
            ass_cols = ASSURANCE_FILTER.get(cfg_assurance, [])
            st.markdown(
                f'<div class="ok-box">🎯 Assurance niveau: <strong>{cfg_assurance}</strong><br>'
                f'Controls getoond: {" + ".join(ass_cols)}</div>',
                unsafe_allow_html=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 3 — PROCESSEN & ASSETS
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "📋 Processen & Assets":
    st.title("📋 Processen & Assets")
    df_proc = data["df_proc"]
    df_ia   = data["df_ia"]
    df_da   = data["df_da"]

    tab1, tab2, tab3 = st.tabs(["Processen", "Informatieassets", "Afhankelijke Assets"])

    with tab1:
        st.markdown('<div class="section-hdr">Overzicht processen</div>', unsafe_allow_html=True)
        if not df_proc.empty:
            df_show = df_proc[["ID","Naam","Eigenaar","Integriteit","Beschikbaarheid"]].copy()
            df_show["Integriteit"]     = df_show["Integriteit"].map(level_lbl)
            df_show["Beschikbaarheid"] = df_show["Beschikbaarheid"].map(level_lbl)
            st.dataframe(df_show, use_container_width=True, hide_index=True)
        else:
            st.info("Geen processen gevonden")

    with tab2:
        st.markdown('<div class="section-hdr">Informatieassets</div>', unsafe_allow_html=True)
        if not df_ia.empty:
            df_ia_show = df_ia[["ID","Naam","Eigenaar","Confidentialiteit"]].copy()
            df_ia_show["Confidentialiteit"] = df_ia_show["Confidentialiteit"].map(level_lbl)
            st.dataframe(df_ia_show, use_container_width=True, hide_index=True)
        else:
            st.info("Geen informatieassets gevonden")

    with tab3:
        st.markdown('<div class="section-hdr">Afhankelijke assets</div>', unsafe_allow_html=True)
        if not df_da.empty:
            # Leesbare CIA-labels
            df_da_show = df_da[["ID","Naam","Omschrijving","Eigenaar",
                                "C_req","I_req","A_req",
                                "C_obj","I_obj","A_obj",
                                "Opmerkingen","Commentaar","_row"]].copy()
            for col in ["C_req","I_req","A_req","C_obj","I_obj","A_obj"]:
                df_da_show[col] = df_da_show[col].map(level_lbl)
            df_da_show = df_da_show.rename(columns={
                "C_req": "C vereist", "I_req": "I vereist", "A_req": "A vereist",
                "C_obj": "C objectief", "I_obj": "I objectief", "A_obj": "A objectief",
            })
            edited_da = st.data_editor(
                df_da_show.drop(columns=["_row"]),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ID":           st.column_config.TextColumn("ID",           disabled=True, width="small"),
                    "Naam":         st.column_config.TextColumn("Naam",         disabled=True, width="medium"),
                    "Omschrijving": st.column_config.TextColumn("Omschrijving", disabled=True, width="large"),
                    "Eigenaar":     st.column_config.TextColumn("Eigenaar",     disabled=True, width="small"),
                    "C vereist":    st.column_config.TextColumn("C vereist",    disabled=True, width="small"),
                    "I vereist":    st.column_config.TextColumn("I vereist",    disabled=True, width="small"),
                    "A vereist":    st.column_config.TextColumn("A vereist",    disabled=True, width="small"),
                    "C objectief":  st.column_config.TextColumn("C objectief",  disabled=True, width="small"),
                    "I objectief":  st.column_config.TextColumn("I objectief",  disabled=True, width="small"),
                    "A objectief":  st.column_config.TextColumn("A objectief",  disabled=True, width="small"),
                    "Opmerkingen":  st.column_config.TextColumn("Opmerkingen",  width="large"),
                    "Commentaar":   st.column_config.TextColumn("Commentaar",   width="large"),
                },
                key="da_editor",
            )

            if st.button("💾 Opmerkingen & commentaar opslaan", type="primary"):
                wb_w = openpyxl.load_workbook(_active_path, keep_vba=True)
                ws_da_w = wb_w["Dependent Assets"]
                for i, row_edit in edited_da.iterrows():
                    xl_r = int(df_da_show["_row"].iloc[i])
                    _safe_write(ws_da_w, xl_r,  7, row_edit["Opmerkingen"] or None)
                    _safe_write(ws_da_w, xl_r, 25, row_edit["Commentaar"]  or None)
                wb_w.save(_active_path)
                wb_w.close()
                load_all.clear()
                st.success("✅ Opmerkingen en commentaar opgeslagen!")
                st.rerun()
        else:
            st.info("Geen afhankelijke assets gevonden")


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 4 — MAATREGELEN
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "🎯 Maatregelen":
    st.title("🎯 Maatregelen — Risicobeheerplan")
    st.caption("Per dependent asset: status van elke CyFun-maatregel gefilterd op assurance niveau.")

    df_rarm    = data["df_rarm"]
    df_risico  = data["df_risico"]
    da_names   = data.get("da_names", [])
    assurance  = data.get("cfg_assurance", "")

    # Waarschuwing als assurance niveau niet ingesteld
    if not assurance:
        st.markdown(
            '<div class="err-box">⚠️ Stel eerst het assurance niveau in via ⚙️ Configuratie.</div>',
            unsafe_allow_html=True,
        )
        st.stop()

    allowed_levels = ASSURANCE_FILTER.get(assurance, [])
    st.info(f"🎯 Assurance niveau: **{assurance}** — toont controls: {' + '.join(allowed_levels)}")

    # Filter RARM op assurance niveau
    if not df_rarm.empty and "Assurance" in df_rarm.columns:
        df_filtered = df_rarm[df_rarm["Assurance"].isin(allowed_levels)].copy()
    else:
        df_filtered = df_rarm.copy()

    if df_filtered.empty:
        st.warning("Geen controls gevonden na filteren op assurance niveau.")
        st.stop()

    st.markdown(f"**{len(df_filtered)} controls** relevant voor dit assurance niveau.")
    st.divider()

    # Selecteer DA
    if not da_names:
        st.warning("Geen Dependent Assets gevonden in RARM-sheet.")
        st.stop()

    selected_da = st.selectbox("Selecteer Dependent Asset", options=da_names, key="maatr_da_sel")

    # Haal bestaande risicobeheer-rijen op voor dit DA
    da_col = f"da_{selected_da}"
    if da_col in df_filtered.columns:
        df_da_controls = df_filtered[df_filtered[da_col] == True].copy()
    else:
        df_da_controls = df_filtered.copy()

    if df_da_controls.empty:
        st.info(f"Geen controls gemarkeerd voor **{selected_da}** in de RARM-sheet.")
        st.stop()

    # ── CIA-classificatie blok ────────────────────────────────────────────────
    df_da      = data["df_da"]
    da_name_to_row = data.get("da_name_to_row", {})
    da_match   = df_da[df_da["Naam"] == selected_da]
    da_row_cia = da_match.iloc[0] if not da_match.empty else None

    def _cia_int(field, default=1):
        if da_row_cia is not None:
            v = da_row_cia.get(field)
            if v is not None:
                try:
                    return int(v)
                except Exception:
                    pass
        return default

    c_req = _cia_int("C_req", None)
    i_req = _cia_int("I_req", None)
    a_req = _cia_int("A_req", None)

    # Session-state sleutels voor objectieven (per DA)
    _ck = f"c_obj_{selected_da}"
    _ik = f"i_obj_{selected_da}"
    _ak = f"a_obj_{selected_da}"

    # Initialiseer vanuit xlsm als nog niet in session_state
    if _ck not in st.session_state:
        st.session_state[_ck] = _cia_int("C_obj", 1)
    if _ik not in st.session_state:
        st.session_state[_ik] = _cia_int("I_obj", 1)
    if _ak not in st.session_state:
        st.session_state[_ak] = _cia_int("A_obj", 1)

    st.markdown('<div class="section-hdr">CIA-classificatie</div>', unsafe_allow_html=True)

    cia_cols = st.columns(3)
    cia_labels = {
        None: "—",
        1: "1 — Laag", 2: "2 — Gemiddeld", 3: "3 — Hoog",
        4: "4 — Zeer Hoog", 5: "5 — Kritiek",
    }
    cia_names = ["C (Vertrouwelijkheid)", "I (Integriteit)", "A (Beschikbaarheid)"]
    cia_reqs  = [c_req, i_req, a_req]
    cia_keys  = [_ck,   _ik,   _ak]

    c_obj_val = st.session_state[_ck]
    i_obj_val = st.session_state[_ik]
    a_obj_val = st.session_state[_ak]

    for col_ui, lbl, req, skey in zip(cia_cols, cia_names, cia_reqs, cia_keys):
        with col_ui:
            req_str = cia_labels.get(req, "—")
            st.markdown(f"**{lbl}**")
            st.caption(f"Benodigde classificatie: **{req_str}**")
            st.number_input(
                "Objectief",
                min_value=1, max_value=5,
                key=skey,
                label_visibility="collapsed",
            )

    # Herlaad na widget-render
    c_obj_val = st.session_state[_ck]
    i_obj_val = st.session_state[_ik]
    a_obj_val = st.session_state[_ak]

    # Auto-commentaar als objectief < vereist
    gaps = []
    if c_req is not None and c_obj_val < c_req:
        gaps.append(f"C: objectief {level_lbl(c_obj_val)} < vereist {level_lbl(c_req)}")
    if i_req is not None and i_obj_val < i_req:
        gaps.append(f"I: objectief {level_lbl(i_obj_val)} < vereist {level_lbl(i_req)}")
    if a_req is not None and a_obj_val < a_req:
        gaps.append(f"A: objectief {level_lbl(a_obj_val)} < vereist {level_lbl(a_req)}")

    auto_comment = ("CIA-gat: " + "; ".join(gaps)) if gaps else ""

    if gaps:
        st.warning(f"⚠️ CIA-gat gedetecteerd — {auto_comment}")

    st.divider()

    # ── Merge met bestaande statussen uit Risicobeheer-sheet ──────────────────
    if not df_risico.empty:
        risico_da = df_risico[df_risico["Dependent Asset"] == selected_da][
            ["Control ID","Status","Datum","Verantwoordelijke","Opmerkingen","_row"]
        ].set_index("Control ID")
    else:
        risico_da = pd.DataFrame(columns=["Status","Datum","Verantwoordelijke","Opmerkingen","_row"])
        risico_da.index.name = "Control ID"

    # Bouw bewerkbare tabel
    edit_data = []
    for _, ctrl in df_da_controls.iterrows():
        ctrl_id  = ctrl["Control ID"]
        existing = risico_da.loc[ctrl_id] if ctrl_id in risico_da.index else None
        saved_opmerking = existing["Opmerkingen"] if existing is not None else ""
        # Pre-fill commentaar met CIA-gat als het nog leeg is
        opmerking = saved_opmerking if (saved_opmerking and str(saved_opmerking).strip()) else auto_comment
        edit_data.append({
            "Control ID":        ctrl_id,
            "Richtlijn":         ctrl.get("Richtlijn", ""),
            "Assurance":         ctrl.get("Assurance", ""),
            "Status":            existing["Status"] if existing is not None else "",
            "Datum":             existing["Datum"]  if existing is not None else None,
            "Verantwoordelijke": existing["Verantwoordelijke"] if existing is not None else "",
            "Opmerkingen":       opmerking,
        })

    df_edit = pd.DataFrame(edit_data)

    # Voortgangsbar
    n_uitg = (df_edit["Status"] == "Uitgevoerd").sum()
    n_tot  = len(df_edit)
    pct    = round(100 * n_uitg / n_tot) if n_tot else 0
    st.markdown(f"**Voortgang {selected_da}:** {n_uitg}/{n_tot} maatregelen uitgevoerd ({pct}%)")
    st.progress(pct / 100)

    st.markdown('<div class="section-hdr">Bewerk maatregelen-status</div>', unsafe_allow_html=True)
    edited = st.data_editor(
        df_edit,
        column_config={
            "Control ID":  st.column_config.TextColumn("Control ID",  disabled=True, width="small"),
            "Richtlijn":   st.column_config.TextColumn("Richtlijn",   disabled=True, width="large"),
            "Assurance":   st.column_config.TextColumn("Assurance",   disabled=True, width="small"),
            "Status": st.column_config.SelectboxColumn(
                "Status", options=["", "Uitgevoerd", "Gepland", "Niet uitvoeren"], width="medium"
            ),
            "Datum":       st.column_config.DateColumn("Datum",       width="small"),
            "Verantwoordelijke": st.column_config.TextColumn("Verantwoordelijke", width="medium"),
            "Opmerkingen": st.column_config.TextColumn("Opmerkingen", width="large"),
        },
        hide_index=True,
        use_container_width=True,
        key=f"maatr_editor_{selected_da}",
    )

    if st.button("💾 Opslaan voor " + selected_da, type="primary"):
        wb_write = openpyxl.load_workbook(_active_path, keep_vba=True)
        ws_risico_w = wb_write["Risicobeheer"]

        # Bepaal hoogste gebruikte rij
        max_row = ws_risico_w.max_row
        if max_row < 3:
            max_row = 2

        # Verwijder placeholder
        if max_row == 3:
            first_val = ws_risico_w.cell(row=3, column=1).value
            if first_val and str(first_val).startswith("(leeg"):
                max_row = 2

        # Bouw index van bestaande rijen (Control ID + DA → rijnummer)
        existing_idx = {}
        for r in range(3, ws_risico_w.max_row + 1):
            cid = ws_risico_w.cell(row=r, column=1).value
            da  = ws_risico_w.cell(row=r, column=4).value
            if cid and da:
                existing_idx[(str(cid), str(da))] = r

        for _, row in edited.iterrows():
            key = (str(row["Control ID"]), selected_da)
            datum_val = row["Datum"]
            if isinstance(datum_val, str) and datum_val:
                try:
                    datum_val = datetime.strptime(datum_val, "%Y-%m-%d").date()
                except Exception:
                    pass

            if key in existing_idx:
                r = existing_idx[key]
            else:
                max_row += 1
                r = max_row
                ws_risico_w.cell(row=r, column=1).value = row["Control ID"]
                ws_risico_w.cell(row=r, column=2).value = row["Richtlijn"]
                ws_risico_w.cell(row=r, column=3).value = row["Assurance"]
                ws_risico_w.cell(row=r, column=4).value = selected_da

            ws_risico_w.cell(row=r, column=5).value = row["Status"] or None
            ws_risico_w.cell(row=r, column=6).value = datum_val if datum_val else None
            ws_risico_w.cell(row=r, column=7).value = row["Verantwoordelijke"] or None
            ws_risico_w.cell(row=r, column=8).value = row["Opmerkingen"] or None

        # ── Schrijf CIA-objectieven terug naar Dependent Assets-sheet ─────────
        da_xl_row = da_name_to_row.get(selected_da)
        if da_xl_row:
            ws_da_w = wb_write["Dependent Assets"]
            _safe_write(ws_da_w, da_xl_row, 16, c_obj_val)  # C objectief
            _safe_write(ws_da_w, da_xl_row, 18, i_obj_val)  # I objectief
            _safe_write(ws_da_w, da_xl_row, 20, a_obj_val)  # A objectief

        wb_write.save(_active_path)
        wb_write.close()
        load_all.clear()
        # Reset session-state objectieven zodat ze opnieuw vanuit xlsm geladen worden
        for _k in [_ck, _ik, _ak]:
            st.session_state.pop(_k, None)
        st.success(f"✅ {len(edited)} maatregelen opgeslagen voor {selected_da}!")
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 5 — PERIODIEKE CONTROLES
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "🔍 Periodieke Controles":
    st.title("🔍 Periodieke Controles")
    st.caption("Beheer het controle-schema per afhankelijk asset.")

    df_controles = data["df_controles"]
    da_names     = data.get("da_names", [])

    # Statistieken
    if not df_controles.empty:
        n_ok    = (df_controles["Status"] == "OK").sum()
        n_te    = (df_controles["Status"] == "Te controleren").sum()
        n_verv  = (df_controles["Status"] == "Vervallen").sum()
        c1, c2, c3 = st.columns(3)
        kpi(c1, n_ok,   "OK",             "#22c55e")
        kpi(c2, n_te,   "Te controleren", "#eab308")
        kpi(c3, n_verv, "Vervallen",      "#ef4444")
        st.markdown("<br>", unsafe_allow_html=True)

    # Tabbladen per DA
    da_with_ctrl = sorted(df_controles["DA"].unique().tolist()) if not df_controles.empty else []
    alle_da = sorted(set(da_names + da_with_ctrl)) if da_names else da_with_ctrl

    if alle_da:
        tabs = st.tabs(alle_da + ["➕ Nieuw"])
        for i, da in enumerate(alle_da):
            with tabs[i]:
                da_df = df_controles[df_controles["DA"] == da].copy() if not df_controles.empty else pd.DataFrame()

                if da_df.empty:
                    st.info(f"Geen controleregels voor **{da}**.")
                else:
                    # Per rij een "Controle uitvoeren"-knop
                    for _, row in da_df.iterrows():
                        with st.expander(
                            f"📌 {row['Control ID']} — {row['Maatregel'][:60]}  "
                            f"| {row['Frequentie']} | {row['Status']}",
                            expanded=False
                        ):
                            col_a, col_b = st.columns([2, 1])
                            with col_a:
                                st.write(f"**Methode:** {row['Methode']}")
                                st.write(f"**Artefacten:** {row['Artefacten locatie']}")
                                lc = row["Laatste controle"]
                                vc = row["Volgende controle"]
                                st.write(f"**Laatste controle:** {lc if lc else '—'}")
                                st.write(f"**Volgende controle:** {vc if vc else '—'}")
                            with col_b:
                                if st.button(
                                    f"✅ Controle uitgevoerd",
                                    key=f"ctrl_btn_{row['_row']}",
                                ):
                                    vandaag = date.today()
                                    volgende = calc_volgende(vandaag, row["Frequentie"])

                                    wb_w = openpyxl.load_workbook(_active_path, keep_vba=True)
                                    ws_ctrl = wb_w["Controles"]
                                    ws_ctrl.cell(row=row["_row"], column=7).value  = vandaag
                                    ws_ctrl.cell(row=row["_row"], column=8).value  = volgende
                                    ws_ctrl.cell(row=row["_row"], column=9).value  = "OK"

                                    # Automatisch actie aanmaken
                                    ws_act = wb_w["Acties"]
                                    max_r  = ws_act.max_row
                                    if max_r < 3:
                                        max_r = 2
                                    # verwijder placeholder
                                    if max_r == 3:
                                        first = ws_act.cell(row=3, column=1).value
                                        if first and str(first).startswith("(leeg"):
                                            max_r = 2

                                    # Genereer actie ID
                                    act_df_fresh = data["df_acties"]
                                    new_id = next_actie_id(act_df_fresh)
                                    if volgende:
                                        max_r += 1
                                        ws_act.cell(row=max_r, column=1).value = new_id
                                        ws_act.cell(row=max_r, column=2).value = "Periodieke controle"
                                        ws_act.cell(row=max_r, column=3).value = da
                                        ws_act.cell(row=max_r, column=4).value = row["Control ID"]
                                        ws_act.cell(row=max_r, column=5).value = (
                                            f"Periodieke controle: {row['Maatregel'][:80]}")
                                        ws_act.cell(row=max_r, column=6).value = volgende
                                        ws_act.cell(row=max_r, column=7).value = ""
                                        ws_act.cell(row=max_r, column=8).value = "Open"
                                        ws_act.cell(row=max_r, column=9).value = vandaag

                                    wb_w.save(_active_path)
                                    wb_w.close()
                                    load_all.clear()
                                    st.success(f"✅ Controle geregistreerd. Volgende: {volgende}")
                                    if volgende:
                                        st.info(f"📋 Actie {new_id} aangemaakt voor {volgende}")
                                    st.rerun()

                st.divider()

                # Nieuwe controleregel toevoegen
                with st.form(f"nieuw_ctrl_{da}"):
                    st.markdown(f"**Nieuwe controleregel voor {da}**")
                    nc1, nc2 = st.columns(2)
                    with nc1:
                        new_ctrl_id  = st.text_input("Control ID", key=f"nci_{da}")
                        new_maatr    = st.text_input("Maatregel",  key=f"nm_{da}")
                        new_freq     = st.selectbox("Frequentie",
                            ["Maandelijks","Kwartaal","Halfjaarlijks","Jaarlijks","Ad hoc"],
                            key=f"nf_{da}")
                    with nc2:
                        new_methode  = st.text_input("Methode",    key=f"nme_{da}")
                        new_artef    = st.text_input("Artefacten locatie", key=f"na_{da}")
                    if st.form_submit_button("➕ Toevoegen"):
                        wb_w = openpyxl.load_workbook(_active_path, keep_vba=True)
                        ws_ctrl = wb_w["Controles"]
                        max_r   = ws_ctrl.max_row
                        if max_r < 3: max_r = 2
                        if max_r == 3:
                            first = ws_ctrl.cell(row=3, column=1).value
                            if first and str(first).startswith("(leeg"): max_r = 2
                        max_r += 1
                        ws_ctrl.cell(row=max_r, column=1).value = da
                        ws_ctrl.cell(row=max_r, column=2).value = new_ctrl_id
                        ws_ctrl.cell(row=max_r, column=3).value = new_maatr
                        ws_ctrl.cell(row=max_r, column=4).value = new_freq
                        ws_ctrl.cell(row=max_r, column=5).value = new_methode
                        ws_ctrl.cell(row=max_r, column=6).value = new_artef
                        ws_ctrl.cell(row=max_r, column=9).value = "Te controleren"
                        wb_w.save(_active_path)
                        wb_w.close()
                        load_all.clear()
                        st.success("Controleregel toegevoegd!")
                        st.rerun()

        # Tab: Nieuw (voor DA niet in lijst)
        with tabs[-1]:
            st.markdown("**Nieuwe controleregel — kies DA vrij**")
            with st.form("nieuw_ctrl_vrij"):
                c1, c2 = st.columns(2)
                with c1:
                    new_da_vrij    = st.text_input("DA", key="ndav")
                    new_ctrl_vrij  = st.text_input("Control ID", key="nciv")
                    new_maatr_vrij = st.text_input("Maatregel",  key="nmv")
                    new_freq_vrij  = st.selectbox("Frequentie",
                        ["Maandelijks","Kwartaal","Halfjaarlijks","Jaarlijks","Ad hoc"],
                        key="nfv")
                with c2:
                    new_meth_vrij  = st.text_input("Methode",            key="nmev")
                    new_art_vrij   = st.text_input("Artefacten locatie", key="nav")
                if st.form_submit_button("➕ Toevoegen"):
                    wb_w = openpyxl.load_workbook(_active_path, keep_vba=True)
                    ws_ctrl = wb_w["Controles"]
                    max_r = ws_ctrl.max_row
                    if max_r < 3: max_r = 2
                    if max_r == 3:
                        first = ws_ctrl.cell(row=3, column=1).value
                        if first and str(first).startswith("(leeg"): max_r = 2
                    max_r += 1
                    ws_ctrl.cell(row=max_r, column=1).value = new_da_vrij
                    ws_ctrl.cell(row=max_r, column=2).value = new_ctrl_vrij
                    ws_ctrl.cell(row=max_r, column=3).value = new_maatr_vrij
                    ws_ctrl.cell(row=max_r, column=4).value = new_freq_vrij
                    ws_ctrl.cell(row=max_r, column=5).value = new_meth_vrij
                    ws_ctrl.cell(row=max_r, column=6).value = new_art_vrij
                    ws_ctrl.cell(row=max_r, column=9).value = "Te controleren"
                    wb_w.save(_active_path)
                    wb_w.close()
                    load_all.clear()
                    st.success("Controleregel toegevoegd!")
                    st.rerun()
    else:
        st.info("Nog geen controleregels. Voeg hieronder een eerste toe.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 6 — ACTIES
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "📋 Acties":
    st.title("📋 Actielijst")
    st.caption("Opvolging van open acties voor maatregelen en periodieke controles.")

    df_acties = data["df_acties"]
    da_names  = data.get("da_names", [])

    # KPI
    n_open = (df_acties["Status"] == "Open").sum()           if not df_acties.empty else 0
    n_prog = (df_acties["Status"] == "In uitvoering").sum()  if not df_acties.empty else 0
    n_ges  = (df_acties["Status"] == "Gesloten").sum()       if not df_acties.empty else 0
    today  = date.today()

    # Vervallen = Open + vervaldatum < vandaag
    if not df_acties.empty:
        df_acties["Vervaldatum_dt"] = pd.to_datetime(df_acties["Vervaldatum"], errors="coerce")
        n_verv = ((df_acties["Status"] == "Open") &
                  (df_acties["Vervaldatum_dt"].dt.date < today)).sum()
    else:
        n_verv = 0

    c1, c2, c3, c4 = st.columns(4)
    kpi(c1, n_open, "Open",          "#eab308")
    kpi(c2, n_prog, "In uitvoering", "#2563eb")
    kpi(c3, n_ges,  "Gesloten",      "#22c55e")
    kpi(c4, n_verv, "Vervallen",     "#ef4444" if n_verv > 0 else "#22c55e")

    st.markdown("<br>", unsafe_allow_html=True)

    # Filters
    with st.expander("🔍 Filters", expanded=True):
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            f_status = st.multiselect("Status", ["Open","In uitvoering","Gesloten"],
                                       default=["Open","In uitvoering"], key="act_f_status")
        with fc2:
            alle_da_act = ["(alles)"] + sorted(df_acties["DA"].dropna().unique().tolist()) \
                          if not df_acties.empty else ["(alles)"]
            f_da = st.selectbox("DA", alle_da_act, key="act_f_da")
        with fc3:
            alle_types = ["(alles)"] + sorted(df_acties["Type"].dropna().unique().tolist()) \
                         if not df_acties.empty else ["(alles)"]
            f_type = st.selectbox("Type", alle_types, key="act_f_type")

    df_filtered = df_acties.copy() if not df_acties.empty else pd.DataFrame()
    if not df_filtered.empty:
        if f_status:
            df_filtered = df_filtered[df_filtered["Status"].isin(f_status)]
        if f_da != "(alles)":
            df_filtered = df_filtered[df_filtered["DA"] == f_da]
        if f_type != "(alles)":
            df_filtered = df_filtered[df_filtered["Type"] == f_type]

    # Bewerkbare tabel
    if not df_filtered.empty:
        st.markdown('<div class="section-hdr">Actielijst</div>', unsafe_allow_html=True)
        show_cols = ["Actie ID","Type","DA","Control ID","Omschrijving",
                     "Vervaldatum","Verantwoordelijke","Status"]
        edited_acts = st.data_editor(
            df_filtered[show_cols],
            column_config={
                "Actie ID":     st.column_config.TextColumn("Actie ID",     disabled=True, width="small"),
                "Type":         st.column_config.TextColumn("Type",         disabled=True, width="medium"),
                "DA":           st.column_config.TextColumn("DA",           disabled=True, width="medium"),
                "Control ID":   st.column_config.TextColumn("Control ID",   disabled=True, width="small"),
                "Omschrijving": st.column_config.TextColumn("Omschrijving", disabled=True, width="large"),
                "Vervaldatum":  st.column_config.DateColumn("Vervaldatum",  width="small"),
                "Verantwoordelijke": st.column_config.TextColumn("Verantwoordelijke", width="medium"),
                "Status": st.column_config.SelectboxColumn(
                    "Status", options=["Open","In uitvoering","Gesloten"], width="medium"
                ),
            },
            hide_index=True,
            use_container_width=True,
            key="acties_editor",
        )

        if st.button("💾 Statuswijzigingen opslaan", type="primary"):
            # Match terug op Actie ID + _row
            id_to_row = df_acties.set_index("Actie ID")["_row"].to_dict()
            wb_w = openpyxl.load_workbook(_active_path, keep_vba=True)
            ws_act = wb_w["Acties"]
            saved = 0
            for _, row in edited_acts.iterrows():
                act_id = row["Actie ID"]
                if act_id in id_to_row:
                    r = id_to_row[act_id]
                    ws_act.cell(row=r, column=8).value = row["Status"]
                    ws_act.cell(row=r, column=6).value = row["Vervaldatum"]
                    ws_act.cell(row=r, column=7).value = row["Verantwoordelijke"]
                    saved += 1
            wb_w.save(_active_path)
            wb_w.close()
            load_all.clear()
            st.success(f"✅ {saved} acties bijgewerkt!")
            st.rerun()
    else:
        st.info("Geen acties gevonden met huidige filters.")

    # Nieuwe actie aanmaken
    st.divider()
    st.markdown('<div class="section-hdr">➕ Nieuwe actie aanmaken</div>', unsafe_allow_html=True)
    with st.form("nieuwe_actie_form"):
        na1, na2 = st.columns(2)
        with na1:
            na_da     = st.selectbox("DA", options=([""] + da_names), key="na_da")
            na_ctrl   = st.text_input("Control ID",    key="na_ctrl")
            na_type   = st.selectbox("Type",
                options=["Maatregel","Periodieke controle","Opvolging"], key="na_type")
        with na2:
            na_omschr = st.text_input("Omschrijving",  key="na_omschr")
            na_datum  = st.date_input("Vervaldatum",   value=date.today() + timedelta(days=30),
                                       key="na_datum")
            na_verant = st.text_input("Verantwoordelijke", key="na_verant")
        if st.form_submit_button("➕ Actie aanmaken"):
            new_id = next_actie_id(df_acties)
            wb_w   = openpyxl.load_workbook(_active_path, keep_vba=True)
            ws_act = wb_w["Acties"]
            max_r  = ws_act.max_row
            if max_r < 3: max_r = 2
            if max_r == 3:
                first = ws_act.cell(row=3, column=1).value
                if first and str(first).startswith("(leeg"): max_r = 2
            max_r += 1
            ws_act.cell(row=max_r, column=1).value = new_id
            ws_act.cell(row=max_r, column=2).value = na_type
            ws_act.cell(row=max_r, column=3).value = na_da
            ws_act.cell(row=max_r, column=4).value = na_ctrl
            ws_act.cell(row=max_r, column=5).value = na_omschr
            ws_act.cell(row=max_r, column=6).value = na_datum
            ws_act.cell(row=max_r, column=7).value = na_verant
            ws_act.cell(row=max_r, column=8).value = "Open"
            ws_act.cell(row=max_r, column=9).value = date.today()
            wb_w.save(_active_path)
            wb_w.close()
            load_all.clear()
            st.success(f"✅ Actie {new_id} aangemaakt!")
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 7 — KWETSBAARHEDEN
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "🛡️ Kwetsbaarheden":
    st.title("🛡️ Kwetsbaarheden")
    df_vuln = data["df_vuln"]

    if df_vuln.empty:
        st.info("Geen kwetsbaarheden gevonden.")
    else:
        st.markdown(f"**{len(df_vuln)} kwetsbaarheden** gekoppeld aan CyFun controls.")
        top10 = df_vuln.sort_values("Aantal controls", ascending=True).tail(10)
        fig = px.bar(top10, x="Aantal controls", y="Kwetsbaarheid", orientation="h",
                     color="Aantal controls",
                     color_continuous_scale=["#dbeafe","#2563eb"],
                     title="Controls per kwetsbaarheid (top 10)")
        fig.update_layout(height=340, margin=dict(t=40,b=0,l=0,r=0),
                          coloraxis_showscale=False)
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(df_vuln, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGINA 8 — INSTELLINGEN
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "🔧 Instellingen":
    st.title("🔧 Instellingen")
    st.caption("Selecteer het .xlsm-bestand dat als gegevensbron dient.")
    st.divider()

    cfg = _load_config()
    _cur = Path(st.session_state.get("xlsm_path", str(_DEFAULT_XLSM)))

    st.markdown('<div class="section-hdr">Huidig actief bestand</div>', unsafe_allow_html=True)
    if _cur.exists():
        _stat = _cur.stat()
        _size_kb  = round(_stat.st_size / 1024, 1)
        _modified = datetime.fromtimestamp(_stat.st_mtime).strftime("%d/%m/%Y %H:%M")
        st.markdown(
            f'<div class="ok-box">📄 <strong>{_cur.name}</strong><br>'
            f'📁 {_cur.parent}<br>'
            f'💾 {_size_kb} KB &nbsp;·&nbsp; 🕒 Gewijzigd: {_modified}</div>',
            unsafe_allow_html=True,
        )
        try:
            _wb_check = openpyxl.load_workbook(str(_cur), read_only=True, keep_vba=True)
            _sheets   = _wb_check.sheetnames
            _wb_check.close()
            st.caption(f"Sheets: {', '.join(_sheets)}")
        except Exception as _e:
            st.caption(f"(sheets niet leesbaar: {_e})")
    else:
        st.markdown(
            f'<div class="err-box">❌ Bestand niet gevonden: {_cur}</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-hdr">Ander bestand laden</div>', unsafe_allow_html=True)
    _new_path = st.text_input(
        "Volledig pad naar .xlsm of .xlsx bestand",
        value=str(_cur),
        key="instellingen_path_input",
    )

    _path_obj = Path(_new_path.strip()) if _new_path.strip() else None
    _path_ok  = False
    if _path_obj:
        if not _path_obj.exists():
            st.markdown('<div class="err-box">⚠️ Bestand bestaat niet.</div>',
                        unsafe_allow_html=True)
        elif _path_obj.suffix.lower() not in (".xlsm", ".xlsx"):
            st.markdown('<div class="err-box">⚠️ Moet .xlsm of .xlsx zijn.</div>',
                        unsafe_allow_html=True)
        else:
            _stat2 = _path_obj.stat()
            st.markdown(
                f'<div class="ok-box">✅ Gevonden: <strong>{_path_obj.name}</strong> — '
                f'{round(_stat2.st_size/1024,1)} KB</div>',
                unsafe_allow_html=True,
            )
            _path_ok = True

    _cb1, _cb2 = st.columns([1, 4])
    with _cb1:
        if st.button("✅ Laad bestand", disabled=not _path_ok, type="primary"):
            _new_str = str(_path_obj)
            st.session_state["xlsm_path"] = _new_str
            cfg["xlsm_path"] = _new_str
            cfg = _add_recent(cfg, _new_str)
            _save_config(cfg)
            load_all.clear()
            st.success(f"Gewijzigd naar: {_path_obj.name}")
            st.rerun()
    with _cb2:
        if st.button("↩️ Reset naar standaard"):
            _def_str = str(_DEFAULT_XLSM)
            st.session_state["xlsm_path"] = _def_str
            cfg["xlsm_path"] = _def_str
            _save_config(cfg)
            load_all.clear()
            st.success("Teruggezet.")
            st.rerun()

    _recent = cfg.get("recent", [])
    if _recent:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-hdr">Recente bestanden</div>', unsafe_allow_html=True)
        for _rp in _recent:
            _rp_obj = Path(_rp)
            _icon   = "✅" if _rp_obj.exists() else "❌"
            _actief = " **← actief**" if _rp == st.session_state.get("xlsm_path") else ""
            rc1, rc2 = st.columns([5, 1])
            with rc1:
                st.caption(f"{_icon} {_rp_obj.name} — `{_rp_obj.parent}`{_actief}")
            with rc2:
                if _rp_obj.exists() and _rp != st.session_state.get("xlsm_path"):
                    if st.button("Laden", key=f"rec_{_rp}"):
                        st.session_state["xlsm_path"] = _rp
                        cfg["xlsm_path"] = _rp
                        cfg = _add_recent(cfg, _rp)
                        _save_config(cfg)
                        load_all.clear()
                        st.rerun()
