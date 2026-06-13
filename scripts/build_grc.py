"""
GRC Tool — Excel Workbook Builder v0.3
Voor overheidsorganisaties

Dit script bouwt volledig automatisch een macro-enabled Excel-bestand (.xlsm) op
voor het beheer van informatieveiligheid (GRC) binnen Belgische overheidsdiensten.

Werkwijze in twee stappen:
  1. openpyxl  → alle sheets, stijlen, formules, validaties en vertalingen worden
                 statisch gebouwd en opgeslagen als tussenbestand (.xlsx).
  2. win32com  → het .xlsx-bestand wordt heropend in Excel via COM-automatisering
                 en VBA-code (modules, UserForms, Sheet-events) wordt geïnjecteerd.
                 Resultaat wordt opgeslagen als .xlsm en het .xlsx-tussenbestand
                 wordt verwijderd.

Sheets (volgorde): Config | Info | Processen | Informatieassets | Verantwoordelijken
                   | Import & Export | Referentiewaarden | _Lang (verborgen)
Output: ../data/template/GRC_Tool.xlsm
"""

import datetime
import io
import os
import re
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# ── Paden & metadata ──────────────────────────────────────────────────────────
# OUT_XLSX is een tijdelijk .xlsx-bestand; na VBA-injectie wordt het verwijderd.
# CYFUN_SRC   : CyFun 2025 ESSENTIAL bronbestand (controls en assurance-niveaus)
# MAPPING_SRC : mapping CyFun 2023 ↔ 2025 (beide richtingen)
# CYFUN23_SRC : CyFun 2023 bronbestand (voor "Enkel 2023"-details)
OUT      = Path(__file__).parent.parent / "data" / "template" / "GRC_Tool.xlsm"
OUT_XLSX = OUT.with_suffix(".xlsx")   # tussenstap voor win32com
CYFUN_SRC   = Path(__file__).parent.parent / "data" / "repositories" / "CyFun2025_Self-Assessment_tool_ESSENTIAL_v3.1.xlsx"
MAPPING_SRC = Path(__file__).parent.parent / "data" / "repositories" / "Mapping_CyFun2023-CyFun2025_v2026-02-25.xlsx"
CYFUN23_SRC = Path(__file__).parent.parent / "data" / "repositories" / "CyFun Self-Assessment tool V2025-08-04.xlsx"
USERNAME = os.environ.get("USERNAME", os.environ.get("USER", "Onbekend"))
NOW_STR  = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

# ── Kleurenpalet ──────────────────────────────────────────────────────────────
# Alle kleursleutels zijn hex-strings zonder '#'. Gebruik via fill(key) of C[key].
# navy/blue_*  : primaire UI-kleuren (titelbalk, koppen, actieve velden)
# grey_*       : neutrale achtergronden en randen
# green/yellow/orange/red/purple : classificatieniveaus 1 t/m 5
# accent/*     : tips-blokken, hint-teksten
C = {
    "navy":          "0F2B46",
    "blue_mid":      "2563EB",
    "blue_light":    "DBEAFE",
    "blue_xlight":   "EFF6FF",
    "white":         "FFFFFF",
    "grey_border":   "CBD5E1",
    "grey_light":    "F8FAFC",
    "grey_mid":      "E2E8F0",
    "grey_dark":     "475569",
    "text":          "0F172A",
    "subtext":       "64748B",
    # Classificatieniveaus (1=Laag/groen … 4=Zeer Hoog/rood, 5=Classified/paars)
    "green":         "15803D",  "green_light":   "DCFCE7",
    "yellow":        "854D0E",  "yellow_light":  "FEF9C3",
    "orange":        "B45309",  "orange_light":  "FEF3C7",
    "red":           "B91C1C",  "red_light":     "FEE2E2",
    "kritiek":       "991B1B",  "kritiek_bg":    "991B1B",
    # UI-accenten
    "accent":        "0EA5E9",  "accent_light":  "E0F2FE",
    "purple":        "6B21A8",  "purple_light":  "F3E8FF",
}

# Vijf classificatieniveaus: index 0 = niveau 1 (Laag) … index 4 = niveau 5 (Classified)
# Worden gebruikt door add_cls_pairs() en ref_dim_table() om CF-regels en legenda's op te bouwen.
LEVEL_FILLS = ["green_light", "yellow_light", "orange_light", "red_light", "purple_light"]
LEVEL_FONTS = ["green",       "yellow",       "orange",       "red",       "purple"]

# ── Stijlhulpfuncties ─────────────────────────────────────────────────────────
def fill(key):
    """Geeft een effen PatternFill terug voor kleursleutel key uit het palet C."""
    return PatternFill("solid", fgColor=C[key])

def font(size=10, bold=False, color="text", italic=False):
    """
    Maakt een Calibri Font-object aan.

    Parameters
    ----------
    size   : int   — tekengrootte in punten (standaard 10)
    bold   : bool  — vetgedrukt (standaard False)
    color  : str   — sleutel in kleurenpalet C (standaard 'text' = bijna-zwart)
    italic : bool  — cursief (standaard False)
    """
    return Font(name="Calibri", size=size, bold=bold, color=C[color], italic=italic)

def align(h="left", v="center", wrap=False, indent=0):
    """
    Maakt een Alignment-object aan.

    Parameters
    ----------
    h      : str  — horizontaal uitlijnen ('left', 'center', 'right')
    v      : str  — verticaal uitlijnen ('top', 'center', 'bottom')
    wrap   : bool — regelterugloop (wrap_text)
    indent : int  — inspringing in tekeneenheden
    """
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def border_all(color="grey_border"):
    """
    Geeft een Border-object terug met dunne rand rondom (alle 4 zijden gelijk).

    Parameters
    ----------
    color : str — sleutel in kleurenpalet C (standaard 'grey_border')
    """
    s = Side(style="thin", color=C[color])
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    """
    Stelt kolombreedtes in voor worksheet ws op basis van een lijst met breedtes.
    Kolom 1 (A) krijgt widths[0], kolom 2 (B) krijgt widths[1], enz.

    Parameters
    ----------
    ws     : openpyxl.Worksheet
    widths : list[float] — kolombreedtes in tekeneenheden
    """
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def no_gridlines(ws):
    """Verbergt het rasterpatroon op een worksheet voor een cleaner uiterlijk."""
    ws.sheet_view.showGridLines = False

# ── Taalformule ───────────────────────────────────────────────────────────────
# LANG_CELL verwijst naar de dropdowncel op het Config-blad waar de gebruiker
# NL, FR of EN kiest. Alle zichtbare teksten in het werkboek zijn formules
# die via t(key) dynamisch de juiste kolom opzoeken in het verborgen _Lang-blad.
LANG_CELL = "Config!$D$9"

def t(key):
    """
    Genereert een Excel-formule die de vertaling van 'key' opzoekt in het
    verborgen _Lang-blad en weergeeft in de taal die is ingesteld op Config!$D$9.

    Werking
    -------
    - MATCH zoekt de sleutelwaarde (key) op in kolom A van _Lang (rijen 2–300).
    - INDEX haalt de waarde op uit kolom B (NL), C (FR) of D (EN) naargelang
      de taalinstelling in Config!$D$9.
    - IFERROR toont '[key]' als de sleutel niet bestaat in _Lang — handig
      om ontbrekende vertalingen snel te detecteren tijdens ontwikkeling.

    Parameters
    ----------
    key : str — unieke sleutelwaarde zoals gedefinieerd in TRANSLATIONS

    Geeft terug
    -----------
    str — een Excel-formulestring die in een cel geplaatst kan worden
    """
    return (
        f'=IFERROR(INDEX(_Lang!$B$2:$D$300,'
        f'MATCH("{key}",_Lang!$A$2:$A$300,0),'
        f'IF({LANG_CELL}="NL",1,IF({LANG_CELL}="FR",2,3))),"[{key}]")'
    )

# ── Classificatiewaarden (5 niveaus) ─────────────────────────────────────────
# CLS, ROLES en ASSET_TYPES worden opgeslagen in het verborgen _Lang-blad
# (kolommen F/G/H voor NL/FR/EN) en via benoemde bereiken (cls_NL, roles_FR, …)
# beschikbaar gesteld aan dropdownvalidaties in de gegevensbladen.
# De dropdown-formule gebruikt INDIRECT("cls_"&Config!$D$9) zodat de lijst
# automatisch wisselt wanneer de gebruiker de taal verandert.
CLS = {
    "NL": ["Laag", "Gemiddeld", "Hoog", "Zeer Hoog", "Classified"],
    "FR": ["Faible", "Moyen", "Élevé", "Très élevé", "Classified"],
    "EN": ["Low",  "Medium",  "High",  "Very High",  "Classified"],
}
ROLES = {
    "NL": ["CISO", "DPO / FG", "Proceseigenaar", "Informatiebeheerder",
           "Systeembeheerder", "Intern auditor", "Directie / Management", "Andere"],
    "FR": ["RSSI", "DPD / FG", "Propriétaire du processus", "Gestionnaire de l'information",
           "Administrateur système", "Auditeur interne", "Direction / Management", "Autre"],
    "EN": ["CISO", "DPO", "Process Owner", "Information Manager",
           "System Administrator", "Internal Auditor", "Management", "Other"],
}
ASSET_TYPES = {
    "NL": ["Applicatie", "Databank", "Server", "Hardware", "Netwerkapparatuur",
           "Dienst", "Document", "Persoon", "Proces", "Andere"],
    "FR": ["Application", "Base de données", "Serveur", "Matériel", "Équipement réseau",
           "Service", "Document", "Personne", "Processus", "Autre"],
    "EN": ["Application", "Database", "Server", "Hardware", "Network Equipment",
           "Service", "Document", "Person", "Process", "Other"],
}
# ── Vertalingstabel ───────────────────────────────────────────────────────────
# TRANSLATIONS is een lijst van tuples: (sleutel, NL-tekst, FR-tekst, EN-tekst).
# Bij het opbouwen van het werkboek worden alle tuples weggeschreven naar het
# verborgen _Lang-blad (kolom A = sleutel, B = NL, C = FR, D = EN).
# Elke cel die gebruikerstekst bevat roept t(sleutel) aan; de formule zoekt
# op basis van de taalinstelling de juiste kolom op.
TRANSLATIONS = [
    # Paginatitel Config / Chrome
    ("cfg_title_main",
        "GRC Tool — Informatieveiligheid voor Overheidsdiensten",
        "Outil GRC — Sécurité de l'information pour les services publics",
        "GRC Tool — Information Security for Government Services"),
    ("app_tagline", "Governance · Risico · Compliance",
        "Gouvernance · Risque · Conformité", "Governance · Risk · Compliance"),
    ("ui_tips",         "TIPS",              "CONSEILS",         "TIPS"),
    ("ui_quicklinks",   "SNELKOPPELINGEN",   "RACCOURCIS",       "QUICK LINKS"),
    ("ui_goto_config",  "→  Configuratie",   "→  Configuration", "→  Configuration"),
    ("ui_goto_ref",     "→  Referentiewaarden", "→  Valeurs de référence", "→  Reference Values"),
    ("ui_goto_proc",    "→  Processen",      "→  Processus",     "→  Processes"),
    ("ui_goto_assets",  "→  Informatieassets","→  Actifs informationnels","→  Information Assets"),
    ("ui_goto_verant",  "→  Verantwoordelijken","→  Responsables","→  Responsible Persons"),
    ("ui_goto_import",  "→  Import & Export","→  Import & Export","→  Import & Export"),
    ("ui_goto_info",    "→  Informatie",     "→  Information",   "→  Information"),
    # Config
    ("cfg_section_general", "ALGEMENE INSTELLINGEN",     "PARAMÈTRES GÉNÉRAUX",      "GENERAL SETTINGS"),
    ("cfg_section_display", "WEERGAVE-INSTELLINGEN",     "PARAMÈTRES D'AFFICHAGE",   "DISPLAY SETTINGS"),
    ("cfg_section_meta",    "OVER DIT BESTAND",          "À PROPOS DE CE FICHIER",   "ABOUT THIS FILE"),
    ("cfg_org_label",       "Naam overheidsinstelling",  "Nom de l'institution publique", "Name of public institution"),
    ("cfg_dept_label",      "Dienst / Entiteit",         "Service / Entité",          "Department / Entity"),
    ("cfg_lang_label",      "Taal / Langue / Language",  "Taal / Langue / Language",  "Taal / Langue / Language"),
    ("cfg_lang_hint",
        "Kies NL, FR of EN. Dropdowns en vertalingen passen zich automatisch aan.",
        "Choisissez NL, FR ou EN. Les listes et traductions s'adaptent automatiquement.",
        "Choose NL, FR or EN. Dropdowns and translations adapt automatically."),
    ("cfg_lang_change_hint",
        "Na taalwijziging: herselect bestaande classificatiewaarden in de gegevensbladen.",
        "Après changement de langue : re-sélectionnez les valeurs dans les feuilles de données.",
        "After language change: re-select existing classification values in data sheets."),
    ("cfg_version_label",    "Versie",          "Version",        "Version"),
    ("cfg_updated_label",    "Laatste update",  "Dernière mise à jour", "Last updated"),
    ("cfg_updated_by_label", "Bijgewerkt door", "Mis à jour par", "Updated by"),
    ("cfg_save_hint",
        "Sla het bestand op na elke wijziging (Ctrl+S).",
        "Enregistrez le fichier après chaque modification (Ctrl+S).",
        "Save the file after every change (Ctrl+S)."),
    # Info
    ("info_title",          "Over dit instrument",        "À propos de cet outil",     "About this tool"),
    ("info_section_doel",   "DOEL",                       "OBJECTIF",                  "PURPOSE"),
    ("info_doel_txt",
        "Deze GRC-tool ondersteunt overheidsorganisaties bij het systematisch beheren van informatieveiligheid. "
        "Processen en informatieassets worden geclassificeerd op integriteit, beschikbaarheid en confidentialiteit (5 niveaus: Laag → Kritiek).",
        "Cet outil GRC aide les organisations publiques à gérer systématiquement la sécurité de l'information. "
        "Les processus et actifs sont classifiés sur l'intégrité, la disponibilité et la confidentialité (5 niveaux : Faible → Critique).",
        "This GRC tool supports government organisations in systematically managing information security. "
        "Processes and assets are classified on integrity, availability and confidentiality (5 levels: Low → Critical)."),
    ("info_section_scope",  "TOEPASSINGSGEBIED",  "CHAMP D'APPLICATION", "SCOPE"),
    ("info_scope_txt",
        "Gericht op Belgische en Nederlandse overheidsdiensten (NIS2, GDPR, ISO 27001:2022, BIO).",
        "Destiné aux services publics belges et néerlandais (NIS2, RGPD, ISO 27001:2022, BIO).",
        "Aimed at Belgian and Dutch government services (NIS2, GDPR, ISO 27001:2022, BIO)."),
    ("info_section_modules", "MODULES IN DEZE VERSIE", "MODULES DANS CETTE VERSION", "MODULES IN THIS VERSION"),
    ("info_mod_proc",
        "Processen — Classificeer bedrijfsprocessen op I · B · C.",
        "Processus — Classifiez les processus métier sur I · D · C.",
        "Processes — Classify business processes on I · A · C."),
    ("info_mod_assets",
        "Informatieassets — Classificeer informatieactiva op I · B · C.",
        "Actifs informationnels — Classifiez les actifs sur I · D · C.",
        "Information Assets — Classify information assets on I · A · C."),
    ("info_mod_verant",
        "Verantwoordelijken — Contactregister van GRC-verantwoordelijken.",
        "Responsables — Registre de contacts des responsables GRC.",
        "Responsible Persons — Contact register of GRC responsible persons."),
    ("info_mod_import",
        "Import & Export — Macro's om gegevens te importeren uit externe bestanden.",
        "Import & Export — Macros pour importer des données depuis des fichiers externes.",
        "Import & Export — Macros to import data from external files."),
    ("info_section_guide",  "HOE GEBRUIKEN", "MODE D'EMPLOI", "HOW TO USE"),
    ("info_guide_1",
        "1. Configuratie: vul de naam van de instelling in en kies de taal (NL / FR / EN).",
        "1. Configuration : saisissez le nom et choisissez la langue (NL / FR / EN).",
        "1. Configuration: enter the name of the institution and choose the language (NL / FR / EN)."),
    ("info_guide_2",
        "2. Referentiewaarden: raadpleeg de definitie van elk classificatieniveau (1–5).",
        "2. Valeurs de référence : consultez la définition de chaque niveau de classification (1–5).",
        "2. Reference Values: consult the definition of each classification level (1–5)."),
    ("info_guide_3",
        "3. Processen & Informatieassets: vul de gegevens in. Kies niveau 1–5 per dimensie.",
        "3. Processus & Actifs : saisissez les données. Choisissez le niveau 1–5 par dimension.",
        "3. Processes & Assets: enter data. Choose level 1–5 per dimension."),
    ("info_guide_4",
        "4. Import & Export: gebruik de macro's om gegevens in bulk te importeren.",
        "4. Import & Export : utilisez les macros pour importer des données en masse.",
        "4. Import & Export: use the macros to import data in bulk."),
    ("info_section_version", "VERSIE-INFORMATIE", "INFORMATIONS DE VERSION", "VERSION INFORMATION"),
    ("info_v_tool",      "GRC Tool v0.3",                           "Outil GRC v0.3",                           "GRC Tool v0.3"),
    ("info_v_status",    "Status: Prototype / Fase 3",              "Statut : Prototype / Phase 3",             "Status: Prototype / Phase 3"),
    ("info_v_frameworks","Kaders: ISO 27001:2022 · NIS2 · GDPR · BIO",
                         "Cadres : ISO 27001:2022 · NIS2 · RGPD · BIO",
                         "Frameworks: ISO 27001:2022 · NIS2 · GDPR · BIO"),
    # Referentiewaarden
    ("ref_title",   "Referentiewaarden en Classificatieschalen",
                    "Valeurs de référence et échelles de classification",
                    "Reference Values and Classification Scales"),
    ("ref_subtitle",
        "Gebruik deze schalen bij het invullen van de classificatievelden (1 = Laag … 5 = Kritiek).",
        "Utilisez ces échelles lors du remplissage des champs de classification (1 = Faible … 5 = Critique).",
        "Use these scales when filling in classification fields (1 = Low … 5 = Critical)."),
    ("ref_level",       "Niveau",           "Niveau",           "Level"),
    ("ref_label",       "Benaming",         "Désignation",      "Designation"),
    ("ref_description", "Omschrijving",     "Description",      "Description"),
    ("ref_lvl1_label",  "Laag",             "Faible",           "Low"),
    ("ref_lvl2_label",  "Gemiddeld",        "Moyen",            "Medium"),
    ("ref_lvl3_label",  "Hoog",             "Élevé",            "High"),
    ("ref_lvl4_label",  "Zeer Hoog",        "Très élevé",       "Very High"),
    ("ref_lvl5_label",  "Classified",       "Classified",       "Classified"),
    ("ref_lvl1_desc",
        "Beperkte gevolgen. Verstoring of verlies heeft een verwaarloosbaar effect.",
        "Conséquences limitées. La perturbation ou la perte a un effet négligeable.",
        "Limited consequences. Disruption or loss has a negligible effect."),
    ("ref_lvl2_desc",
        "Merkbare gevolgen. Beperkt maar waarneembaar effect op de dienstverlening.",
        "Conséquences notables. Effet limité mais perceptible sur le service.",
        "Noticeable consequences. Limited but perceptible effect on service delivery."),
    ("ref_lvl3_desc",
        "Ernstige gevolgen. Significante schade aan de dienstverlening of betrokkenen.",
        "Conséquences graves. Dommages significatifs au service ou aux parties concernées.",
        "Serious consequences. Significant damage to service delivery or stakeholders."),
    ("ref_lvl4_desc",
        "Zeer ernstige gevolgen. Grote schade, gevaar voor personen of vitale functies.",
        "Conséquences très graves. Dommages importants, danger pour des personnes ou fonctions vitales.",
        "Very serious consequences. Major damage, danger to persons or vital functions."),
    ("ref_lvl5_desc",
        "Kritieke gevolgen. Onherstelbare schade, gevaar voor nationale veiligheid of grondrechten, of volledige uitval van vitale overheidsfuncties.",
        "Conséquences critiques. Dommages irréparables, danger pour la sécurité nationale ou les droits fondamentaux, ou arrêt total des fonctions gouvernementales vitales.",
        "Critical consequences. Irreparable damage, danger to national security or fundamental rights, or complete failure of vital government functions."),
    ("ref_section_int",   "INTEGRITEIT",      "INTÉGRITÉ",        "INTEGRITY"),
    ("ref_int_subtitle",  "Mate van correctheid, volledigheid en betrouwbaarheid van gegevens.",
                          "Degré de correction, d'exhaustivité et de fiabilité des données.",
                          "Degree of correctness, completeness and reliability of data."),
    ("ref_int_1",  "Fouten hebben geen impact op beslissingen of dienstverlening.",
                   "Les erreurs n'ont aucun impact sur les décisions ou le service.",
                   "Errors have no impact on decisions or service delivery."),
    ("ref_int_2",  "Fouten kunnen leiden tot beperkte fouten in beslissingen.",
                   "Les erreurs peuvent entraîner des erreurs limitées dans les décisions.",
                   "Errors may lead to limited mistakes in decisions."),
    ("ref_int_3",  "Fouten leiden tot verkeerde beslissingen met significante gevolgen.",
                   "Les erreurs entraînent des décisions erronées avec des conséquences significatives.",
                   "Errors lead to wrong decisions with significant consequences."),
    ("ref_int_4",  "Fouten leiden tot grove schade: gevaar, juridische gevolgen, schending van rechten.",
                   "Les erreurs entraînent des dommages graves : danger, conséquences juridiques, violation des droits.",
                   "Errors lead to serious damage: danger, legal consequences, violation of rights."),
    ("ref_int_5",  "Fouten kunnen leiden tot onherstelbare schade, gevaar voor nationale veiligheid of fundamentele schending van grondrechten.",
                   "Les erreurs peuvent entraîner des dommages irréparables, un danger pour la sécurité nationale ou une violation fondamentale des droits.",
                   "Errors may lead to irreparable damage, danger to national security or fundamental violation of rights."),
    ("ref_section_avail", "BESCHIKBAARHEID",  "DISPONIBILITÉ",    "AVAILABILITY"),
    ("ref_avail_subtitle","Mate van toegankelijkheid wanneer dat nodig is.",
                          "Degré de disponibilité lorsque nécessaire.",
                          "Degree of accessibility when needed."),
    ("ref_avail_1","Uitval heeft geen merkbare impact. Processen lopen handmatig verder.",
                   "La panne n'a pas d'impact. Les processus continuent manuellement.",
                   "Outage has no noticeable impact. Processes continue manually."),
    ("ref_avail_2","Uitval veroorzaakt hinder maar geen kritieke verstoring.",
                   "La panne cause des inconvénients mais pas de perturbation critique.",
                   "Outage causes inconvenience but no critical disruption."),
    ("ref_avail_3","Uitval leidt tot significante verstoring van de dienstverlening.",
                   "La panne entraîne une perturbation significative du service.",
                   "Outage leads to significant disruption of services."),
    ("ref_avail_4","Uitval veroorzaakt grote schade of stillegging van belangrijke diensten.",
                   "La panne provoque des dommages importants ou l'arrêt de services importants.",
                   "Outage causes major damage or shutdown of important services."),
    ("ref_avail_5","Uitval leidt tot volledige stopzetting van vitale overheidsfuncties of noodsituaties met gevaar voor leven.",
                   "La panne entraîne l'arrêt total des fonctions gouvernementales vitales ou des situations d'urgence mettant des vies en danger.",
                   "Outage leads to complete failure of vital government functions or emergencies endangering lives."),
    ("ref_section_conf",  "CONFIDENTIALITEIT","CONFIDENTIALITÉ",  "CONFIDENTIALITY"),
    ("ref_conf_subtitle", "Toegankelijkheid enkel voor bevoegde personen of systemen.",
                          "Accessibilité uniquement aux personnes ou systèmes autorisés.",
                          "Accessibility only to authorised persons or systems."),
    ("ref_conf_1", "Openbare gegevens. Verspreiding heeft geen nadelige gevolgen.",
                   "Données publiques. La diffusion n'a pas de conséquences négatives.",
                   "Public data. Disclosure has no adverse consequences."),
    ("ref_conf_2", "Interne gegevens. Verspreiding kan leiden tot beperkte reputatieschade.",
                   "Données internes. La diffusion peut entraîner des dommages à la réputation.",
                   "Internal data. Disclosure may lead to limited reputational damage."),
    ("ref_conf_3", "Vertrouwelijke gegevens. Ongeoorloofde verspreiding heeft ernstige gevolgen.",
                   "Données confidentielles. La diffusion non autorisée a de graves conséquences.",
                   "Confidential data. Unauthorised disclosure has serious consequences."),
    ("ref_conf_4", "Strikt vertrouwelijke gegevens. Verspreiding bedreigt personen, organisaties of vitale processen.",
                   "Données strictement confidentielles. La diffusion menace des personnes, organisations ou processus vitaux.",
                   "Strictly confidential data. Disclosure threatens persons, organisations or vital processes."),
    ("ref_conf_5", "Geheime gegevens. Verspreiding kan nationale veiligheid, grondrechten of vitale staatsfuncties schaden.",
                   "Données secrètes. La diffusion peut nuire à la sécurité nationale, aux droits fondamentaux ou aux fonctions vitales de l'État.",
                   "Secret data. Disclosure may harm national security, fundamental rights or vital state functions."),
    # Processen
    ("proc_title",          "Processen",         "Processus",        "Processes"),
    ("proc_subtitle",
        "Voer de bedrijfsprocessen in en ken de veiligheidclassificatie toe (1 = Laag … 5 = Kritiek).",
        "Saisissez les processus et attribuez la classification de sécurité (1 = Faible … 5 = Critique).",
        "Enter the business processes and assign the security classification (1 = Low … 5 = Critical)."),
    ("proc_col_id",           "ID",                "ID",               "ID"),
    ("proc_col_naam",         "Naam",              "Nom",              "Name"),
    ("proc_col_omschrijving", "Omschrijving",      "Description",      "Description"),
    ("proc_col_eigenaar",     "Proceseigenaar",    "Propriétaire",     "Process Owner"),
    ("proc_col_dienst",       "Dienst / Entiteit", "Service / Entité", "Department / Entity"),
    ("proc_col_integriteit",  "Integriteit",       "Intégrité",        "Integrity"),
    ("proc_col_beschikbaar",  "Beschikbaarheid",   "Disponibilité",    "Availability"),
    ("proc_col_conf",         "Confidentialiteit", "Confidentialité",  "Confidentiality"),
    ("proc_col_assets",       "Informatieassets",  "Actifs informationnels", "Information Assets"),
    ("proc_col_opmerkingen",  "Opmerkingen",       "Remarques",        "Remarks"),
    # Informatieassets
    ("asset_title",           "Informatieassets",  "Actifs informationnels", "Information Assets"),
    ("asset_subtitle",
        "Voer de informatieactiva in en ken de veiligheidclassificatie toe (1 = Laag … 5 = Kritiek).",
        "Saisissez les actifs et attribuez la classification de sécurité (1 = Faible … 5 = Critique).",
        "Enter the information assets and assign the security classification (1 = Low … 5 = Critical)."),
    ("asset_col_id",          "ID",                "ID",               "ID"),
    ("asset_col_naam",        "Naam",              "Nom",              "Name"),
    ("asset_col_type",        "Type",              "Type",             "Type"),
    ("asset_col_omschrijving","Omschrijving",      "Description",      "Description"),
    ("asset_col_eigenaar",    "Eigenaar",          "Propriétaire",     "Owner"),
    ("asset_col_dienst",      "Dienst / Entiteit", "Service / Entité", "Department / Entity"),
    ("asset_col_integriteit", "Integriteit",       "Intégrité",        "Integrity"),
    ("asset_col_beschikbaar", "Beschikbaarheid",   "Disponibilité",    "Availability"),
    ("asset_col_conf",        "Confidentialiteit", "Confidentialité",  "Confidentiality"),
    ("asset_col_opmerkingen", "Opmerkingen",       "Remarques",        "Remarks"),
    # Verantwoordelijken
    ("verant_title",           "Verantwoordelijken", "Responsables",    "Responsible Persons"),
    ("verant_subtitle",
        "Overzicht van contactpersonen en hun GRC-rol.",
        "Aperçu des personnes de contact et de leur rôle GRC.",
        "Overview of contact persons and their GRC role."),
    ("verant_col_id",          "ID",                "ID",               "ID"),
    ("verant_col_naam",        "Naam",              "Nom",              "Name"),
    ("verant_col_voornaam",    "Voornaam",          "Prénom",           "First Name"),
    ("verant_col_functie",     "Functietitel",      "Titre de fonction","Job Title"),
    ("verant_col_dienst",      "Dienst / Entiteit", "Service / Entité", "Department / Entity"),
    ("verant_col_email",       "E-mailadres",       "Adresse e-mail",   "Email Address"),
    ("verant_col_tel",         "Telefoon",          "Téléphone",        "Phone"),
    ("verant_col_rol",         "Rol in GRC",        "Rôle dans le GRC", "Role in GRC"),
    ("verant_col_opmerkingen", "Opmerkingen",       "Remarques",        "Remarks"),
    # Import & Export
    ("ie_title",     "Import & Export",  "Import & Export",  "Import & Export"),
    ("ie_subtitle",
        "Importeer gegevens vanuit externe bronnen of exporteer het GRC-register.",
        "Importez des données depuis des sources externes ou exportez le registre GRC.",
        "Import data from external sources or export the GRC register."),
    ("ie_section_import",  "IMPORTEREN",    "IMPORTATION",    "IMPORT"),
    ("ie_section_export",  "EXPORTEREN",    "EXPORTATION",    "EXPORT"),
    ("ie_section_info",    "TABELSTRUCTUUR BRONBESTAND", "STRUCTURE DU FICHIER SOURCE", "SOURCE FILE TABLE STRUCTURE"),
    ("ie_import_all_btn",
        "Importeer Alles",
        "Importer tout",
        "Import All"),
    ("ie_import_all_desc",
        "Importeert informatieassets, afhankelijke assets, processen én kwetsbaarheden-matrix in één stap.",
        "Importe actifs informationnels, actifs dépendants, processus et matrice des vulnérabilités en une seule étape.",
        "Imports information assets, dependent assets, processes and vulnerability matrix in one step."),
    ("ie_import_ia_btn",
        "Importeer Informatieassets",
        "Importer actifs informationnels",
        "Import Information Assets"),
    ("ie_import_ia_desc",
        "Leest de tabel 'T - Information Assets' uit de Access-database en vult de tab Informatieassets.",
        "Lit la table 'T - Information Assets' depuis la base Access et remplit l'onglet Actifs informationnels.",
        "Reads the 'T - Information Assets' table from the Access database and fills the Information Assets tab."),
    ("ie_import_da_btn",
        "Importeer Afhankelijke Assets",
        "Importer actifs dépendants",
        "Import Dependent Assets"),
    ("ie_import_da_desc",
        "Leest de tabel 'T - Dependent assets' en vult de tab Afhankelijke Assets inclusief CIA-objectives.",
        "Lit la table 'T - Dependent assets' et remplit l'onglet Actifs dépendants avec les objectifs CIA.",
        "Reads the 'T - Dependent assets' table and fills the Dependent Assets tab including CIA objectives."),
    ("ie_import_proc_btn",
        "Importeer Processen",
        "Importer processus",
        "Import Processes"),
    ("ie_import_proc_desc",
        "Leest 'T - Processes in scope' en vult de tab Processen inclusief koppelingen aan assets.",
        "Lit 'T - Processes in scope' et remplit l'onglet Processus avec les liens aux actifs.",
        "Reads 'T - Processes in scope' and fills the Processes tab including asset links."),
    ("ie_import_kwets_btn",
        "Importeer Kwetsbaarheden",
        "Importer vulnérabilités",
        "Import Vulnerabilities"),
    ("ie_import_kwets_desc",
        "Laadt de kwetsbaarheden en hun koppeling aan CyFun-controls. Vereist dat de tab Kwetsbaarheden al aangemaakt is.",
        "Charge les vulnérabilités et leur lien aux contrôles CyFun. Nécessite que l'onglet Vulnérabilités soit déjà créé.",
        "Loads vulnerabilities and their link to CyFun controls. Requires the Vulnerabilities tab to already exist."),
    ("ie_import_links_btn",
        "Importeer Links DA / Kwetsbaarheden",
        "Importer liens actifs dépendants / vulnérabilités",
        "Import DA / Vulnerability Links"),
    ("ie_import_links_desc",
        "Voer dit uit NA de import van kwetsbaarheden. Leest de koppelingen tussen afhankelijke assets en kwetsbaarheden uit de Access-database en vult de RARM-tab aan.",
        "À exécuter APRÈS l'import des vulnérabilités. Lit les liens entre actifs dépendants et vulnérabilités depuis la base Access et complète l'onglet RARM.",
        "Run this AFTER importing vulnerabilities. Reads links between dependent assets and vulnerabilities from the Access database and fills in the RARM tab."),
    ("ie_export_btn",      "Exporteer alle gegevens",     "Exporter toutes les données",   "Export all data"),
    ("ie_export_desc",
        "Exporteert de bladen Processen, Informatieassets en Verantwoordelijken naar een nieuw Excel-bestand.",
        "Exporte les feuilles Processus, Actifs et Responsables vers un nouveau fichier Excel.",
        "Exports the Processes, Information Assets and Responsible Persons sheets to a new Excel file."),
    ("ie_table_proc",  "T - Processes in scope",    "T - Processes in scope",    "T - Processes in scope"),
    ("ie_table_asset", "T - Information Assets",    "T - Information Assets",    "T - Information Assets"),
    ("ie_table_dep",   "T - Dependent assets",      "T - Dependent assets",      "T - Dependent assets"),
    # Afhankelijke assets sheet
    ("dep_title",    "Afhankelijke Assets",
                     "Actifs dépendants",
                     "Dependent Assets"),
    ("dep_subtitle",
        "Overzicht van systemen, infrastructuur en diensten waarvan processen afhankelijk zijn",
        "Vue d'ensemble des systèmes, infrastructures et services dont dépendent les processus",
        "Overview of systems, infrastructure and services that processes depend on"),
    ("dep_col_id",           "ID",                  "ID",                      "ID"),
    ("dep_col_naam",         "Naam",                "Nom",                     "Name"),
    ("dep_col_omschrijving", "Omschrijving",        "Description",             "Description"),
    ("dep_col_eigenaar",     "Eigenaar",            "Propriétaire",            "Owner"),
    ("dep_col_dienst",       "Dienst / Entiteit",   "Service / Entité",        "Department / Entity"),
    ("dep_col_overarching",  "Overarching",         "Transversal",             "Overarching"),
    ("dep_col_opmerkingen",  "Opmerkingen",         "Remarques",               "Remarks"),
    ("dep_col_processes",
        "Gebruikt door processen",
        "Utilisé par les processus",
        "Used by Processes"),
    ("dep_col_linked_info",
        "Informatieassets (via processen)",
        "Actifs informationnels (via processus)",
        "Information Assets (via processes)"),
    ("dep_sec_req_header",
        "Security Requirements",
        "Exigences de sécurité",
        "Security Requirements"),
    ("dep_col_conf",
        "Confidentialiteit",
        "Confidentialité",
        "Confidentiality"),
    ("dep_col_integ",
        "Integriteit",
        "Intégrité",
        "Integrity"),
    ("dep_col_avail",
        "Beschikbaarheid",
        "Disponibilité",
        "Availability"),
    ("dep_sec_obj_header",
        "Security Objectives",
        "Objectifs de sécurité",
        "Security Objectives"),
    ("dep_col_obj_commentaar",
        "Commentaar",
        "Commentaire",
        "Comment"),
    ("dep_gap_header",
        "Gap Analyse",
        "Analyse des écarts",
        "Gap Analysis"),
    ("dep_col_gap_conf",
        "Conf. niet gedekt (info assets)",
        "Conf. non couverte (actifs info)",
        "Conf. not covered (info assets)"),
    ("dep_col_gap_integ",
        "Int. niet gedekt (processen)",
        "Int. non couverte (processus)",
        "Int. not covered (processes)"),
    ("dep_col_gap_avail",
        "Beschikb. niet gedekt (processen)",
        "Disp. non couverte (processus)",
        "Avail. not covered (processes)"),
    # Extra kolom in Processen
    ("proc_col_dep_assets",  "Afhankelijke assets", "Actifs dépendants",       "Dependent Assets"),
    # Informatieassets extra kolom
    ("asset_col_processes",
        "Gebruikt in processen",
        "Utilisé dans les processus",
        "Used in Processes"),
    # Navigatie
    ("ui_goto_dep",          "→  Dependent Assets",    "→  Actifs dépendants", "→  Dependent Assets"),
    ("ie_col_hint",
        "Ondersteunde kolomnamen (niet hoofdlettergevoelig):",
        "Noms de colonnes pris en charge (insensibles à la casse) :",
        "Supported column names (case-insensitive):"),
    ("ie_macro_hint",
        "De macro's zijn ingebouwd in dit bestand (Module: GRC_Macros). Klik op de knoppen hierboven om te starten.",
        "Les macros sont intégrées dans ce fichier (Module : GRC_Macros). Cliquez sur les boutons ci-dessus pour démarrer.",
        "The macros are built into this file (Module: GRC_Macros). Click the buttons above to start."),
]

# ── VBA-macrocode ─────────────────────────────────────────────────────────────
# Elke VBA-string wordt via win32com in het juiste VBComponent geïnjecteerd.
# De strings bevatten volledig geldige VBA-code; Python-aanhalingstekens zijn
# hier slechts de "verpakking" voor het transport naar de VBE.

KWETS_SHEET_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module : Kwetsbaarheden (Worksheet)
' Doel   : Laat de gebruiker met een dubbelklik een ✔-vinkje plaatsen of
'          verwijderen in de matrix (rij 3+ / kolom 3+).
'          Rijen 1-2 bevatten de titelregel en kwetsbaarheid-koppen;
'          kolommen A-B bevatten Control ID en Richtlijn — beide zijn
'          beschermde cellen en mogen niet worden gewijzigd.
' ══════════════════════════════════════════════════════════════════════════════

' ── Worksheet_BeforeDoubleClick ───────────────────────────────────────────────
' Onderschept een dubbelklik in de matrix en schakelt het ✔-vinkje (U+2714)
' aan of uit. Cancel = True voorkomt dat Excel het cel-bewerkingsvenster opent.
Private Sub Worksheet_BeforeDoubleClick(ByVal Target As Range, Cancel As Boolean)
    If Target.Cells.Count > 1 Then Exit Sub
    If Target.Row < 3 Or Target.Column < 3 Then Exit Sub
    Cancel = True
    Dim sVal As String
    sVal = ""
    If Not IsError(Target.Value) Then sVal = CStr(Target.Value)
    ' Toggle: vinkje aanwezig → verwijderen, afwezig → plaatsen
    If sVal = ChrW(10004) Then
        Target.Value = ""
    Else
        Target.Value = ChrW(10004)
        Target.HorizontalAlignment = xlCenter
        Target.VerticalAlignment = xlCenter
    End If
End Sub
'''

INFO_ASSET_SHEET_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module : Information Assets (Worksheet)
' Doel   : Vult kolom I ("Gebruikt in processen") automatisch bij
'          bij elke activering van het blad.
' ══════════════════════════════════════════════════════════════════════════════

' ── Worksheet_Activate ───────────────────────────────────────────────────────
' Wanneer de gebruiker het blad opent, worden kolom I (rijen 6-105) opnieuw
' berekend door het Processes-blad te scannen.
' Voor elk informatieasset (kolom B) wordt gecontroleerd of de naam voorkomt
' in kolom K (Informatieassets-kolom) van elk proces.
' De match is hoofdletterongevoelig (LCase) om tikfouten te vermijden.
' AutoFit past de rijhoogte aan wanneer meerdere processen worden gevonden.
Private Sub Worksheet_Activate()
    Dim procWs As Worksheet
    Dim r As Long, c As Long
    Dim assetNaam As String, pNaam As String, procList As String
    Application.ScreenUpdating = False
    Set procWs = ThisWorkbook.Sheets("Processes")
    Me.Range("I6:I105").ClearContents
    ' ── Itereer over alle informatieassets (rij 6-105) ──────────────────────
    For r = 6 To 105
        assetNaam = CStr(Me.Cells(r, 2).Value)
        If assetNaam <> "" Then
            procList = ""
            ' ── Zoek asset-naam terug in kolom K van elk proces ─────────────
            ' Processen slagen meerdere asset-namen op als newline-gescheiden
            ' tekst in kolom K; InStr detecteert deelstringmatches.
            For c = 6 To 105
                pNaam = CStr(procWs.Cells(c, 2).Value)
                If pNaam <> "" Then
                    If InStr(LCase(CStr(procWs.Cells(c, 11).Value)), LCase(assetNaam)) > 0 Then
                        If procList <> "" Then procList = procList & Chr(10)
                        procList = procList & pNaam
                    End If
                End If
            Next c
            Me.Cells(r, 9) = procList
            If procList <> "" Then Me.Rows(r).AutoFit
        End If
    Next r
    Application.ScreenUpdating = True
End Sub
'''

DEP_ASSET_SHEET_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module : Dependent Assets (Worksheet)
' Doel   : Berekent automatisch de volgende kolommen bij activering van het blad
'          én bij wijziging van Security Objective-cellen (P/R/T = cols 16/18/20):
'
'   H (8)  = Gebruikt door processen (uit Processes kolom L, omgekeerde lookup)
'   I (9)  = Gekoppelde informatieassets (via de proces-asset-koppeling)
'   J (10) = Max. Confidentialiteit-requirement (max C van gekoppelde IA's)
'   L (12) = Max. Integriteit-requirement (max I van gekoppelde processen)
'   N (14) = Max. Beschikbaarheid-requirement (max A van gekoppelde processen)
'   V (22) = Gap Confidentialiteit: IA-namen waarvoor C > C-objective
'   W (23) = Gap Integriteit: procesnamen waarvoor I > I-objective
'   X (24) = Gap Beschikbaarheid: procesnamen waarvoor A > A-objective
'
' Beveiligingsmodel: het blad is beveiligd (Contents=True), maar UserInterfaceOnly
' laat VBA toe te schrijven zonder dat de gebruiker handmatig kan bewerken.
' ══════════════════════════════════════════════════════════════════════════════

' ── Worksheet_Activate ───────────────────────────────────────────────────────
' Herberekent alle afgeleide kolommen voor alle afhankelijke assets (rij 6-105).
' Volgorde van berekening:
'   1. Verzamel gekoppelde processen (col H) via omgekeerde lookup in Processes!L
'   2. Verzamel gekoppelde IA's (col I) via de informatieassets-kolommen van die processen
'   3. Bereken max C/I/A requirements op basis van IA's en processen
'   4. Bereken gap-analyse op basis van Security Objectives (cols P/R/T)
Private Sub Worksheet_Activate()
    Dim procWs As Worksheet, iaWs As Worksheet
    Dim r As Long, c As Long, p As Long, ia As Long
    Dim depNaam As String, pNaam As String, iaName As String
    Dim procList As String, infoList As String, iaStr As String
    Dim parts() As String
    Dim maxConf As Integer, maxInt As Integer, maxAvail As Integer
    Dim objConf As Integer, objInt As Integer, objAvail As Integer
    Dim gapConf As String, gapInt As String, gapAvail As String
    Dim v As Variant
    Application.ScreenUpdating = False
    Me.Unprotect
    Set procWs = ThisWorkbook.Sheets("Processes")
    Set iaWs = ThisWorkbook.Sheets("Information Assets")
    ' Reset de berekende kolommen vooraleer opnieuw te vullen
    Me.Range("H6:J105").ClearContents
    Me.Range("L6:L105").ClearContents
    Me.Range("N6:N105").ClearContents
    Me.Range("V6:X105").ClearContents
    For r = 6 To 105
        depNaam = CStr(Me.Cells(r, 2).Value)
        If depNaam <> "" Then
            procList = "": infoList = ""
            maxConf = 0: maxInt = 0: maxAvail = 0
            ' ── Stap 1: zoek alle processen die deze DA gebruiken ────────────
            ' Processes!L (kolom 12) bevat newline-gescheiden DA-namen per proces.
            ' Omgekeerde lookup: als depNaam voorkomt in die kolom → proces is gekoppeld.
            For c = 6 To 105
                pNaam = CStr(procWs.Cells(c, 2).Value)
                If pNaam <> "" Then
                    If InStr(LCase(CStr(procWs.Cells(c, 12).Value)), LCase(depNaam)) > 0 Then
                        If procList <> "" Then procList = procList & Chr(10)
                        procList = procList & pNaam
                        ' ── Stap 2a: max I en A via gekoppelde processen ─────
                        ' Processen dragen I (col 6) en A (col 8) requirements bij.
                        v = procWs.Cells(c, 6).Value
                        If IsNumeric(v) And CInt(v) > maxInt Then maxInt = CInt(v)
                        v = procWs.Cells(c, 8).Value
                        If IsNumeric(v) And CInt(v) > maxAvail Then maxAvail = CInt(v)
                        ' ── Stap 2b: verzamel gekoppelde IA's (deduplicatie) ─
                        ' Processes!K (kolom 11) bevat newline-gescheiden IA-namen.
                        iaStr = CStr(procWs.Cells(c, 11).Value)
                        If iaStr <> "" Then
                            parts = Split(iaStr, Chr(10))
                            For p = 0 To UBound(parts)
                                iaName = Trim(parts(p))
                                If iaName <> "" Then
                                    ' Voeg IA toe aan infoList als nog niet aanwezig
                                    If InStr(Chr(10) & infoList & Chr(10), Chr(10) & iaName & Chr(10)) = 0 Then
                                        If infoList <> "" Then infoList = infoList & Chr(10)
                                        infoList = infoList & iaName
                                    End If
                                    ' ── Stap 2c: max C via IA-confidentialiteit ──
                                    ' Informatieassets!F (kolom 6) = C-score.
                                    For ia = 6 To 105
                                        If LCase(CStr(iaWs.Cells(ia, 2).Value)) = LCase(iaName) Then
                                            v = iaWs.Cells(ia, 6).Value
                                            If IsNumeric(v) And CInt(v) > maxConf Then maxConf = CInt(v)
                                            Exit For
                                        End If
                                    Next ia
                                End If
                            Next p
                        End If
                    End If
                End If
            Next c
            Me.Cells(r, 8) = procList
            Me.Cells(r, 9) = infoList
            If maxConf > 0 Then Me.Cells(r, 10) = maxConf
            If maxInt > 0 Then Me.Cells(r, 12) = maxInt
            If maxAvail > 0 Then Me.Cells(r, 14) = maxAvail
            ' ── Stap 3: lees Security Objectives (editeerbaar door gebruiker) ─
            ' Kolom P(16)=C-obj, R(18)=I-obj, T(20)=A-obj
            objConf = 0: objInt = 0: objAvail = 0
            If IsNumeric(Me.Cells(r, 16).Value) Then objConf = CInt(Me.Cells(r, 16).Value)
            If IsNumeric(Me.Cells(r, 18).Value) Then objInt = CInt(Me.Cells(r, 18).Value)
            If IsNumeric(Me.Cells(r, 20).Value) Then objAvail = CInt(Me.Cells(r, 20).Value)
            ' ── Stap 4: Gap-analyse per CIA-dimensie ─────────────────────────
            ' Gaplogica: als requirement > objective én de lijst niet leeg is →
            ' toon de namen van de entiteiten die het tekort veroorzaken.
            ' Confidentialiteit-gap: welke IA's hebben C > C-objective?
            gapConf = ""
            If objConf > 0 And maxConf > objConf And infoList <> "" Then
                parts = Split(infoList, Chr(10))
                For p = 0 To UBound(parts)
                    iaName = Trim(parts(p))
                    If iaName <> "" Then
                        For ia = 6 To 105
                            If LCase(CStr(iaWs.Cells(ia, 2).Value)) = LCase(iaName) Then
                                v = iaWs.Cells(ia, 6).Value
                                If IsNumeric(v) And CInt(v) > objConf Then
                                    If gapConf <> "" Then gapConf = gapConf & Chr(10)
                                    gapConf = gapConf & iaName
                                End If
                                Exit For
                            End If
                        Next ia
                    End If
                Next p
            End If
            Me.Cells(r, 22) = gapConf
            ' Integriteit-gap: welke processen hebben I > I-objective?
            gapInt = ""
            If objInt > 0 And maxInt > objInt And procList <> "" Then
                parts = Split(procList, Chr(10))
                For p = 0 To UBound(parts)
                    pNaam = Trim(parts(p))
                    If pNaam <> "" Then
                        For c = 6 To 105
                            If LCase(CStr(procWs.Cells(c, 2).Value)) = LCase(pNaam) Then
                                v = procWs.Cells(c, 6).Value
                                If IsNumeric(v) And CInt(v) > objInt Then
                                    If gapInt <> "" Then gapInt = gapInt & Chr(10)
                                    gapInt = gapInt & pNaam
                                End If
                                Exit For
                            End If
                        Next c
                    End If
                Next p
            End If
            Me.Cells(r, 23) = gapInt
            ' Beschikbaarheid-gap: welke processen hebben A > A-objective?
            gapAvail = ""
            If objAvail > 0 And maxAvail > objAvail And procList <> "" Then
                parts = Split(procList, Chr(10))
                For p = 0 To UBound(parts)
                    pNaam = Trim(parts(p))
                    If pNaam <> "" Then
                        For c = 6 To 105
                            If LCase(CStr(procWs.Cells(c, 2).Value)) = LCase(pNaam) Then
                                v = procWs.Cells(c, 8).Value
                                If IsNumeric(v) And CInt(v) > objAvail Then
                                    If gapAvail <> "" Then gapAvail = gapAvail & Chr(10)
                                    gapAvail = gapAvail & pNaam
                                End If
                                Exit For
                            End If
                        Next c
                    End If
                Next p
            End If
            Me.Cells(r, 24) = gapAvail
            If procList <> "" Or infoList <> "" Then Me.Rows(r).AutoFit
        End If
    Next r
    ' Herstel bladbeveiliging — UserInterfaceOnly laat toekomstige macro-aanroepen toe
    Me.Protect DrawingObjects:=True, Contents:=True, Scenarios:=True, UserInterfaceOnly:=True, AllowFiltering:=True, AllowSorting:=True
    Application.ScreenUpdating = True
End Sub

' ── Worksheet_BeforeDoubleClick (Overarching-kolom F) ────────────────────────
' Laat de gebruiker dubbelklikken op kolom F (Overarching) om het ✔-vinkje
' te plaatsen of te verwijderen. Overarching = True betekent dat deze DA
' een organisatorische of infrastructurele asset is die van toepassing is
' op alle processen, ook als er geen expliciete koppeling bestaat.
Private Sub Worksheet_BeforeDoubleClick(ByVal Target As Range, Cancel As Boolean)
    If Target.Cells.Count > 1 Then Exit Sub
    If Target.Row < 6 Or Target.Row > 105 Then Exit Sub
    If Target.Column <> 6 Then Exit Sub
    Cancel = True
    If CStr(Target.Value) = ChrW(10004) Then
        Target.Value = ""
    Else
        Target.Value = ChrW(10004)
        Target.HorizontalAlignment = xlCenter
        Target.VerticalAlignment = xlCenter
    End If
End Sub

' ── Worksheet_Change ─────────────────────────────────────────────────────────
' Herberekent de gap-analyse wanneer de gebruiker een Security Objective-waarde
' (kolommen P/R/T = 16/18/20) aanpast — zonder het volledige blad opnieuw te
' activeren. rowDone() voorkomt dubbele berekening wanneer meerdere cellen
' tegelijk worden gewijzigd (bv. via Plakken).
Private Sub Worksheet_Change(ByVal Target As Range)
    Dim cell As Range
    Dim hasObjChange As Boolean
    hasObjChange = False
    For Each cell In Target.Cells
        If cell.Row >= 6 And cell.Row <= 105 Then
            If cell.Column = 16 Or cell.Column = 18 Or cell.Column = 20 Then
                hasObjChange = True
                Exit For
            End If
        End If
    Next cell
    If Not hasObjChange Then Exit Sub
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    Me.Unprotect
    Dim procWs As Worksheet, iaWs As Worksheet
    Dim r As Long, c As Long, p As Long, ia As Long
    Dim pNaam As String, iaName As String
    Dim procList As String, infoList As String
    Dim parts() As String
    Dim v As Variant
    Dim maxConf As Integer, maxInt As Integer, maxAvail As Integer
    Dim objConf As Integer, objInt As Integer, objAvail As Integer
    Dim gapConf As String, gapInt As String, gapAvail As String
    Dim rowDone(6 To 105) As Boolean
    Set procWs = ThisWorkbook.Sheets("Processes")
    Set iaWs = ThisWorkbook.Sheets("Information Assets")
    For Each cell In Target.Cells
        r = cell.Row
        If r >= 6 And r <= 105 Then
            If (cell.Column = 16 Or cell.Column = 18 Or cell.Column = 20) And Not rowDone(r) Then
                rowDone(r) = True
                procList = CStr(Me.Cells(r, 8).Value)
                infoList = CStr(Me.Cells(r, 9).Value)
                maxConf = 0: maxInt = 0: maxAvail = 0
                If IsNumeric(Me.Cells(r, 10).Value) Then maxConf = CInt(Me.Cells(r, 10).Value)
                If IsNumeric(Me.Cells(r, 12).Value) Then maxInt = CInt(Me.Cells(r, 12).Value)
                If IsNumeric(Me.Cells(r, 14).Value) Then maxAvail = CInt(Me.Cells(r, 14).Value)
                objConf = 0: objInt = 0: objAvail = 0
                If IsNumeric(Me.Cells(r, 16).Value) Then objConf = CInt(Me.Cells(r, 16).Value)
                If IsNumeric(Me.Cells(r, 18).Value) Then objInt = CInt(Me.Cells(r, 18).Value)
                If IsNumeric(Me.Cells(r, 20).Value) Then objAvail = CInt(Me.Cells(r, 20).Value)
                gapConf = ""
                If objConf > 0 And maxConf > objConf And infoList <> "" Then
                    parts = Split(infoList, Chr(10))
                    For p = 0 To UBound(parts)
                        iaName = Trim(parts(p))
                        If iaName <> "" Then
                            For ia = 6 To 105
                                If LCase(CStr(iaWs.Cells(ia, 2).Value)) = LCase(iaName) Then
                                    v = iaWs.Cells(ia, 6).Value
                                    If IsNumeric(v) And CInt(v) > objConf Then
                                        If gapConf <> "" Then gapConf = gapConf & Chr(10)
                                        gapConf = gapConf & iaName
                                    End If
                                    Exit For
                                End If
                            Next ia
                        End If
                    Next p
                End If
                Me.Cells(r, 22) = gapConf
                gapInt = ""
                If objInt > 0 And maxInt > objInt And procList <> "" Then
                    parts = Split(procList, Chr(10))
                    For p = 0 To UBound(parts)
                        pNaam = Trim(parts(p))
                        If pNaam <> "" Then
                            For c = 6 To 105
                                If LCase(CStr(procWs.Cells(c, 2).Value)) = LCase(pNaam) Then
                                    v = procWs.Cells(c, 6).Value
                                    If IsNumeric(v) And CInt(v) > objInt Then
                                        If gapInt <> "" Then gapInt = gapInt & Chr(10)
                                        gapInt = gapInt & pNaam
                                    End If
                                    Exit For
                                End If
                            Next c
                        End If
                    Next p
                End If
                Me.Cells(r, 23) = gapInt
                gapAvail = ""
                If objAvail > 0 And maxAvail > objAvail And procList <> "" Then
                    parts = Split(procList, Chr(10))
                    For p = 0 To UBound(parts)
                        pNaam = Trim(parts(p))
                        If pNaam <> "" Then
                            For c = 6 To 105
                                If LCase(CStr(procWs.Cells(c, 2).Value)) = LCase(pNaam) Then
                                    v = procWs.Cells(c, 8).Value
                                    If IsNumeric(v) And CInt(v) > objAvail Then
                                        If gapAvail <> "" Then gapAvail = gapAvail & Chr(10)
                                        gapAvail = gapAvail & pNaam
                                    End If
                                    Exit For
                                End If
                            Next c
                        End If
                    Next p
                End If
                Me.Cells(r, 24) = gapAvail
                Me.Rows(r).AutoFit
            End If
        End If
    Next cell
    Me.Protect DrawingObjects:=True, Contents:=True, Scenarios:=True, UserInterfaceOnly:=True, AllowFiltering:=True, AllowSorting:=True
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub
'''

USERFORM_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module  : AssetPicker (UserForm)
' Doel    : Popup-dialoogvenster waarmee de gebruiker één of meer assets
'           koppelt aan een proces.
' Gebruik : Verschijnt automatisch wanneer de gebruiker op kolom K
'           (Informatieassets) of kolom L (Afhankelijke assets) klikt in
'           het Processes-blad. De bronsheet en doelkolom worden doorgegeven
'           via globale variabelen in GRC_Macros:
'             g_TargetRow       — rijnummer in Processes
'             g_ProcesNaam      — naam van het geselecteerde proces (voor caption)
'             g_PickerSourceSheet — "Information Assets" of "Dependent Assets"
'             g_PickerTargetCol — 11 (K) of 12 (L)
' ══════════════════════════════════════════════════════════════════════════════

' ── UserForm_Initialize ───────────────────────────────────────────────────────
' Vult de listbox met alle asset-namen uit de bronsheet (rijen 6-105, kolom B).
' Reeds geselecteerde assets (aanwezig in de huidige celwaarde) worden
' automatisch aangevinkt. De vergelijking is hoofdletterongevoelig.
Private Sub UserForm_Initialize()
    Me.Caption = "Assets voor: " & GRC_Macros.g_ProcesNaam
    lstAssets.Clear
    Dim assetWs As Worksheet
    Set assetWs = ThisWorkbook.Sheets(GRC_Macros.g_PickerSourceSheet)
    Dim huidigStr As String
    huidigStr = LCase(CStr(ThisWorkbook.Sheets("Processes").Cells(GRC_Macros.g_TargetRow, GRC_Macros.g_PickerTargetCol).Value))
    Dim r As Long
    For r = 6 To 105
        Dim assetNaam As String
        assetNaam = CStr(assetWs.Cells(r, 2).Value)
        If assetNaam <> "" Then
            lstAssets.AddItem assetNaam
            Dim idx As Long: idx = lstAssets.ListCount - 1
            Dim lines() As String
            lines = Split(huidigStr, Chr(10))
            Dim li As Long, isSelected As Boolean
            isSelected = False
            For li = 0 To UBound(lines)
                If Trim(lines(li)) = LCase(Trim(assetNaam)) Then isSelected = True: Exit For
            Next li
            If isSelected Then lstAssets.Selected(idx) = True
        End If
    Next r
End Sub

' ── btnOK_Click ──────────────────────────────────────────────────────────────
' Bouwt een newline-gescheiden string van alle aangevinkte asset-namen en
' schrijft die terug naar de doelcel in het Processes-blad.
' AutoFit past de rijhoogte aan voor meerdere gekoppelde assets.
Private Sub btnOK_Click()
    Dim result As String: result = ""
    Dim i As Long
    For i = 0 To lstAssets.ListCount - 1
        If lstAssets.Selected(i) Then
            If result <> "" Then result = result & Chr(10)
            result = result & lstAssets.List(i)
        End If
    Next i
    Dim procWs As Worksheet
    Set procWs = ThisWorkbook.Sheets("Processes")
    procWs.Cells(GRC_Macros.g_TargetRow, GRC_Macros.g_PickerTargetCol) = result
    procWs.Rows(GRC_Macros.g_TargetRow).AutoFit
    Unload Me
    procWs.Activate
End Sub

' ── btnAnnuleer_Click ─────────────────────────────────────────────────────────
' Sluit het formulier zonder wijzigingen op te slaan.
Private Sub btnAnnuleer_Click()
    Unload Me
    ThisWorkbook.Sheets("Processes").Activate
End Sub
'''

RARM_SHEET_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module  : RARM (Worksheet)
' Doel    : Risk Assessment & Remediation Matrix — koppelt afhankelijke assets
'           (DA-kolommen) aan CyFun 2025 controls (rijen).
'           Rij 2 = DA-namen  |  Rij 3 = geselecteerde kwetsbaarheden per DA
'           Rij 4+ = CyFun 2025 control-rijen
' ══════════════════════════════════════════════════════════════════════════════

' ── Worksheet_Activate ───────────────────────────────────────────────────────
' Bij elke activering: synchroniseer de DA-kolommen met de Dependent Assets-sheet
' (SyncRARMKolommen) en herkleur de matrix (KleurAlleRARMKolommen).
' EnableEvents = False voorkomt dat de SyncRARMKolommen-aanroep zelf opnieuw
' een Activate-event triggert.
Private Sub Worksheet_Activate()
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    GRC_Macros.SyncRARMKolommen
    GRC_Macros.KleurAlleRARMKolommen
    Application.EnableEvents = True
    Application.ScreenUpdating = True
End Sub

' ── Worksheet_BeforeDoubleClick ───────────────────────────────────────────────
' Schakelt een ✔-vinkje aan/uit in de datacellen van de RARM-matrix (rij 4+,
' kolom 4+). Cellen in rijen 1-3 of kolommen 1-3 zijn beschermde headers.
' Extra veiligheid: als kolom A van de gerichte rij leeg is, is er geen
' control gedefinieerd → vinkje wordt genegeerd.
Private Sub Worksheet_BeforeDoubleClick(ByVal Target As Range, Cancel As Boolean)
    If Target.Cells.Count > 1 Then Exit Sub
    If Target.Row < 8 Then Exit Sub
    If Target.Column < 6 Then Exit Sub
    If CStr(Me.Cells(Target.Row, 1).Value) = "" Then Exit Sub
    Cancel = True
    If CStr(Target.Value) = ChrW(10004) Then
        Target.Value = ""
    Else
        Target.Value = ChrW(10004)
        Target.HorizontalAlignment = xlCenter
        Target.VerticalAlignment = xlCenter
    End If
End Sub

' ── Worksheet_SelectionChange ─────────────────────────────────────────────────
' Wanneer de gebruiker een cel in rij 3 (kwetsbaarheden-rij) selecteert vanaf
' kolom E (5+), wordt automatisch het VulnPicker-formulier geopend.
' g_RARMCol wordt ingesteld zodat het formulier weet voor welke DA-kolom
' de kwetsbaarheden worden geselecteerd.
' EnableEvents = False voorkomt een herhaalde trigger vanuit het formulier.
Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    If Target.Cells.Count > 1 Then Exit Sub
    If Target.Row <> 3 Then Exit Sub
    If Target.Column < 6 Then Exit Sub
    Application.EnableEvents = False
    GRC_Macros.TonenVulnPicker Target.Column
    Application.EnableEvents = True
End Sub
'''

VULNPICKER_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module  : VulnPicker (UserForm)
' Doel    : Popup-dialoogvenster om kwetsbaarheden + kanswaarden (1-4) te
'           koppelen aan een afhankelijke asset (DA) in de RARM-matrix.
'           Rij 3 van de RARM-sheet bevat per DA-kolom de geselecteerde
'           kwetsbaarheden in het formaat: "VulnNaam (prob), VulnNaam2 (prob)"
' Opbouw  : Het formulier bevat een scrollbaar frame (frmVulns) dat dynamisch
'           wordt gevuld met CheckBox + ComboBox per kwetsbaarheid.
'           De kwetsbaarheden-namen staan in rij 2 van het Kwetsbaarheden-blad
'           vanaf kolom C.
' ══════════════════════════════════════════════════════════════════════════════

' ── UserForm_Initialize ───────────────────────────────────────────────────────
' Vult het scrollbare frame dynamisch:
'   - Leest kwetsbaarheden-namen uit Kwetsbaarheden!rij 2, kolom C+
'   - Parseert de bestaande celwaarde in RARM!rij 3 naar een dictionary
'     {kwetsbaarheid → kans (1-4)} zodat eerder gekozen selecties bewaard blijven
'   - Voegt per kwetsbaarheid een CheckBox (naam) + ComboBox (kans 1-4) toe
'   - IIf-constructie voor caption: toon DA-naam als die beschikbaar is,
'     anders een generieke titel
Private Sub UserForm_Initialize()
    Dim wsK As Worksheet, wsR As Worksheet
    On Error Resume Next
    Set wsK = ThisWorkbook.Sheets("Kwetsbaarheden")
    Set wsR = ThisWorkbook.Sheets("RARM")
    On Error GoTo 0
    If wsK Is Nothing Then Exit Sub

    Dim daName As String
    daName = ""
    If Not wsR Is Nothing Then
        daName = CStr(wsR.Cells(2, GRC_Macros.g_RARMCol).Value)
    End If
    ' IIf is VBA's ternaire operator: als daName niet leeg → "voor: DA-naam", anders generieke titel
    Me.Caption = IIf(daName <> "", "Kwetsbaarheden voor: " & daName, "Kwetsbaarheden selecteren")
    lblTitel.Caption = Me.Caption

    ' ── Bestaande selectie inlezen naar Dictionary: vname → kans (1-4) ──────
    ' Opslagformaat in RARM rij 3: "VulnName (2), VulnName2 (3)"
    ' InStrRev zoekt het laatste " (" om kans-suffix te scheiden van naam,
    ' wat robuuster is dan een eenvoudige Split (namen kunnen spaties bevatten).
    Dim selDict As Object
    Set selDict = CreateObject("Scripting.Dictionary")
    If Not wsR Is Nothing Then
        Dim curVal As String
        curVal = CStr(wsR.Cells(3, GRC_Macros.g_RARMCol).Value)
        If curVal <> "" Then
            Dim parts() As String
            parts = Split(curVal, ", ")
            Dim pi As Integer
            For pi = 0 To UBound(parts)
                Dim entry As String
                entry = Trim(parts(pi))
                If entry <> "" Then
                    Dim parenOpen As Integer
                    parenOpen = InStrRev(entry, " (")
                    If parenOpen > 0 Then
                        Dim vKey As String, pNum As String
                        vKey = Left(entry, parenOpen - 1)
                        pNum = Mid(entry, parenOpen + 2, Len(entry) - parenOpen - 2)
                        selDict(vKey) = pNum
                    Else
                        selDict(entry) = "2"
                    End If
                End If
            Next pi
        End If
    End If

    ' ── Dynamisch CheckBox + ComboBox toevoegen per kwetsbaarheid ───────────
    ' Elke kwetsbaarheid krijgt een rij in het scrollbare frame:
    '   - CheckBox (links, 220 pt breed): naam van de kwetsbaarheid
    '   - ComboBox (rechts, 162 pt breed): kans 1–4 (standaard = 2)
    ' Rijen worden gestapeld met stap 22 punten; ScrollHeight past zich aan.
    Dim rowTop As Long
    rowTop = 4
    Dim col As Long
    col = 3
    Do While wsK.Cells(2, col).Value <> ""
        Dim vName As String
        vName = CStr(wsK.Cells(2, col).Value)

        Dim chk As Object
        Set chk = frmVulns.Controls.Add("Forms.CheckBox.1")
        chk.Name = "chk_" & col
        chk.Caption = vName
        chk.Tag = vName
        chk.Left = 4
        chk.Top = rowTop
        chk.Width = 220
        chk.Height = 18
        chk.Font.Size = 9

        Dim cmb As Object
        Set cmb = frmVulns.Controls.Add("Forms.ComboBox.1")
        cmb.Name = "cmb_" & col
        cmb.Tag = vName
        cmb.Left = 228
        cmb.Top = rowTop
        cmb.Width = 162
        cmb.Height = 18
        cmb.Font.Size = 9
        cmb.Style = 2   ' fmStyleDropDownList — voorkomt vrije tekstinvoer
        cmb.AddItem "1 - Not probable"
        cmb.AddItem "2 - Low probability"
        cmb.AddItem "3 - Medium probability"
        cmb.AddItem "4 - High probability"
        cmb.ListIndex = 1  ' standaard = index 1 = "2 - Low probability"

        Dim isSelected As Boolean
        isSelected = False
        If selDict.Exists(vName) Then
            isSelected = True
            Dim savedProb As String
            savedProb = CStr(selDict(vName))
            Dim savedIdx As Integer
            savedIdx = CInt(savedProb) - 1
            If savedIdx >= 0 And savedIdx <= 3 Then cmb.ListIndex = savedIdx
        End If
        chk.Value = isSelected

        rowTop = rowTop + 22
        col = col + 1
    Loop

    frmVulns.ScrollHeight = rowTop + 4
End Sub

' ── btnOK_Click ──────────────────────────────────────────────────────────────
' Itereert over alle dynamisch aangemaakte controls in frmVulns.
' Voor elke aangevinkte CheckBox wordt "VulnNaam (kans)" aan het resultaat
' toegevoegd. De kans wordt gehaald uit de gekoppelde ComboBox (ListIndex + 1).
' Resultaat wordt geschreven naar RARM!rij3!g_RARMCol en de kleuring
' wordt hersteld via KleurAlleRARMKolommen.
Private Sub btnOK_Click()
    Dim wsR As Worksheet, wsK As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    Set wsK = ThisWorkbook.Sheets("Kwetsbaarheden")
    On Error GoTo 0
    If wsR Is Nothing Then Unload Me: Exit Sub

    Dim result As String
    result = ""
    Dim col As Long
    col = 3
    Do While wsK.Cells(2, col).Value <> ""
        Dim chk As Object, cmb As Object
        Set chk = Nothing: Set cmb = Nothing
        On Error Resume Next
        Set chk = frmVulns.Controls("chk_" & col)
        Set cmb = frmVulns.Controls("cmb_" & col)
        On Error GoTo 0
        If Not chk Is Nothing Then
            If chk.Value = True Then
                Dim probNum As String
                probNum = "2"
                If Not cmb Is Nothing Then
                    If cmb.ListIndex >= 0 Then probNum = CStr(cmb.ListIndex + 1)
                End If
                If result <> "" Then result = result & ", "
                result = result & chk.Tag & " (" & probNum & ")"
            End If
        End If
        col = col + 1
    Loop
    wsR.Cells(3, GRC_Macros.g_RARMCol).Value = result
    Unload Me
    GRC_Macros.KleurAlleRARMKolommen
End Sub

' ── btnAnnuleer_Click ─────────────────────────────────────────────────────────
' Sluit het formulier zonder wijzigingen.
Private Sub btnAnnuleer_Click()
    Unload Me
End Sub
'''

PROC_SHEET_CODE = '''\
' ══════════════════════════════════════════════════════════════════════════════
' Module  : Processes (Worksheet)
' Doel    : Opent automatisch de AssetPicker-popup wanneer de gebruiker
'           klikt op kolom K (Informatieassets, index 11) of kolom L
'           (Afhankelijke assets, index 12) in het Processes-blad.
' ══════════════════════════════════════════════════════════════════════════════

' ── Worksheet_SelectionChange ─────────────────────────────────────────────────
' Detecteert een selectie in de data-kolommen K of L en delegeert naar
' GRC_Macros.TonenAssetPicker met de juiste bronsheet en doelkolom.
' EnableEvents = False voorkomt herhaalde triggers terwijl het formulier actief is.
Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    If Target.Cells.Count = 1 Then
        If Target.Row >= 6 And Target.Row <= 105 Then
            If Target.Column = 11 Or Target.Column = 12 Then
                Application.EnableEvents = False
                If Target.Column = 11 Then
                    ' Kolom K = Informatieassets: bronsheet is "Information Assets"
                    GRC_Macros.TonenAssetPicker Target.Row, "Information Assets", 11
                Else
                    ' Kolom L = Afhankelijke assets: bronsheet is "Dependent Assets"
                    GRC_Macros.TonenAssetPicker Target.Row, "Dependent Assets", 12
                End If
                Application.EnableEvents = True
            End If
        End If
    End If
End Sub
'''

VBA_CODE = r'''Attribute VB_Name = "GRC_Macros"
Option Explicit

' ══════════════════════════════════════════════════════════════════════════════
' Module  : GRC_Macros (Standaardmodule)
' Doel    : Centrale module voor alle import/export-macros en hulpfuncties
'           van de GRC Tool. Alle importoperaties maken gebruik van ADO
'           (ADODB.Connection / ADODB.Recordset) om gegevens uit een MS
'           Access-database te lezen.
' Afhankelijkheden:
'   - Microsoft Access Database Engine (ACE 12.0 of Jet 4.0)
'   - UserForms: AssetPicker, VulnPicker
'   - Bladen: Processes, Information Assets, Dependent Assets, RARM,
'             Kwetsbaarheden, CyFun Controls
' ══════════════════════════════════════════════════════════════════════════════

' ── Globale toestandsvariabelen ───────────────────────────────────────────────
' Deze variabelen worden ingesteld door de Worksheet_SelectionChange-handlers
' en doorgegeven aan de UserForms.
Public g_TargetRow As Long           ' Onthoudt welke rij in Processen geklikt werd
Public g_ProcesNaam As String        ' Naam van het geselecteerde proces (voor caption AssetPicker)
Public g_PickerSourceSheet As String ' Bronblad voor de picker (Informatieassets / Afhankelijke assets)
Public g_PickerTargetCol As Long     ' Doelkolom in Processen (11 = info assets, 12 = afhankelijke assets)
Public g_RARMCol As Long             ' Kolom in RARM-sheet waarop geklikt werd (vuln picker)

' GRC Tool v0.3 - Import & Export via MS Access (ADO)

' ── CIALbl ────────────────────────────────────────────────────────────────────
' Zet een CIA-niveau (1-5) om naar een leesbaar label (NL).
Function CIALbl(val As Variant) As String
    On Error GoTo fallback
    If IsEmpty(val) Or IsNull(val) Or CStr(val) = "" Then CIALbl = "": Exit Function
    Select Case CLng(val)
        Case 1: CIALbl = "Laag"
        Case 2: CIALbl = "Gemiddeld"
        Case 3: CIALbl = "Hoog"
        Case 4: CIALbl = "Zeer Hoog"
        Case 5: CIALbl = "Kritiek"
        Case Else: CIALbl = CStr(val)
    End Select
    Exit Function
fallback:
    CIALbl = CStr(val)
End Function

' ── MapCls ────────────────────────────────────────────────────────────────────
' Zet een classificatietekst of -cijfer (uit Access) om naar een geheel getal 1-5.
' Accepteert zowel numerieke strings ("3") als tekstwaarden in NL/FR/EN.
' Geeft 0 terug als de waarde niet herkend wordt (zodat de cel leeg blijft).
' De meertalige InStr-checks dekken de volledige CLS-woordenschat (zie Python).
Private Function MapCls(val As String) As Integer
    Dim s As String
    s = LCase(Trim(val))
    Dim n As Integer
    n = 0
    ' Probeer eerst numerieke conversie (snelste pad voor databases die cijfers opslaan)
    On Error Resume Next
    n = CInt(s)
    On Error GoTo 0
    If n >= 1 And n <= 5 Then MapCls = n: Exit Function
    ' Tekstuele match in NL, FR en EN (niet-casegevoelig door LCase hierboven)
    If InStr(s, "kritiek") > 0 Or InStr(s, "critique") > 0 Or InStr(s, "critical") > 0 Then
        MapCls = 5
    ElseIf InStr(s, "zeer hoog") > 0 Or InStr(s, "very high") > 0 Or InStr(s, "tr") > 0 And InStr(s, "s " & Chr(233)) > 0 Then
        MapCls = 4   ' "Très élevé" → Chr(233) = é
    ElseIf InStr(s, "hoog") > 0 Or InStr(s, "high") > 0 Or InStr(s, Chr(233) & "lev") > 0 Then
        MapCls = 3   ' "Élevé" → Chr(233) = é
    ElseIf InStr(s, "gemiddeld") > 0 Or InStr(s, "moyen") > 0 Or InStr(s, "medium") > 0 Then
        MapCls = 2
    ElseIf InStr(s, "laag") > 0 Or InStr(s, "faible") > 0 Or InStr(s, "low") > 0 Then
        MapCls = 1
    Else
        MapCls = 0   ' onbekende waarde → cel blijft leeg
    End If
End Function

' ── FieldVal ──────────────────────────────────────────────────────────────────
' Leest een veldwaarde uit een ADODB.Recordset op basis van één of meer
' kandidaat-namen (ParamArray).
' Twee-staps zoekstrategie (twee passen) voor robuuste kolomherkenning:
'   Pass 1 (exacte match): LCase(fld.Name) = LCase(kandidaat)
'          — snel en precies voor bekende kolomnamen
'   Pass 2 (gedeeltelijke match): InStr(LCase(fld.Name), LCase(kandidaat)) > 0
'          — als fallback voor databases met langere of afwijkende kolomnamen
' Geeft "" terug als geen match gevonden of als de veldwaarde NULL is.
'
' Voorbeeld:
'   FieldVal(rs, "naam", "name", "nom") vindt zowel "Naam", "ProcessNaam"
'   (pass 2) als "Nom du processus" (pass 2) in de recordset.
Private Function FieldVal(rs As Object, ParamArray names() As Variant) As String
    Dim n As Variant, fld As Object
    ' ── Pass 1: exacte naamvergelijking ──────────────────────────────────────
    For Each n In names
        For Each fld In rs.Fields
            If LCase(Trim(fld.Name)) = LCase(Trim(CStr(n))) Then
                If IsNull(fld.Value) Then FieldVal = "" Else FieldVal = CStr(fld.Value)
                Exit Function
            End If
        Next fld
    Next n
    ' ── Pass 2: deelstringvergelijking (fallback) ─────────────────────────────
    For Each n In names
        For Each fld In rs.Fields
            If InStr(LCase(fld.Name), LCase(Trim(CStr(n)))) > 0 Then
                If IsNull(fld.Value) Then FieldVal = "" Else FieldVal = CStr(fld.Value)
                Exit Function
            End If
        Next fld
    Next n
    FieldVal = ""
End Function

' ── VeldenTonen ─────────────────────────────────────────────────────────────
' Diagnostisch hulpprogramma: opent een Access-database, leest het schema
' (tabellen via adSchemaTables=20, queries via adSchemaViews=23) en
' schrijft alle tabel-/veldnamen en query-SQL naar een nieuw werkblad.
' Nuttig bij het debuggen van importproblemen met een onbekende databasestructuur.
' Het resultaat wordt niet via MsgBox getoond (te lang) maar weggeschreven naar
' een nieuw Workbook.
Sub VeldenTonen()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database voor diagnostiek"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Exit Sub

    Dim msg As String
    Dim schemars As Object, rs As Object, fld As Object

    ' ─ Tabellen ─
    msg = "=== TABELLEN ===" & vbCrLf & vbCrLf
    Set schemars = conn.OpenSchema(20)   ' adSchemaTables
    Do While Not schemars.EOF
        If schemars.Fields("TABLE_TYPE").Value = "TABLE" Then
            Dim tname As String
            tname = schemars.Fields("TABLE_NAME").Value
            msg = msg & "[" & tname & "]" & vbCrLf
            Set rs = CreateObject("ADODB.Recordset")
            On Error Resume Next
            rs.Open "SELECT TOP 1 * FROM [" & tname & "]", conn, 0, 1
            If Err.Number = 0 Then
                For Each fld In rs.Fields
                    msg = msg & "    " & fld.Name & vbCrLf
                Next fld
                rs.Close
            Else
                Err.Clear
            End If
            On Error GoTo 0
            msg = msg & vbCrLf
        End If
        schemars.MoveNext
    Loop
    schemars.Close

    ' ─ Queries ─
    msg = msg & "=== QUERIES ===" & vbCrLf & vbCrLf
    On Error Resume Next
    Set schemars = conn.OpenSchema(23)   ' adSchemaViews
    On Error GoTo 0
    If Not schemars Is Nothing Then
        Do While Not schemars.EOF
            Dim qname As String, qsql As String
            qname = schemars.Fields("TABLE_NAME").Value
            On Error Resume Next
            qsql = schemars.Fields("VIEW_DEFINITION").Value
            On Error GoTo 0
            msg = msg & "[" & qname & "]" & vbCrLf
            msg = msg & "    SQL: " & Left(qsql, 200) & vbCrLf & vbCrLf
            schemars.MoveNext
        Loop
        schemars.Close
    End If

    conn.Close

    ' Schrijf resultaat naar nieuw werkblad (te lang voor MsgBox)
    Dim owb As Workbook
    Set owb = Workbooks.Add
    Dim ows As Worksheet
    Set ows = owb.Sheets(1)
    ows.Name = "DB Diagnostiek"
    ows.Cells(1, 1) = msg
    ows.Columns(1).ColumnWidth = 120
    ows.Rows(1).WrapText = True
    ows.Rows(1).AutoFit
    MsgBox "Diagnostiek geschreven naar nieuw werkblad.", vbInformation, "GRC Diagnostiek"
End Sub

' ── OpenAccess ───────────────────────────────────────────────────────────────
' Opent een verbinding naar een Access-database via ADODB.
' Probeert eerst de modernere ACE 12.0-driver; valt terug op de oudere
' Jet 4.0-driver voor compatibiliteit met systemen zonder Access geïnstalleerd.
' Chr(239) = ï (geïnstalleerd) — nodig omdat VBA-strings geen UTF-8 zijn.
' Geeft Nothing terug en toont een foutmelding als beide drivers falen.
Private Function OpenAccess(dbPath As String) As Object
    Dim conn As Object
    Set conn = CreateObject("ADODB.Connection")
    On Error Resume Next
    conn.Open "Provider=Microsoft.ACE.OLEDB.12.0;Data Source=" & dbPath & ";"
    If Err.Number <> 0 Then
        Err.Clear
        ' Fallback: Jet 4.0 voor oudere Windows-installaties
        conn.Open "Provider=Microsoft.Jet.OLEDB.4.0;Data Source=" & dbPath & ";"
    End If
    On Error GoTo 0
    If conn.State = 0 Then
        MsgBox "Verbinding met Access-database mislukt." & vbCrLf & _
               "Zorg dat Microsoft Access Database Engine (ACE) is ge" & Chr(239) & "nstalleerd.", _
               vbExclamation, "GRC Import"
        Set OpenAccess = Nothing
    Else
        Set OpenAccess = conn
    End If
End Function

' ── ImportProcessen ──────────────────────────────────────────────────────────
' Importeert processen uit de Access-tabel "T - Processes in scope" naar het
' Processes-blad (rijen 6-105). Leegt eerst de doelkolommen B-L vooraleer
' nieuwe data te schrijven.
' Na het inlezen van de basisgegevens worden via VulGekoppeldeAssets en
' VulGekoppeldeAfhankelijkeAssets ook de gekoppelde assets opgehaald.
Sub ImportProcessen()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database (processen)"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Application.ScreenUpdating = False

    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub

    Dim rs As Object
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Processes in scope]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Processes in scope' niet gevonden in de database.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0

    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("Processes")
    ws.Range("B6:E105").ClearContents
    ws.Range("F6:F105").ClearContents  ' Integriteit code
    ws.Range("H6:H105").ClearContents  ' Beschikbaarheid code
    ws.Range("J6:J105").ClearContents  ' Opmerkingen
    ws.Range("K6:K105").ClearContents  ' Gekoppelde informatieassets
    ws.Range("L6:L105").ClearContents  ' Gekoppelde afhankelijke assets

    Dim destRow As Long
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "procesnaam", "process naam", "name", "process name", "nom", "nom du processus", "processus", "titel", "title")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr", "toelichting", "detail", "details")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "proceseigenaar", "process eigenaar", "process owner", "verantwoordelijke", "owner", "propri" & Chr(233) & "taire", "responsible")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "organisatie", "entiteit", "department", "service", "business unit", "entity")
        ws.Cells(destRow, 6) = MapCls(FieldVal(rs, "integriteit", "integrity", "int" & Chr(233) & "grit" & Chr(233), "integr"))
        ws.Cells(destRow, 8) = MapCls(FieldVal(rs, "beschikbaarheid", "availability", "disponibilit" & Chr(233), "beschikb", "availab"))
        destRow = destRow + 1
        rs.MoveNext
    Loop

    rs.Close

    ' Haal gekoppelde assets op via de koppelingstabellen
    VulGekoppeldeAssets conn, ws, 6, 105
    VulGekoppeldeAfhankelijkeAssets conn, ws, 6, 105
    ws.Rows("6:105").AutoFit

    conn.Close
    Application.ScreenUpdating = True
    MsgBox destRow - 6 & " processen ge" & Chr(239) & "mporteerd.", vbInformation, "GRC Import"
End Sub

' ── VulGekoppeldeLijst ───────────────────────────────────────────────────────
' Generieke helper: koppelt een asset-tabel aan processen via een many-to-many
' koppelingstabel (junction table). Schrijft de gekoppelde asset-namen als
' newline-gescheiden string naar een doelkolom in het Processes-blad.
'
' Werkwijze in 6 stappen:
'   1. Detecteer de FK-veldnamen in de koppelingstabel via kolomnaam-heuristieken
'   2. Detecteer de naam-kolom in "T - Processes in scope"
'   3. Detecteer de naam-kolom in de asset-tabel
'   4. Voer een JOIN-query uit om procesvelde (PNaam) en asset-naam (ANaam) op te halen
'   5. Bouw een in-memory map {procesnaam → asset-namen} op met array-buffers
'      (geen Scripting.Dictionary voor betere compatibiliteit)
'   6. Schrijf de asset-namen per process naar de doelkolom
'
' Parameters
' ----------
' conn          : ADODB.Connection — open verbinding met de Access-database
' ws            : Worksheet        — doelblad (Processes)
' startRow/endRow : Long           — bereik van datarijen in ws (bv. 6-105)
' junctionTable : String           — naam van de koppelingstabel (bv. "LT - Information Assets to Processes")
' assetTable    : String           — naam van de asset-tabel (bv. "T - Information Assets")
' assetKwd1/2   : String           — zoekwoorden voor het FK-veld in junctionTable
' targetCol     : Long             — doelkolom in ws (11 = K of 12 = L)
Private Sub VulGekoppeldeLijst(conn As Object, ws As Worksheet, startRow As Long, endRow As Long, _
                                junctionTable As String, assetTable As String, _
                                assetKwd1 As String, assetKwd2 As String, targetCol As Long)
    Dim tmpRs As Object, f As Object, fn As String

    ' Stap 1: FK-velden in koppelingstabel
    Set tmpRs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    tmpRs.Open "SELECT TOP 1 * FROM [" & junctionTable & "]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel '" & junctionTable & "' niet gevonden.", vbExclamation, "GRC Import"
        On Error GoTo 0: Exit Sub
    End If
    On Error GoTo 0
    Dim procFld As String, assetFld As String
    For Each f In tmpRs.Fields
        fn = LCase(f.Name)
        If procFld = "" And (InStr(fn, "process") > 0 Or InStr(fn, "proces") > 0) Then procFld = f.Name
        If assetFld = "" And InStr(fn, assetKwd1) > 0 Then assetFld = f.Name
        If assetFld = "" And assetKwd2 <> "" And InStr(fn, assetKwd2) > 0 Then assetFld = f.Name
    Next f
    tmpRs.Close
    If procFld = "" Or assetFld = "" Then
        MsgBox "FK-velden niet herkend in '" & junctionTable & "'." & vbCrLf & _
               "procFld='" & procFld & "'  assetFld='" & assetFld & "'", vbExclamation, "GRC Import"
        Exit Sub
    End If

    ' Stap 2: naamveld in T - Processes in scope
    Dim procNameFld As String
    Set tmpRs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    tmpRs.Open "SELECT TOP 1 * FROM [T - Processes in scope]", conn, 0, 1
    On Error GoTo 0
    For Each f In tmpRs.Fields
        fn = LCase(f.Name)
        If procNameFld = "" And fn <> "id" Then
            If InStr(fn, "name") > 0 Or InStr(fn, "naam") > 0 Or InStr(fn, "nom") > 0 Then
                procNameFld = f.Name
            End If
        End If
    Next f
    tmpRs.Close
    If procNameFld = "" Then procNameFld = "ProcessName"

    ' Stap 3: naamveld in asset-tabel
    Dim assetNameFld As String
    Set tmpRs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    tmpRs.Open "SELECT TOP 1 * FROM [" & assetTable & "]", conn, 0, 1
    On Error GoTo 0
    For Each f In tmpRs.Fields
        fn = LCase(f.Name)
        If assetNameFld = "" And fn <> "id" Then
            If InStr(fn, "name") > 0 Or InStr(fn, "naam") > 0 Or InStr(fn, "nom") > 0 Then
                assetNameFld = f.Name
            End If
        End If
    Next f
    tmpRs.Close
    If assetNameFld = "" Then assetNameFld = "Name"

    ' Stap 4: JOIN
    Dim rs As Object
    Set rs = CreateObject("ADODB.Recordset")
    Dim sql As String
    sql = "SELECT p.[" & procNameFld & "] AS PNaam, a.[" & assetNameFld & "] AS ANaam " & _
          "FROM ([T - Processes in scope] AS p " & _
          "INNER JOIN [" & junctionTable & "] AS lt ON lt.[" & procFld & "] = p.ID) " & _
          "INNER JOIN [" & assetTable & "] AS a ON a.ID = lt.[" & assetFld & "] " & _
          "ORDER BY p.[" & procNameFld & "], a.[" & assetNameFld & "]"
    On Error Resume Next
    rs.Open sql, conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "JOIN mislukt." & vbCrLf & sql, vbExclamation, "GRC Import"
        On Error GoTo 0: Exit Sub
    End If
    On Error GoTo 0

    ' Stap 5: bouw process→assets map
    Dim pkeys(500) As String, pvals(500) As String
    Dim pcount As Long: pcount = 0
    Dim pref As String, aref As String, fi As Long, pfound As Boolean
    Do While Not rs.EOF
        If IsNull(rs.Fields("PNaam").Value) Then pref = "" Else pref = CStr(rs.Fields("PNaam").Value)
        If IsNull(rs.Fields("ANaam").Value) Then aref = "" Else aref = CStr(rs.Fields("ANaam").Value)
        If pref <> "" And aref <> "" Then
            pfound = False
            For fi = 0 To pcount - 1
                If LCase(pkeys(fi)) = LCase(pref) Then
                    pvals(fi) = pvals(fi) & Chr(10) & aref: pfound = True: Exit For
                End If
            Next fi
            If Not pfound And pcount <= 500 Then
                pkeys(pcount) = pref: pvals(pcount) = aref: pcount = pcount + 1
            End If
        End If
        rs.MoveNext
    Loop
    rs.Close

    ' Stap 6: schrijf naar doelkolom
    Dim r As Long, pn As String
    For r = startRow To endRow
        pn = CStr(ws.Cells(r, 2).Value)
        If pn <> "" Then
            For fi = 0 To pcount - 1
                If LCase(pkeys(fi)) = LCase(pn) Then
                    ws.Cells(r, targetCol) = pvals(fi)
                    ws.Rows(r).AutoFit
                    Exit For
                End If
            Next fi
        End If
    Next r
End Sub

' ── VulGekoppeldeAssets ──────────────────────────────────────────────────────
' Specifieke wrapper: vult kolom K (11) van het Processes-blad met de namen
' van gekoppelde informatieassets via koppelingstabel "LT - Information Assets to Processes".
Private Sub VulGekoppeldeAssets(conn As Object, ws As Worksheet, startRow As Long, endRow As Long)
    VulGekoppeldeLijst conn, ws, startRow, endRow, _
        "LT - Information Assets to Processes", "T - Information Assets", "asset", "informat", 11
End Sub

' ── VulGekoppeldeAfhankelijkeAssets ──────────────────────────────────────────
' Specifieke wrapper: vult kolom L (12) van het Processes-blad met de namen
' van gekoppelde afhankelijke assets via "LT - Dependent assets to Processes".
Private Sub VulGekoppeldeAfhankelijkeAssets(conn As Object, ws As Worksheet, startRow As Long, endRow As Long)
    VulGekoppeldeLijst conn, ws, startRow, endRow, _
        "LT - Dependent assets to Processes", "T - Dependent assets", "depend", "", 12
End Sub

' ── ImportAlles ──────────────────────────────────────────────────────────────
' Voert alle importoperaties uit in één workflow vanuit één Access-database:
'   1. Informatieassets   → blad "Information Assets"
'   2. Afhankelijke assets → blad "Dependent Assets" (incl. C/I/A objectives)
'   3. Processen          → blad "Processes" (incl. gekoppelde assets)
'   4. RefreshRARMKolommen — herlaadt DA-kolomkoppen in RARM
'   5. CoreImportKwetsbaarheden — kwetsbaarheden + controls importeren
' Aan het einde worden de RARM-kleuren bijgewerkt via KleurAlleRARMKolommen.
Sub ImportAlles()
    Dim fd As FileDialog
    Dim conn As Object
    Dim rs As Object
    Dim ws As Worksheet
    Dim destRow As Long
    Dim nProc As Long
    Dim nAsset As Long
    Dim nDep As Long
    Dim nCls As Integer

    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Application.ScreenUpdating = False
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub

    ' --- 1. Informatieassets ---
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Information Assets]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Information Assets' niet gevonden.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0
    Set ws = ThisWorkbook.Sheets("Information Assets")
    ws.Range("B6:E105").ClearContents
    ws.Range("F6:F105").ClearContents
    ws.Range("H6:H105").ClearContents
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "assetnaam", "asset naam", "name", "asset name", "nom", "actif", "titel", "title")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr", "toelichting", "detail")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "asseteigenaar", "asset eigenaar", "owner", "propri" & Chr(233) & "taire", "verantwoordelijke")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "organisatie", "entiteit", "department", "service", "business unit")
        ws.Cells(destRow, 6) = MapCls(FieldVal(rs, "confidentialiteit", "confidentiality", "confidentialit" & Chr(233), "conf"))
        destRow = destRow + 1
        rs.MoveNext
    Loop
    rs.Close
    nAsset = destRow - 6

    ' --- 2. Afhankelijke assets ---
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Dependent assets]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Dependent assets' niet gevonden.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0
    Set ws = ThisWorkbook.Sheets("Dependent Assets")
    ws.Unprotect
    ws.Range("B6:E105").ClearContents
    ws.Range("F6:F105").ClearContents   ' Overarching
    ws.Range("G6:G105").ClearContents   ' Opmerkingen
    ws.Range("P6:P105").ClearContents   ' Conf objective (col 16)
    ws.Range("R6:R105").ClearContents   ' Int objective (col 18)
    ws.Range("T6:T105").ClearContents   ' Avail objective (col 20)
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "name", "nom", "afhankelijk", "dependent", "asset naam", "asset name")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "owner", "propri" & Chr(233) & "taire", "verantwoordelijke")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "department", "service", "entiteit")
        Dim oarchFld2 As Object
        On Error Resume Next: Set oarchFld2 = rs.Fields("Overarching"): On Error GoTo 0
        If Not oarchFld2 Is Nothing Then
            If Not IsNull(oarchFld2.Value) Then
                If CBool(oarchFld2.Value) Then ws.Cells(destRow, 6) = ChrW(10004)
            End If
            Set oarchFld2 = Nothing
        End If
        nCls = MapCls(FieldVal(rs, "c-objective", "confidentiality objective", "conf objective", "c obj"))
        If nCls > 0 Then ws.Cells(destRow, 16) = nCls
        nCls = MapCls(FieldVal(rs, "i-objective", "integrity objective", "int objective", "i obj"))
        If nCls > 0 Then ws.Cells(destRow, 18) = nCls
        nCls = MapCls(FieldVal(rs, "a-objective", "availability objective", "avail objective", "a obj"))
        If nCls > 0 Then ws.Cells(destRow, 20) = nCls
        destRow = destRow + 1
        rs.MoveNext
    Loop
    ws.Protect DrawingObjects:=True, Contents:=True, Scenarios:=True, _
        UserInterfaceOnly:=True, AllowFiltering:=True, AllowSorting:=True
    rs.Close
    nDep = destRow - 6

    ' --- 3. Processen ---
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Processes in scope]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Processes in scope' niet gevonden.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0
    Set ws = ThisWorkbook.Sheets("Processes")
    ws.Range("B6:E105").ClearContents
    ws.Range("F6:F105").ClearContents
    ws.Range("H6:H105").ClearContents
    ws.Range("J6:J105").ClearContents
    ws.Range("K6:K105").ClearContents
    ws.Range("L6:L105").ClearContents
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "procesnaam", "process naam", "name", "process name", "nom", "nom du processus", "processus", "titel", "title")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr", "toelichting", "detail", "details")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "proceseigenaar", "process eigenaar", "process owner", "verantwoordelijke", "owner", "propri" & Chr(233) & "taire", "responsible")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "organisatie", "entiteit", "department", "service", "business unit", "entity")
        ws.Cells(destRow, 6) = MapCls(FieldVal(rs, "integriteit", "integrity", "int" & Chr(233) & "grit" & Chr(233), "integr"))
        ws.Cells(destRow, 8) = MapCls(FieldVal(rs, "beschikbaarheid", "availability", "disponibilit" & Chr(233), "beschikb", "availab"))
        destRow = destRow + 1
        rs.MoveNext
    Loop
    rs.Close
    nProc = destRow - 6
    VulGekoppeldeAssets conn, ws, 6, 105
    VulGekoppeldeAfhankelijkeAssets conn, ws, 6, 105
    ws.Rows("6:105").AutoFit

    ' Herlaad RARM-kolomkoppen
    RefreshRARMKolommen

    ' Kwetsbaarheden en controls matrix mee importeren
    CoreImportKwetsbaarheden conn

    conn.Close
    KleurAlleRARMKolommen
    Application.ScreenUpdating = True
    MsgBox nProc & " processen, " & nAsset & " informatieassets en " & nDep & _
           " afhankelijke assets ge" & Chr(239) & "mporteerd.", vbInformation, "GRC Import"
End Sub

' ── ImportKoppelingen ────────────────────────────────────────────────────────
' Importeert uitsluitend de koppelingen tussen processen en assets
' (kolommen K en L van het Processes-blad), zonder de basisgegevens
' opnieuw te laden. Handig na een handmatige correctie van procesnamen.
Sub ImportKoppelingen()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database (koppelingen)"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Application.ScreenUpdating = False
    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub

    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("Processes")
    ws.Range("K6:K105").ClearContents
    ws.Range("L6:L105").ClearContents
    VulGekoppeldeAssets conn, ws, 6, 105
    VulGekoppeldeAfhankelijkeAssets conn, ws, 6, 105

    conn.Close
    Application.ScreenUpdating = True
    MsgBox "Koppelingen ge" & Chr(239) & "mporteerd (kolom K: informateassets, kolom L: afhankelijke assets).", vbInformation, "GRC Import"
End Sub

' ── ImportAfhankelijkeAssets ─────────────────────────────────────────────────
' Importeert afhankelijke assets uit "T - Dependent assets" naar het
' "Dependent Assets"-blad. Verwerkt ook het Overarching-veld (Boolean →
' ✔-vinkje in kolom F) en de C/I/A Security Objectives.
' Na import: herlaad RARM-kolomkoppen en bijwerk kleuring.
Sub ImportAfhankelijkeAssets()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database (afhankelijke assets)"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Application.ScreenUpdating = False

    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub

    Dim rs As Object
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Dependent assets]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Dependent assets' niet gevonden in de database.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0

    Dim ws As Worksheet
    Dim nCls As Integer
    Set ws = ThisWorkbook.Sheets("Dependent Assets")
    ws.Unprotect
    ws.Range("B6:E105").ClearContents
    ws.Range("F6:F105").ClearContents   ' Overarching
    ws.Range("G6:G105").ClearContents   ' Opmerkingen
    ws.Range("P6:P105").ClearContents   ' Conf objective (col 16)
    ws.Range("R6:R105").ClearContents   ' Int objective (col 18)
    ws.Range("T6:T105").ClearContents   ' Avail objective (col 20)

    Dim destRow As Long
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "name", "nom", "afhankelijk", "dependent", "asset naam", "asset name")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "owner", "propri" & Chr(233) & "taire", "verantwoordelijke")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "department", "service", "entiteit")
        Dim oarchFld As Object
        On Error Resume Next: Set oarchFld = rs.Fields("Overarching"): On Error GoTo 0
        If Not oarchFld Is Nothing Then
            If Not IsNull(oarchFld.Value) Then
                If CBool(oarchFld.Value) Then ws.Cells(destRow, 6) = ChrW(10004)
            End If
            Set oarchFld = Nothing
        End If
        nCls = MapCls(FieldVal(rs, "c-objective", "confidentiality objective", "conf objective", "c obj"))
        If nCls > 0 Then ws.Cells(destRow, 16) = nCls
        nCls = MapCls(FieldVal(rs, "i-objective", "integrity objective", "int objective", "i obj"))
        If nCls > 0 Then ws.Cells(destRow, 18) = nCls
        nCls = MapCls(FieldVal(rs, "a-objective", "availability objective", "avail objective", "a obj"))
        If nCls > 0 Then ws.Cells(destRow, 20) = nCls
        destRow = destRow + 1
        rs.MoveNext
    Loop

    ws.Protect DrawingObjects:=True, Contents:=True, Scenarios:=True, _
        UserInterfaceOnly:=True, AllowFiltering:=True, AllowSorting:=True

    rs.Close

    ' Koppel ook meteen aan Processen-sheet
    Dim procWs As Worksheet
    Set procWs = ThisWorkbook.Sheets("Processes")
    procWs.Range("L6:L105").ClearContents
    VulGekoppeldeAfhankelijkeAssets conn, procWs, 6, 105

    ' Herlaad RARM-kolomkoppen
    RefreshRARMKolommen

    conn.Close
    Application.ScreenUpdating = True
    KleurAlleRARMKolommen

    MsgBox destRow - 6 & " afhankelijke assets ge" & Chr(239) & "mporteerd.", vbInformation, "GRC Import"
End Sub

' ── ImportLinksKwetsbaarheden ────────────────────────────────────────────────
' Importeert de geselecteerde controls per DA (ImportGeselecteerdeControls)
' en de kwetsbaarheden per DA in RARM rij 3 (ImportRARMKwetsbaarheden).
' OPGELET: voer dit altijd UIT NA ImportKwetsbaarheden / ImportAlles,
' want de kwetsbaarheden-sheet moet al gevuld zijn voor correcte koppeling.
Sub ImportLinksKwetsbaarheden()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub
    Application.ScreenUpdating = False
    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub
    RefreshRARMKolommen
    ImportRARMKwetsbaarheden conn
    ImportGeselecteerdeControls conn
    conn.Close
    Application.ScreenUpdating = True
    KleurAlleRARMKolommen
    MsgBox "Links kwetsbaarheden / controls per afhankelijke asset ge" & Chr(239) & "mporteerd.", vbInformation, "GRC Import"
End Sub

' ── ImportGeselecteerdeControls ──────────────────────────────────────────────
' Importeert de geselecteerde controls per afhankelijke asset (DA) vanuit de
' koppelingstabel "LT - Selected controls to DA" en schrijft ✔-vinkjes in de
' overeenkomstige RARM-datacellen.
'
' Vertaalketen (stap voor stap):
'   ctrlRef  (LT.ControlReference = Long)
'     → refNrMap(ctrlRef) = CyFun 2023 ID-string  (uit T - CyFunEssentiel.Requirement)
'       bv. "IMPORTANT_RS.CO-3.2: The organization..."
'     → NormId23() → id23  = genormaliseerde 2023-ID (bv. "rs.co-3.2")
'     → rev23to25(id23) = id25  = 2025-ID (bv. "rs.co-3.2" als ongewijzigd,
'                                  of nieuw 2025-ID als hernoemd)
'     → ctrlRowMap(id25) = rijnummer in RARM
'   ltDaId (LT.DAID = Long)
'     → daIdColMap(ltDaId) = kolomnummer in RARM (via DA-naam in rij 2)
'
' Koppelingssleutel DA-naam ↔ RARM: T-Dependent assets.DAName (NIET .ID)
' omdat RARM-kolommen op naam zijn aangemaakt (RefreshRARMKolommen).
'
' Afwijkingen (2023-controls zonder 2025-equivalent) worden weggeschreven naar
' het blad "Controls 2023 - DA" voor rapportage aan de CISO.
Sub ImportGeselecteerdeControls(conn As Object)
    Const RARM_ROW_DA   As Long = 2
    Const RARM_ROW_DATA As Long = 8
    Const RARM_COL_ID   As Long = 1
    Const RARM_COL_DA   As Long = 6
    Const COL_CF25      As Integer = 6
    Const COL_CF23      As Integer = 13
    Const ROW_CF_DATA   As Integer = 4

    Dim wsR As Worksheet, wsCC As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    Set wsCC = ThisWorkbook.Sheets("CyFun Controls")
    On Error GoTo 0
    If wsR Is Nothing Then Exit Sub

    Dim rarmLast As Long
    rarmLast = wsR.Cells(wsR.Rows.Count, RARM_COL_ID).End(xlUp).Row
    Dim lastCol As Long
    lastCol = wsR.Cells(RARM_ROW_DA, wsR.Columns.Count).End(xlToLeft).Column

    ' ── 1. ctrlRowMap: LCase(2025-ID) → rijnummer in RARM ───────────────────────
    Dim ctrlRowMap As Object
    Set ctrlRowMap = CreateObject("Scripting.Dictionary")
    Dim rr As Long
    For rr = RARM_ROW_DATA To rarmLast
        Dim ck As String
        ck = LCase(Trim(CStr(wsR.Cells(rr, RARM_COL_ID).Value)))
        If ck <> "" And Not ctrlRowMap.Exists(ck) Then ctrlRowMap.Add ck, rr
    Next rr

    ' ── 2. refNrMap: RefNr (Long) → CyFun 2023 ID-string ───────────────────────
    Dim refNrMap As Object
    Set refNrMap = CreateObject("Scripting.Dictionary")
    Dim rsCtrl As Object
    Set rsCtrl = CreateObject("ADODB.Recordset")
    rsCtrl.Open "SELECT RefNr, Requirement FROM [T - CyFunEssentiel]", conn
    Do While Not rsCtrl.EOF
        Dim refNr As Long
        refNr = CLng(rsCtrl.Fields("RefNr").Value)
        If Not refNrMap.Exists(refNr) Then
            refNrMap.Add refNr, CStr(rsCtrl.Fields("Requirement").Value)
        End If
        rsCtrl.MoveNext
    Loop
    rsCtrl.Close

    ' ── 3. rev23to25: LCase(2023-ID) → LCase(2025-ID) via CyFun Controls sheet ──
    '       enkel23Set: 2023-IDs zonder 2025-equivalent
    Dim rev23to25 As Object, enkel23Set As Object
    Set rev23to25  = CreateObject("Scripting.Dictionary")
    Set enkel23Set = CreateObject("Scripting.Dictionary")
    If Not wsCC Is Nothing Then
        Dim ccLast As Long
        ccLast = Application.Max( _
            wsCC.Cells(wsCC.Rows.Count, COL_CF25).End(xlUp).Row, _
            wsCC.Cells(wsCC.Rows.Count, COL_CF23).End(xlUp).Row)
        Dim rCC As Long
        For rCC = ROW_CF_DATA To ccLast
            Dim r25t As String, r23t As String
            r25t = Trim(CStr(wsCC.Cells(rCC, COL_CF25).Value))
            r23t = Trim(CStr(wsCC.Cells(rCC, COL_CF23).Value))
            If r25t <> "" And r23t <> "" Then
                Dim p25 As Integer, p23 As Integer
                p25 = InStr(r25t, ":"): p23 = InStr(r23t, ":")
                Dim i25 As String, i23 As String
                i25 = LCase(Trim(Replace(IIf(p25 > 0, Left(r25t, p25 - 1), r25t), Chr(9), "")))
                i23 = NormId23(r23t)
                If i23 <> "" And i25 <> "" And Not rev23to25.Exists(i23) Then
                    rev23to25.Add i23, i25
                End If
            ElseIf r25t = "" And r23t <> "" Then
                Dim e23 As String
                e23 = NormId23(r23t)
                If e23 <> "" And Not enkel23Set.Exists(e23) Then
                    enkel23Set.Add e23, True
                End If
            End If
        Next rCC
    End If

    ' ── 4. daIdColMap: DAID (Long) → RARM-kolom ─────────────────────────────────
    ' Koppel via naam: T - Dependent assets.DAName ↔ RARM rij 2 (LCase)
    Dim daNameIdMap As Object, daIdNameMap As Object
    Set daNameIdMap = CreateObject("Scripting.Dictionary")
    Set daIdNameMap = CreateObject("Scripting.Dictionary")
    Dim rsDA As Object
    Set rsDA = CreateObject("ADODB.Recordset")
    rsDA.Open "SELECT ID, DAName FROM [T - Dependent assets]", conn
    Do While Not rsDA.EOF
        Dim daId As Long
        daId = CLng(rsDA.Fields("ID").Value)
        Dim daNm As String
        daNm = LCase(Trim(CStr(rsDA.Fields("DAName").Value)))
        If daNm <> "" And Not daNameIdMap.Exists(daNm) Then
            daNameIdMap.Add daNm, daId
            daIdNameMap.Add daId, CStr(rsDA.Fields("DAName").Value)
        End If
        rsDA.MoveNext
    Loop
    rsDA.Close

    Dim daIdColMap As Object
    Set daIdColMap = CreateObject("Scripting.Dictionary")
    Dim c As Long
    For c = RARM_COL_DA To lastCol
        Dim rarmDaNm As String
        rarmDaNm = LCase(Trim(CStr(wsR.Cells(RARM_ROW_DA, c).Value)))
        If rarmDaNm <> "" And daNameIdMap.Exists(rarmDaNm) Then
            Dim daIdForCol As Long
            daIdForCol = CLng(daNameIdMap(rarmDaNm))
            If Not daIdColMap.Exists(daIdForCol) Then daIdColMap.Add daIdForCol, c
        End If
    Next c

    ' ── 5. Wis bestaande vinkjes in DA-kolommen ──────────────────────────────────
    If lastCol >= RARM_COL_DA And rarmLast >= RARM_ROW_DATA Then
        wsR.Range(wsR.Cells(RARM_ROW_DATA, RARM_COL_DA), _
                  wsR.Cells(rarmLast, lastCol)).ClearContents
    End If

    ' ── 6. Lees LT - Selected controls to DA en schrijf vinkjes in RARM ─────────
    Dim nMatched As Long
    nMatched = 0
    Dim afwList As Object          ' sleutel="DAID|id23" → Array(daNaam, id23, reden)
    Set afwList = CreateObject("Scripting.Dictionary")
    Dim rsLT As Object
    Set rsLT = CreateObject("ADODB.Recordset")
    rsLT.Open "SELECT DAID, ControlReference FROM [LT - Selected controls to DA]", conn
    Do While Not rsLT.EOF
        Dim ltDaId As Long, ctrlRef As Long
        ltDaId  = CLng(rsLT.Fields("DAID").Value)
        ctrlRef = CLng(rsLT.Fields("ControlReference").Value)
        If daIdColMap.Exists(ltDaId) And refNrMap.Exists(ctrlRef) Then
            Dim id23v As String
            id23v = NormId23(CStr(refNrMap.Item(ctrlRef)))
            Dim id25v As String
            If rev23to25.Exists(id23v) Then
                id25v = CStr(rev23to25.Item(id23v))
            Else
                id25v = id23v
            End If
            If ctrlRowMap.Exists(id25v) Then
                Dim ctrlRow As Long
                ctrlRow = CLng(ctrlRowMap.Item(id25v))
                Dim targetCol As Long
                targetCol = CLng(daIdColMap(ltDaId))
                With wsR.Cells(ctrlRow, targetCol)
                    .Value = ChrW(10004)
                    .HorizontalAlignment = xlCenter
                    .VerticalAlignment = xlCenter
                End With
                nMatched = nMatched + 1
            Else
                ' Control niet gevonden in 2025 — registreer afwijking
                Dim afwKey As String
                afwKey = ltDaId & "|" & id23v
                If Not afwList.Exists(afwKey) Then
                    Dim daNaamAfw As String
                    If daIdNameMap.Exists(ltDaId) Then
                        daNaamAfw = CStr(daIdNameMap.Item(ltDaId))
                    Else
                        daNaamAfw = CStr(ltDaId)
                    End If
                    Dim redenAfw As String
                    If enkel23Set.Exists(id23v) Then
                        redenAfw = "Enkel in CyFun 2023"
                    Else
                        redenAfw = "Niet gevonden in 2025"
                    End If
                    afwList.Add afwKey, Array(daNaamAfw, id23v, redenAfw)
                End If
            End If
        End If
        rsLT.MoveNext
    Loop
    rsLT.Close

    ' ── 7. Afwijkingen wegschrijven naar sheet "Controls 2023 - DA" ─────────────
    Dim wsAfw As Worksheet
    On Error Resume Next
    Set wsAfw = ThisWorkbook.Sheets("Controls 2023 - DA")
    On Error GoTo 0
    If wsAfw Is Nothing Then
        Set wsAfw = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        wsAfw.Name = "Controls 2023 - DA"
    End If

    ' Koptekst (enkel bij eerste aanmaak of als rij 1 leeg is)
    If Trim(CStr(wsAfw.Cells(1, 1).Value)) = "" Then
        With wsAfw.Cells(1, 1)
            .Value = "Controls uit CyFun 2023 zonder equivalent in CyFun 2025"
            .Font.Bold = True: .Font.Size = 11
        End With
        With wsAfw.Cells(2, 1): .Value = "Dependent Asset": .Font.Bold = True: End With
        With wsAfw.Cells(2, 2): .Value = "2023 Control-ID":  .Font.Bold = True: End With
        With wsAfw.Cells(2, 3): .Value = "Reden":            .Font.Bold = True: End With
    End If

    ' Wis vorige data (rij 3+)
    Dim afwLast As Long
    afwLast = wsAfw.Cells(wsAfw.Rows.Count, 1).End(xlUp).Row
    If afwLast >= 3 Then wsAfw.Rows("3:" & afwLast).ClearContents

    ' Schrijf afwijkingen gesorteerd op DA-naam
    Dim afwRow As Long: afwRow = 3
    Dim afwK As Variant
    For Each afwK In afwList.Keys
        Dim afwInfo As Variant
        afwInfo = afwList.Item(afwK)
        wsAfw.Cells(afwRow, 1).Value = afwInfo(0)
        wsAfw.Cells(afwRow, 2).Value = afwInfo(1)
        wsAfw.Cells(afwRow, 3).Value = afwInfo(2)
        afwRow = afwRow + 1
    Next afwK

    ' Autofit kolommen
    wsAfw.Columns("A:C").AutoFit

    MsgBox "Import voltooid: " & nMatched & " controls gekoppeld." & vbCrLf & _
           afwList.Count & " 2023-controls zonder 2025-equivalent opgeslagen in 'Controls 2023 - DA'.", _
           vbInformation, "GRC Import"
End Sub

' ── TonenVulnPicker ──────────────────────────────────────────────────────────
' Opent het VulnPicker-formulier voor de opgegeven DA-kolom in de RARM-sheet.
' Slaat de kolomindex op in g_RARMCol zodat het formulier de juiste cel kan lezen/schrijven.
Sub TonenVulnPicker(targetCol As Long)
    g_RARMCol = targetCol
    VulnPicker.Show
End Sub

' ── GetMarkedCtrlIDs ──────────────────────────────────────────────────────────
' Verzamelt alle Control IDs (kolom A) die een ✔ hebben voor minstens één
' van de kwetsbaarheden in selVulns.
' selVulns: kommagescheiden string "VulnNaam (kans), ..." (uit RARM rij 3).
' De kans-suffix wordt gestript via InStrRev voor de match.
' Retourneert een Scripting.Dictionary {lowercase ctrl_id → True} voor snelle
' opzoekingen door KleurRARMKolom.
Private Function GetMarkedCtrlIDs(wsK As Worksheet, selVulns As String) As Object
    Const KWETS_ROW_HDR    As Long = 2
    Const KWETS_ROW_DATA   As Long = 6
    Const KWETS_COL_ID     As Long = 1
    Const KWETS_COL_VSTART As Long = 3
    Dim result As Object
    Set result = CreateObject("Scripting.Dictionary")
    If wsK Is Nothing Or selVulns = "" Then Set GetMarkedCtrlIDs = result: Exit Function
    Dim kwetsLast As Long
    kwetsLast = wsK.Cells(wsK.Rows.Count, KWETS_COL_ID).End(xlUp).Row
    Dim selParts() As String
    selParts = Split(selVulns, ", ")
    Dim p As Integer, kc As Long, r As Long
    Dim vName As String, vCol As Long, hdrVal As String, cv As String, cid As String
    For p = 0 To UBound(selParts)
        vName = Trim(selParts(p))
        ' Strip probability suffix (format: "VulnName (N)")
        Dim parenIdx As Integer
        parenIdx = InStrRev(vName, " (")
        If parenIdx > 0 Then vName = Left(vName, parenIdx - 1)
        If vName <> "" Then
            vCol = 0
            For kc = KWETS_COL_VSTART To KWETS_COL_VSTART + 300
                hdrVal = CStr(wsK.Cells(KWETS_ROW_HDR, kc).Value)
                If hdrVal = "" Then Exit For
                If hdrVal = vName Then vCol = kc: Exit For
            Next kc
            If vCol > 0 Then
                For r = KWETS_ROW_DATA To kwetsLast
                    cv = ""
                    If Not IsError(wsK.Cells(r, vCol).Value) Then cv = CStr(wsK.Cells(r, vCol).Value)
                    If cv = ChrW(10004) Then
                        cid = LCase(Trim(CStr(wsK.Cells(r, KWETS_COL_ID).Value)))
                        If cid <> "" And Not result.Exists(cid) Then result.Add cid, True
                    End If
                Next r
            End If
        End If
    Next p
    Set GetMarkedCtrlIDs = result
End Function

' ── ImportRARMKwetsbaarheden ─────────────────────────────────────────────────
' Vult rij 3 van de RARM-sheet met de geselecteerde kwetsbaarheden (+ kans)
' per afhankelijke asset.
' Gebruikt een directe SQL-JOIN op de onderliggende tabellen in plaats van
' de Access-query "QLT2 - Vulnerabilities for DA - assigned" (die formulier-
' parameters bevat die niet via ADO kunnen worden meegegeven).
'
' Tabellen gebruikt:
'   T - Dependent assets           : DAName
'   LT - Vulnerabilities to Dependent Assets : DAID, VulnerabilityID, Probability
'   T - Vulnerabilities            : ID, Vulnerability (naam)
'
' Resultaatformaat in RARM rij 3: "VulnNaam (kans), VulnNaam2 (kans), ..."
Sub ImportRARMKwetsbaarheden(conn As Object)
    Const RARM_ROW_DA   As Long = 2
    Const RARM_ROW_VULN As Long = 3
    Const RARM_COL_DA   As Long = 6

    Dim wsR As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    On Error GoTo 0
    If wsR Is Nothing Then Exit Sub

    ' ── 1. Directe SQL op de onderliggende tabellen (geen Access-query nodig) ─────
    ' Tabel LT - Vulnerabilities to Dependent Assets: DAID, VulnerabilityID, Probability
    Dim qSQL As String
    qSQL = "SELECT [T - Dependent assets].DAName, [T - Vulnerabilities].Vulnerability, " & _
           "[LT - Vulnerabilities to Dependent Assets].Probability " & _
           "FROM ([T - Dependent assets] " & _
           "INNER JOIN [LT - Vulnerabilities to Dependent Assets] " & _
           "ON [T - Dependent assets].ID = [LT - Vulnerabilities to Dependent Assets].DAID) " & _
           "INNER JOIN [T - Vulnerabilities] " & _
           "ON [T - Vulnerabilities].ID = [LT - Vulnerabilities to Dependent Assets].VulnerabilityID"

    ' ── 2. Voer SQL uit en groepeer per DA-naam ──────────────────────────────────
    ' qaMap: LCase(DA-naam) → vbTab-gescheiden "vulnNaam~prob" items
    Dim qaMap As Object
    Set qaMap = CreateObject("Scripting.Dictionary")
    Dim rs As Object
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open qSQL, conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Fout bij uitvoeren query: " & Err.Description, vbExclamation, "GRC Import"
        Exit Sub
    End If
    On Error GoTo 0

    Do While Not rs.EOF
        Dim daNm As String
        daNm = FieldVal(rs, "naam", "name", "da naam", "da name", "dependent asset", _
                        "asset naam", "asset name", "asset", "da")
        Dim vNm As String
        vNm = FieldVal(rs, "vulnerability", "kwetsbaarheid", "vuln naam", "vuln name", "vuln")
        Dim pStr As String
        pStr = "2"
        On Error Resume Next
        If Not IsNull(rs.Fields("Probability").Value) Then pStr = Trim(CStr(rs.Fields("Probability").Value))
        On Error GoTo 0
        If daNm <> "" And vNm <> "" Then
            Dim qKey As String
            qKey = LCase(Trim(daNm))
            Dim qEntry As String
            qEntry = vNm & "~" & pStr
            If qaMap.Exists(qKey) Then
                qaMap(qKey) = qaMap(qKey) & vbTab & qEntry
            Else
                qaMap(qKey) = qEntry
            End If
        End If
        rs.MoveNext
    Loop
    rs.Close

    ' ── 3. Schrijf naar RARM rij 3 op basis van DA-naam in rij 2 (bovenste lijn) ─
    Dim col As Long
    col = RARM_COL_DA
    Do While Trim(CStr(wsR.Cells(RARM_ROW_DA, col).Value)) <> ""
        Dim rarmKey As String
        rarmKey = LCase(Trim(CStr(wsR.Cells(RARM_ROW_DA, col).Value)))
        Dim result As String
        result = ""
        If qaMap.Exists(rarmKey) Then
            Dim pairs() As String
            pairs = Split(CStr(qaMap(rarmKey)), vbTab)
            Dim pi As Integer
            For pi = 0 To UBound(pairs)
                Dim pp() As String
                pp = Split(pairs(pi), "~")
                If UBound(pp) >= 0 And pp(0) <> "" Then
                    If result <> "" Then result = result & ", "
                    result = result & pp(0)
                    If UBound(pp) >= 1 And pp(1) <> "" Then result = result & " (" & pp(1) & ")"
                End If
            Next pi
        End If
        wsR.Cells(RARM_ROW_VULN, col).Value = result
        col = col + 1
    Loop
End Sub

' ── RefreshRARMKolommen ──────────────────────────────────────────────────────
' Herlaadt de DA-kolomkoppen (rij 2) in de RARM-sheet op basis van het
' Dependent Assets-blad (kolom B = naam, kolom F = overarching-vinkje).
' Wist eerst het volledige DA-bereik (koppen + data + opmaak) om restanten
' van verwijderde of herordende DA's te vermijden.
' Overarching-DA's krijgen een oranje achtergrond (RGB 180,83,9);
' gewone DA's krijgen gele achtergrond (RGB 254,249,195).
' Rij 1 (titel-merge) wordt na afloop uitgebreid tot het nieuwe kolombereik.
Sub RefreshRARMKolommen()
    Const RARM_ROW_DA   As Long = 2
    Const RARM_ROW_VULN As Long = 3
    Const RARM_ROW_DATA As Long = 8
    Const RARM_COL_ID   As Long = 1
    Const RARM_COL_DA   As Long = 6
    Const DA_COL_NAAM   As Long = 2
    Const DA_COL_OARCH  As Long = 6
    Dim wsR As Worksheet, wsD As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    Set wsD = ThisWorkbook.Sheets("Dependent Assets")
    On Error GoTo 0
    If wsR Is Nothing Or wsD Is Nothing Then Exit Sub

    ' Bepaal het huidige bereik van DA-kolommen en de laatste datarij
    Dim lastCol As Long, rarmLastData As Long
    lastCol = wsR.Cells(RARM_ROW_DA, wsR.Columns.Count).End(xlToLeft).Column
    rarmLastData = wsR.Cells(wsR.Rows.Count, RARM_COL_ID).End(xlUp).Row

    ' Reset het volledige DA-zone: koppen (rij 2-3) + datacellen — zodat verwijderde
    ' kolommen geen opmaakrestanten achterlaten
    If lastCol >= RARM_COL_DA Then
        Dim hdrRng As Range, dataRng As Range, ciaRng As Range
        Set hdrRng  = wsR.Range(wsR.Cells(RARM_ROW_DA, RARM_COL_DA), _
                                wsR.Cells(RARM_ROW_VULN, lastCol + 5))
        Set ciaRng  = wsR.Range(wsR.Cells(4, RARM_COL_DA), wsR.Cells(7, lastCol + 5))
        Set dataRng = wsR.Range(wsR.Cells(RARM_ROW_DATA, RARM_COL_DA), _
                                wsR.Cells(rarmLastData, lastCol + 5))
        hdrRng.ClearContents
        hdrRng.Interior.ColorIndex  = xlNone
        hdrRng.Borders.LineStyle    = xlNone
        ciaRng.ClearContents
        ciaRng.Interior.ColorIndex  = xlNone
        ciaRng.Borders.LineStyle    = xlNone
        dataRng.Interior.ColorIndex = xlNone
        dataRng.Borders.LineStyle   = xlNone
    End If

    Dim col As Long
    col = RARM_COL_DA
    Dim r As Long, daNaam As String, isOarch As Boolean
    Dim rC8 As String, rCB As String
    For r = 6 To 105
        daNaam = Trim(CStr(wsD.Cells(r, DA_COL_NAAM).Value))
        If daNaam <> "" Then
            isOarch = (Trim(CStr(wsD.Cells(r, DA_COL_OARCH).Value)) <> "")
            With wsR.Cells(RARM_ROW_DA, col)
                .Value = daNaam
                .Font.Bold = True: .Font.Size = 10
                .Font.Color = RGB(15, 43, 70)
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .WrapText = True
                If isOarch Then
                    .Interior.Color = RGB(180, 83, 9)
                Else
                    .Interior.Color = RGB(254, 249, 195)
                End If
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With
            With wsR.Cells(RARM_ROW_VULN, col)
                .Font.Italic = True: .Font.Size = 9
                .Font.Color = RGB(100, 116, 139)
                .Interior.Color = RGB(239, 246, 255)
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .WrapText = True
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With
            ' CIA-objectieven (rijen 4-6) vanuit DA-sheet; aangevinkt-teller (rij 7)
            rC8 = wsR.Cells(RARM_ROW_DATA, col).Address(False, False)
            rCB = wsR.Cells(1048576, col).Address(False, False)
            wsR.Cells(4, col).Value = CIALbl(wsD.Cells(r, 16).Value)
            wsR.Cells(5, col).Value = CIALbl(wsD.Cells(r, 18).Value)
            wsR.Cells(6, col).Value = CIALbl(wsD.Cells(r, 20).Value)
            wsR.Cells(7, col).Formula = "=COUNTA(" & rC8 & ":" & rCB & ")"
            With wsR.Range(wsR.Cells(4, col), wsR.Cells(7, col))
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .Interior.Color = RGB(239, 246, 255)
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With
            wsR.Columns(col).ColumnWidth = 22
            col = col + 1
        End If
    Next r
    ' Herlaad titel-merge (rij 1) zodat die altijd de volledige breedte dekt
    Dim titleLast As Long
    titleLast = col - 1
    If titleLast < RARM_COL_DA + 9 Then titleLast = RARM_COL_DA + 9  ' min 10 DA-kolommen
    On Error Resume Next
    wsR.Cells(1, 1).MergeArea.UnMerge
    On Error GoTo 0
    wsR.Range(wsR.Cells(1, 1), wsR.Cells(1, titleLast)).Merge
End Sub

' ── KleurAlleRARMKolommen ────────────────────────────────────────────────────
' Itereert over alle DA-kolommen (rij 2, kolom 5+) en roept per kolom
' KleurRARMKolom aan. Stopt bij de eerste lege kolomkop (einde DA-zone).
Sub KleurAlleRARMKolommen()
    Const RARM_ROW_DA As Long = 2
    Const RARM_COL_DA As Long = 6
    Dim wsR As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    On Error GoTo 0
    If wsR Is Nothing Then Exit Sub
    Dim col As Long
    col = RARM_COL_DA
    Do While Trim(CStr(wsR.Cells(RARM_ROW_DA, col).Value)) <> ""
        KleurRARMKolom col
        col = col + 1
    Loop
End Sub

' ── SyncRARMKolommen ─────────────────────────────────────────────────────────
' Synchroniseert de DA-kolommen van de RARM-sheet met de huidige staat van
' het Dependent Assets-blad. Bewaart rij-3-waarden (kwetsbaarheden) en
' ✔-vinkjes voor DA's die nog bestaan; verwijdert kolommen voor weggevallen DA's.
' Herbouwt de kolommen in de volgorde van het Dependent Assets-blad.
' Roep na afloop KleurAlleRARMKolommen aan voor correcte kleuring.
'
' Werkwijze:
'   1. Snapshot: sla rij-3-waarden en vinkjes op per DA-naam (Dictionary)
'   2. Wis de volledige DA-zone (koppen + data + opmaak)
'   3. Herbouw kolommen in DA-sheet-volgorde; herstel opgeslagen data
'   4. Vernieuw titel-merge (rij 1) tot het nieuwe kolombereik
Sub SyncRARMKolommen()
    Const RARM_ROW_DA   As Long = 2
    Const RARM_ROW_VULN As Long = 3
    Const RARM_ROW_DATA As Long = 8
    Const RARM_COL_ID   As Long = 1
    Const RARM_COL_DA   As Long = 6
    Const DA_COL_NAAM   As Long = 2
    Const DA_COL_OARCH  As Long = 6

    Dim wsR As Worksheet, wsD As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    Set wsD = ThisWorkbook.Sheets("Dependent Assets")
    On Error GoTo 0
    If wsR Is Nothing Or wsD Is Nothing Then Exit Sub

    Dim rarmLast As Long
    rarmLast = wsR.Cells(wsR.Rows.Count, RARM_COL_ID).End(xlUp).Row
    Dim lastCol As Long
    lastCol = wsR.Cells(RARM_ROW_DA, wsR.Columns.Count).End(xlToLeft).Column

    ' ── 1. Snapshot: bewaar rij-3-waarde en vinkjes per DA-naam ─────────────────
    Dim oldVuln As Object
    Set oldVuln = CreateObject("Scripting.Dictionary")
    Dim oldData As Object
    Set oldData = CreateObject("Scripting.Dictionary")

    Dim c As Long, dKey As String, dNm As String
    For c = RARM_COL_DA To lastCol
        dNm = Trim(CStr(wsR.Cells(RARM_ROW_DA, c).Value))
        If dNm <> "" Then
            dKey = LCase(dNm)
            If Not oldVuln.Exists(dKey) Then
                oldVuln.Add dKey, CStr(wsR.Cells(RARM_ROW_VULN, c).Value)
                Dim rowDict As Object
                Set rowDict = CreateObject("Scripting.Dictionary")
                Dim r As Long
                For r = RARM_ROW_DATA To rarmLast
                    Dim cv As String
                    cv = CStr(wsR.Cells(r, c).Value)
                    If cv <> "" Then rowDict.Add r, cv
                Next r
                oldData.Add dKey, rowDict
            End If
        End If
    Next c

    ' ── 2. Wis DA-zone (koppen + data, incl. formattering) ──────────────────────
    If lastCol >= RARM_COL_DA Then
        Dim clearTo As Long
        clearTo = lastCol + 5
        With wsR.Range(wsR.Cells(RARM_ROW_DA, RARM_COL_DA), wsR.Cells(RARM_ROW_VULN, clearTo))
            .ClearContents
            .Interior.ColorIndex = xlNone
            .Borders.LineStyle   = xlNone
        End With
        ' Wis CIA-objectief rijen (4-7) en aangevinkt-teller
        With wsR.Range(wsR.Cells(4, RARM_COL_DA), wsR.Cells(7, clearTo))
            .ClearContents
            .Interior.ColorIndex = xlNone
            .Borders.LineStyle   = xlNone
        End With
        With wsR.Range(wsR.Cells(RARM_ROW_DATA, RARM_COL_DA), wsR.Cells(rarmLast, clearTo))
            .ClearContents
            .Interior.ColorIndex = xlNone
            .Borders.LineStyle   = xlNone
        End With
    End If

    ' ── 3. Herbouw DA-kolommen in DA-sheet volgorde; herstel opgeslagen data ─────
    Dim col As Long
    col = RARM_COL_DA
    Dim daR As Long, daNameRaw As String, isOarch As Boolean, dkLow As String
    Dim daS As Long, ciaC As Variant, ciaI As Variant, ciaA As Variant
    Dim rC8 As String, rCB As String
    For daR = 6 To 105
        daNameRaw = Trim(CStr(wsD.Cells(daR, DA_COL_NAAM).Value))
        If daNameRaw <> "" Then
            isOarch = (Trim(CStr(wsD.Cells(daR, DA_COL_OARCH).Value)) <> "")
            dkLow = LCase(daNameRaw)

            With wsR.Cells(RARM_ROW_DA, col)
                .Value = daNameRaw
                .Font.Bold = True: .Font.Size = 10
                .Font.Color = RGB(15, 43, 70)
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .WrapText = True
                .Interior.Color = IIf(isOarch, RGB(180, 83, 9), RGB(254, 249, 195))
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With

            With wsR.Cells(RARM_ROW_VULN, col)
                If oldVuln.Exists(dkLow) Then .Value = oldVuln(dkLow)
                .Font.Italic = True: .Font.Size = 9
                .Font.Color = RGB(100, 116, 139)
                .Interior.Color = RGB(239, 246, 255)
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .WrapText = True
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With

            ' CIA-objectieven (rijen 4-6) vanuit DA-sheet; aangevinkt-teller (rij 7)
            ciaC = "": ciaI = "": ciaA = ""
            For daS = 6 To 105
                If LCase(Trim(CStr(wsD.Cells(daS, DA_COL_NAAM).Value))) = dkLow Then
                    ciaC = wsD.Cells(daS, 16).Value
                    ciaI = wsD.Cells(daS, 18).Value
                    ciaA = wsD.Cells(daS, 20).Value
                    Exit For
                End If
            Next daS
            rC8 = wsR.Cells(RARM_ROW_DATA, col).Address(False, False)
            rCB = wsR.Cells(1048576, col).Address(False, False)
            wsR.Cells(4, col).Value = CIALbl(ciaC)
            wsR.Cells(5, col).Value = CIALbl(ciaI)
            wsR.Cells(6, col).Value = CIALbl(ciaA)
            wsR.Cells(7, col).Formula = "=COUNTA(" & rC8 & ":" & rCB & ")"
            With wsR.Range(wsR.Cells(4, col), wsR.Cells(7, col))
                .HorizontalAlignment = xlCenter
                .VerticalAlignment = xlCenter
                .Interior.Color = RGB(239, 246, 255)
                .Borders.LineStyle = xlContinuous
                .Borders.Color     = RGB(15, 43, 70)
            End With

            If oldData.Exists(dkLow) Then
                Dim rd As Object
                Set rd = oldData(dkLow)
                Dim rk As Variant
                For Each rk In rd.Keys
                    With wsR.Cells(CLng(rk), col)
                        .Value = rd(rk)
                        .HorizontalAlignment = xlCenter
                        .VerticalAlignment = xlCenter
                    End With
                Next rk
            End If

            wsR.Columns(col).ColumnWidth = 22
            col = col + 1
        End If
    Next daR

    ' ── 4. Titel-merge ───────────────────────────────────────────────────────────
    Dim titleLast As Long
    titleLast = col - 1
    If titleLast < RARM_COL_DA + 9 Then titleLast = RARM_COL_DA + 9
    On Error Resume Next
    wsR.Cells(1, 1).MergeArea.UnMerge
    On Error GoTo 0
    wsR.Range(wsR.Cells(1, 1), wsR.Cells(1, titleLast)).Merge
End Sub

' ── KleurRARMKolom ───────────────────────────────────────────────────────────
' Kleurt de datacellen (rij 4+) van één DA-kolom in de RARM-matrix:
'
'   1. Reset: wisselende grijs/wit achtergrond (streepjespatroon)
'   2. Eigen kwetsbaarheden: controls die een ✔ hebben voor minstens één
'      van de kwetsbaarheden in rij 3 van targetCol → oranje (RGB 255,192,0)
'   3. Overarching spill (alleen als targetCol NIET overarching is):
'      controls die ook aangevinkt zijn voor overarching DA's → lichtgeel
'      (RGB 255,230,153). Dit toont welke controls van organisatorisch niveau
'      ook van toepassing zijn op de huidige DA.
'
' Kleuren worden bepaald via GetMarkedCtrlIDs die de ✔-marks in het
' Kwetsbaarheden-blad uitleest voor de opgegeven kwetsbaarheids-namen.
Sub KleurRARMKolom(targetCol As Long)
    Const RARM_ROW_DA      As Long = 2
    Const RARM_ROW_VULN    As Long = 3
    Const RARM_ROW_DATA    As Long = 8
    Const RARM_COL_ID      As Long = 1
    Const RARM_COL_DA      As Long = 6
    Const DA_COL_NAAM      As Long = 2
    Const DA_COL_OARCH     As Long = 6

    Dim wsR As Worksheet, wsK As Worksheet, wsD As Worksheet
    On Error Resume Next
    Set wsR = ThisWorkbook.Sheets("RARM")
    Set wsK = ThisWorkbook.Sheets("Kwetsbaarheden")
    Set wsD = ThisWorkbook.Sheets("Dependent Assets")
    On Error GoTo 0
    If wsR Is Nothing Or wsK Is Nothing Then Exit Sub

    Dim rarmLast As Long
    rarmLast = wsR.Cells(wsR.Rows.Count, RARM_COL_ID).End(xlUp).Row

    ' Reset naar wisselend grijs/wit met uniforme rand
    Dim bClr As Long
    bClr = RGB(203, 213, 225)   ' grey_border
    Dim r As Long
    For r = RARM_ROW_DATA To rarmLast
        With wsR.Cells(r, targetCol)
            .Interior.Color = IIf((r - RARM_ROW_DATA) Mod 2 = 0, RGB(242, 242, 242), RGB(255, 255, 255))
            .Borders.LineStyle = xlContinuous
            .Borders.Color     = bClr
        End With
    Next r

    ' Eigen kwetsbaarheden → volledige kleur
    Dim selVal As String
    selVal = Trim(CStr(wsR.Cells(RARM_ROW_VULN, targetCol).Value))
    Dim ownMarked As Object
    Set ownMarked = GetMarkedCtrlIDs(wsK, selVal)

    Dim rid As String
    For r = RARM_ROW_DATA To rarmLast
        rid = LCase(Trim(CStr(wsR.Cells(r, RARM_COL_ID).Value)))
        If ownMarked.Exists(rid) Then
            wsR.Cells(r, targetCol).Interior.Color = RGB(255, 192, 0)
        End If
    Next r

    ' Bepaal of targetCol zelf overarching is
    Dim targetDaName As String
    targetDaName = LCase(Trim(CStr(wsR.Cells(RARM_ROW_DA, targetCol).Value)))
    Dim isOarch As Boolean
    isOarch = False
    If Not wsD Is Nothing And targetDaName <> "" Then
        Dim daR As Long
        For daR = 6 To 105
            If LCase(Trim(CStr(wsD.Cells(daR, DA_COL_NAAM).Value))) = targetDaName Then
                isOarch = (Trim(CStr(wsD.Cells(daR, DA_COL_OARCH).Value)) <> "")
                Exit For
            End If
        Next daR
    End If

    ' Als targetCol niet overarching is → ook spill-kleur van overarching kolommen
    If Not isOarch And Not wsD Is Nothing Then
        Dim oarchMarked As Object
        Set oarchMarked = CreateObject("Scripting.Dictionary")
        Dim daCol As Long, oaDaName As String, oaVulns As String
        Dim oaIDs As Object, key As Variant
        daCol = RARM_COL_DA
        Do While Trim(CStr(wsR.Cells(RARM_ROW_DA, daCol).Value)) <> ""
            If daCol <> targetCol Then
                oaDaName = LCase(Trim(CStr(wsR.Cells(RARM_ROW_DA, daCol).Value)))
                For daR = 6 To 105
                    If LCase(Trim(CStr(wsD.Cells(daR, DA_COL_NAAM).Value))) = oaDaName Then
                        If Trim(CStr(wsD.Cells(daR, DA_COL_OARCH).Value)) <> "" Then
                            oaVulns = Trim(CStr(wsR.Cells(RARM_ROW_VULN, daCol).Value))
                            If oaVulns <> "" Then
                                Set oaIDs = GetMarkedCtrlIDs(wsK, oaVulns)
                                For Each key In oaIDs.Keys
                                    If Not oarchMarked.Exists(key) Then oarchMarked.Add key, True
                                Next key
                            End If
                        End If
                        Exit For
                    End If
                Next daR
            End If
            daCol = daCol + 1
        Loop

        For r = RARM_ROW_DATA To rarmLast
            rid = LCase(Trim(CStr(wsR.Cells(r, RARM_COL_ID).Value)))
            If oarchMarked.Exists(rid) Then
                If wsR.Cells(r, targetCol).Interior.Color <> RGB(255, 192, 0) Then
                    wsR.Cells(r, targetCol).Interior.Color = RGB(255, 230, 153)
                End If
            End If
        Next r
    End If
End Sub

' ── TonenAssetPicker ─────────────────────────────────────────────────────────
' Initialiseer de globale toestandsvariabelen voor AssetPicker en open het
' formulier. Geeft niets terug als de geselecteerde rij geen procesnaam bevat
' (voorkomen van een lege popup voor lege rijen).
Sub TonenAssetPicker(targetRow As Long, sourceSheet As String, targetCol As Long)
    g_TargetRow = targetRow
    g_ProcesNaam = CStr(ThisWorkbook.Sheets("Processes").Cells(targetRow, 2).Value)
    g_PickerSourceSheet = sourceSheet
    g_PickerTargetCol = targetCol
    If g_ProcesNaam = "" Then Exit Sub   ' lege rij → geen popup
    AssetPicker.Show
End Sub

' ── ImportInformatieassets ───────────────────────────────────────────────────
' Importeert informatieassets uit "T - Information Assets" naar het
' "Information Assets"-blad. Slaat Naam, Omschrijving, Eigenaar, Dienst en
' Confidentialiteit (col F) op. Opmerkingen (col H) worden niet overschreven.
Sub ImportInformatieassets()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database (informatieassets)"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub

    Application.ScreenUpdating = False

    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub

    Dim rs As Object
    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT * FROM [T - Information Assets]", conn, 0, 1
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Information Assets' niet gevonden in de database.", vbExclamation, "GRC Import"
        conn.Close: Application.ScreenUpdating = True: Exit Sub
    End If
    On Error GoTo 0

    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("Information Assets")
    ws.Range("B6:E105").ClearContents   ' Naam, Omschrijving, Eigenaar, Dienst
    ws.Range("F6:F105").ClearContents   ' Confidentialiteit code (col 6)
    ws.Range("H6:H105").ClearContents   ' Opmerkingen (col 8)

    Dim destRow As Long
    destRow = 6
    Do While Not rs.EOF And destRow <= 105
        ws.Cells(destRow, 2) = FieldVal(rs, "naam", "assetnaam", "asset naam", "name", "asset name", "nom", "actif", "titel", "title")
        ws.Cells(destRow, 3) = FieldVal(rs, "omschrijving", "beschrijving", "description", "beschr", "omschr", "toelichting", "detail")
        ws.Cells(destRow, 4) = FieldVal(rs, "eigenaar", "asseteigenaar", "asset eigenaar", "owner", "propri" & Chr(233) & "taire", "verantwoordelijke")
        ws.Cells(destRow, 5) = FieldVal(rs, "dienst", "afdeling", "organisatie", "entiteit", "department", "service", "business unit")
        ws.Cells(destRow, 6) = MapCls(FieldVal(rs, "confidentialiteit", "confidentiality", "confidentialit" & Chr(233), "conf"))
        destRow = destRow + 1
        rs.MoveNext
    Loop

    rs.Close
    conn.Close
    Application.ScreenUpdating = True
    MsgBox destRow - 6 & " informatieassets ge" & Chr(239) & "mporteerd.", vbInformation, "GRC Import"
End Sub

' ── ExporteerAlles ────────────────────────────────────────────────────────────
' Kopieert de bladen Processes, Information Assets, Dependent Assets en
' Responsible Persons naar een nieuw .xlsx-bestand. Verwijdert het standaard
' lege werkblad (Sheet1/Blad1/Feuil1) uit het exportbestand.
Sub ExporteerAlles()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogSaveAs)
    fd.Title = "Sla de export op als..."
    fd.InitialFileName = "GRC_Export_" & Format(Now, "YYYYMMDD") & ".xlsx"
    If fd.Show = False Then Exit Sub

    Dim exportPath As String
    exportPath = fd.SelectedItems(1)
    If LCase(Right(exportPath, 5)) <> ".xlsx" Then exportPath = exportPath & ".xlsx"

    Application.ScreenUpdating = False
    Dim expWb As Workbook
    Set expWb = Workbooks.Add

    Dim sheetsToCopy As Variant
    sheetsToCopy = Array("Processes", "Information Assets", "Dependent Assets", "Responsible Persons")
    Dim sName As Variant
    For Each sName In sheetsToCopy
        On Error Resume Next
        Dim srcSh As Worksheet
        Set srcSh = ThisWorkbook.Sheets(CStr(sName))
        On Error GoTo 0
        If Not srcSh Is Nothing Then
            srcSh.Copy After:=expWb.Sheets(expWb.Sheets.Count)
            expWb.Sheets(expWb.Sheets.Count).Name = CStr(sName)
        End If
    Next sName

    Application.DisplayAlerts = False
    Dim ws As Worksheet
    For Each ws In expWb.Worksheets
        If ws.Name = "Sheet1" Or ws.Name = "Blad1" Or ws.Name = "Feuil1" Then ws.Delete
    Next ws
    Application.DisplayAlerts = True

    expWb.SaveAs exportPath, FileFormat:=xlOpenXMLWorkbook
    expWb.Close False
    Application.ScreenUpdating = True
    MsgBox "Export opgeslagen:" & vbCrLf & exportPath, vbInformation, "GRC Export"
End Sub

' ── CoreImportKwetsbaarheden ─────────────────────────────────────────────────
' Kernroutine voor het importeren van de kwetsbaarheden-matrix.
' Wordt aangeroepen vanuit ImportAlles (workflow) en ImportKwetsbaarheden
' (standalone knop).
'
' ══ FASE 1 — Kwetsbaarheden en CIA-impact bijwerken ══════════════════════════
' Doel: synchroniseer de kwetsbaarheid-kolommen (rij 2+, kolom C+) van het
'       Kwetsbaarheden-blad met de actuele inhoud van "T - Vulnerabilities".
'
'   Stap 1 : Bouw ctrlRowMap: LCase(2025-ID) → rijnummer in Kwetsbaarheden-blad
'   Stap 2 : Bouw rev23to25 + enkel23Set via CyFun Controls-sheet (kolom F=2025, M=2023)
'   Stap 3 : Bouw refNrMap: RefNr (Long) → CyFun 2023 ID-string via T-CyFunEssentiel
'   Stap 4 : Scan bestaande kwetsbaarheid-kolommen → existingVulnCols map
'   Stap 5 : Laad kwetsbaarheden uit T-Vulnerabilities; werk bestaande kolom
'            bij of voeg nieuwe kolom toe; sla vinkjes (fase 2) leeg
'   Stap 6 : Verwijder orphan-kolommen (in blad maar niet meer in database)
'   Stap 7 : Pas titelrij-merge aan tot het nieuwe kolombereik
'
' ══ FASE 2 — Controls markeren per kwetsbaarheid ════════════════════════════
' Doel: vul ✔-vinkjes in de matrix op basis van de koppelingstabel
'       "LT -  Vulnerability to control - fixed".
'
' Vertaalketen per koppeling (vulnID → ctrlRow):
'   ltCRef  = LT.CyFunControl  (Long = RefNr uit T-CyFunEssentiel)
'     → refNrMap(ltCRef)  = CyFun 2023 ID-string
'     → NormId23()        → id23  = genormaliseerde 2023-ID (bv. "rs.co-3.2")
'     → rev23to25(id23)   → id25  = genormaliseerde 2025-ID (of id23 als ongewijzigd)
'     → ctrlRowMap(id25)  → rijnummer in Kwetsbaarheden-blad
'   vinkje wordt gezet in ws.Cells(ctrlRow, vulnCols(v))
'
' Niet-gematchte 2023-controls (ontbreken in 2025 ESSENTIAL) worden
' gerapporteerd in het blad "Afwijkingen".
Private Sub CoreImportKwetsbaarheden(conn As Object)
    Const COL_ID      As Integer = 1
    Const COL_TITLE   As Integer = 2
    Const COL_V_START As Integer = 3
    Const ROW_TITLE   As Integer = 1
    Const ROW_VULN    As Integer = 2
    Const ROW_C       As Integer = 3
    Const ROW_I       As Integer = 4
    Const ROW_A       As Integer = 5
    Const ROW_DATA    As Integer = 6
    Const COL_CF25    As Integer = 6
    Const COL_CF23    As Integer = 13
    Const ROW_CF_DATA As Integer = 4

    Dim ws As Worksheet, wsCC As Worksheet, wsAfw As Worksheet
    Dim rs As Object, rsCtrl As Object, rsLT As Object
    Dim rr As Long, v As Integer
    Dim ctrlRowMap As Object, refNrMap As Object
    Dim rev23to25 As Object, enkel23Set As Object, afwList As Object
    Dim lastRow As Long, nVuln As Integer, nMatched As Long

    Set ws = ThisWorkbook.Sheets("Kwetsbaarheden")

    ' ── 1. Ctrl-rij-map: LCase(ID) → rijnummer ───────────────────────────────
    Set ctrlRowMap = CreateObject("Scripting.Dictionary")
    rr = ROW_DATA
    Do While ws.Cells(rr, COL_ID).Value <> ""
        Dim ck As String
        ck = LCase(Trim(CStr(ws.Cells(rr, COL_ID).Value)))
        If Not ctrlRowMap.Exists(ck) Then ctrlRowMap.Add ck, rr
        rr = rr + 1
    Loop
    If ctrlRowMap.Count = 0 Then
        MsgBox "Kwetsbaarheden-sheet bevat geen controls (rij 6+ van kolom A is leeg)." & vbCrLf & _
               "Herbouw het bestand via build_grc.py en probeer opnieuw.", vbExclamation, "GRC Import"
        Exit Sub
    End If
    lastRow = ws.Cells(ws.Rows.Count, COL_ID).End(xlUp).Row

    ' ── 2. 2023→2025 reverse mapping + "Enkel 2023"-set via CyFun Controls-sheet ──
    Set rev23to25  = CreateObject("Scripting.Dictionary")
    Set enkel23Set = CreateObject("Scripting.Dictionary")
    On Error Resume Next
    Set wsCC = ThisWorkbook.Sheets("CyFun Controls")
    On Error GoTo 0
    If Not wsCC Is Nothing Then
        Dim ccLast As Long
        ccLast = Application.Max( _
            wsCC.Cells(wsCC.Rows.Count, COL_CF25).End(xlUp).Row, _
            wsCC.Cells(wsCC.Rows.Count, COL_CF23).End(xlUp).Row)
        Dim rCC As Long
        For rCC = ROW_CF_DATA To ccLast
            Dim r25t As String, r23t As String
            r25t = Trim(CStr(wsCC.Cells(rCC, COL_CF25).Value))
            r23t = Trim(CStr(wsCC.Cells(rCC, COL_CF23).Value))
            If r25t <> "" And r23t <> "" Then
                Dim p25 As Integer, p23 As Integer
                p25 = InStr(r25t, ":"): p23 = InStr(r23t, ":")
                Dim i25 As String, i23 As String
                i25 = LCase(Trim(Replace(IIf(p25 > 0, Left(r25t, p25 - 1), r25t), Chr(9), "")))
                i23 = NormId23(r23t)
                If i23 <> "" And i25 <> "" And Not rev23to25.Exists(i23) Then
                    rev23to25.Add i23, i25
                End If
            ElseIf r25t = "" And r23t <> "" Then
                Dim e23 As String
                e23 = NormId23(r23t)
                If e23 <> "" And Not enkel23Set.Exists(e23) Then
                    enkel23Set.Add e23, True
                End If
            End If
        Next rCC
    End If

    ' ── 3. RefNr → CyFun 2023 ID via T-CyFunEssentiel ────────────────────────
    Set refNrMap = CreateObject("Scripting.Dictionary")
    Set rsCtrl = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rsCtrl.Open "SELECT RefNr, Requirement FROM [T - CyFunEssentiel]", conn
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - CyFunEssentiel' niet gevonden in de database.", vbExclamation, "GRC Import"
        On Error GoTo 0
        Exit Sub
    End If
    On Error GoTo 0
    Do While Not rsCtrl.EOF
        Dim refNr As Long
        refNr = CLng(rsCtrl.Fields("RefNr").Value)
        Dim reqVal As String
        reqVal = CStr(rsCtrl.Fields("Requirement").Value)
        If Not refNrMap.Exists(refNr) Then refNrMap.Add refNr, reqVal
        rsCtrl.MoveNext
    Loop
    rsCtrl.Close

    ' ══════════════════════════════════════════════════════════════════════════
    ' FASE 1 — Kwetsbaarheden en CIA-impact importeren op bestaande positie
    ' ══════════════════════════════════════════════════════════════════════════

    ' ── 4. Scan bestaande kwetsbaarheid-kolommen (rij 2, C+) ─────────────────
    ' existingVulnCols: LCase(naam) → kolomnummer (van vorige import)
    Dim existingVulnCols As Object
    Set existingVulnCols = CreateObject("Scripting.Dictionary")
    Dim usedEndCol As Long
    usedEndCol = ws.UsedRange.Column + ws.UsedRange.Columns.Count - 1
    Dim scanCol As Long
    For scanCol = COL_V_START To usedEndCol
        Dim scanHdr As String
        scanHdr = Trim(CStr(ws.Cells(ROW_VULN, scanCol).Value))
        If scanHdr <> "" Then
            Dim scanKey As String
            scanKey = LCase(scanHdr)
            If Not existingVulnCols.Exists(scanKey) Then
                existingVulnCols.Add scanKey, scanCol
            End If
        End If
    Next scanCol

    ' Bepaal startkolom voor nieuwe kwetsbaarheden
    Dim nextNewCol As Long
    If existingVulnCols.Count > 0 Then
        Dim maxExistCol As Long: maxExistCol = COL_V_START - 1
        Dim eKey As Variant
        For Each eKey In existingVulnCols.Keys
            Dim ec As Long: ec = CLng(existingVulnCols.Item(eKey))
            If ec > maxExistCol Then maxExistCol = ec
        Next eKey
        nextNewCol = maxExistCol + 1
    Else
        nextNewCol = COL_V_START
    End If

    ' ── 5. Kwetsbaarheden laden; bijwerken op bestaande positie of nieuw toevoegen ──
    Dim vulnIDs()   As Long
    Dim vulnNames() As String
    Dim vulnC()     As Boolean, vulnI() As Boolean, vulnA() As Boolean
    Dim vulnCols()  As Long
    Dim matchedCols As Object  ' bijgehouden kolomnummers (voor orphan-detectie)
    Set matchedCols = CreateObject("Scripting.Dictionary")
    nVuln = 0
    ReDim vulnIDs(0 To 99):   ReDim vulnNames(0 To 99)
    ReDim vulnC(0 To 99):     ReDim vulnI(0 To 99):   ReDim vulnA(0 To 99)
    ReDim vulnCols(0 To 99)

    Set rs = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rs.Open "SELECT Reference, Vulnerability, Confidentialiteit, Integriteit, " & _
            "Beschikbaarheid FROM [T - Vulnerabilities] ORDER BY Reference", conn
    If Err.Number <> 0 Then
        MsgBox "Tabel 'T - Vulnerabilities' niet gevonden of ongeldige query." & vbCrLf & _
               Err.Description, vbExclamation, "GRC Import"
        On Error GoTo 0
        Exit Sub
    End If
    On Error GoTo 0

    Do While Not rs.EOF And nVuln < 100
        Dim vID As Long, vNm As String
        Dim vC As Boolean, vI As Boolean, vA As Boolean
        vID = CLng(rs.Fields("Reference").Value)
        vNm = CStr(rs.Fields("Vulnerability").Value)
        vC  = CBool(rs.Fields("Confidentialiteit").Value)
        vI  = CBool(rs.Fields("Integriteit").Value)
        vA  = CBool(rs.Fields("Beschikbaarheid").Value)

        ' Gebruik bestaande kolom (op naam) of maak nieuwe aan
        Dim targetCol As Long
        Dim vNmKey As String: vNmKey = LCase(vNm)
        If existingVulnCols.Exists(vNmKey) Then
            targetCol = CLng(existingVulnCols.Item(vNmKey))
        Else
            targetCol = nextNewCol
            nextNewCol = nextNewCol + 1
            ws.Columns(targetCol).ColumnWidth = 10
        End If

        vulnIDs(nVuln)   = vID
        vulnNames(nVuln) = vNm
        vulnC(nVuln) = vC: vulnI(nVuln) = vI: vulnA(nVuln) = vA
        vulnCols(nVuln)  = targetCol
        If Not matchedCols.Exists(targetCol) Then matchedCols.Add targetCol, True

        ' Rij 2: naam (bijwerken)
        With ws.Cells(ROW_VULN, targetCol)
            .Value = vNm
            .Interior.Color = IIf(nVuln Mod 2 = 0, RGB(208, 220, 243), RGB(255, 255, 255))
            .Font.Bold = True: .Font.Size = 9
            .HorizontalAlignment = xlCenter: .VerticalAlignment = xlCenter
            .WrapText = True
            .Borders.LineStyle = xlContinuous: .Borders.Weight = xlThin
        End With

        ' Rijen 3–5: CIA-impact (bijwerken)
        Dim cia As Integer
        Dim ciaRows(1 To 3) As Integer
        Dim ciaFlags(1 To 3) As Boolean
        ciaRows(1) = ROW_C: ciaFlags(1) = vC
        ciaRows(2) = ROW_I: ciaFlags(2) = vI
        ciaRows(3) = ROW_A: ciaFlags(3) = vA
        Dim ciaColors(1 To 3) As Long
        ciaColors(1) = RGB(173, 216, 230)
        ciaColors(2) = RGB(198, 239, 206)
        ciaColors(3) = RGB(255, 235, 156)
        For cia = 1 To 3
            With ws.Cells(ciaRows(cia), targetCol)
                If ciaFlags(cia) Then
                    .Value = ChrW(10004)
                    .Interior.Color = ciaColors(cia)
                    .Font.Bold = True: .Font.Size = 9
                Else
                    .Value = ""
                    .Interior.Color = RGB(217, 217, 217)
                End If
                .HorizontalAlignment = xlCenter: .VerticalAlignment = xlCenter
                .Borders.LineStyle = xlContinuous: .Borders.Weight = xlThin
            End With
        Next cia

        ' Wis bestaande ✔-marks (worden in fase 2 opnieuw ingevuld)
        If lastRow >= ROW_DATA Then
            ws.Range(ws.Cells(ROW_DATA, targetCol), _
                     ws.Cells(lastRow, targetCol)).ClearContents
        End If

        nVuln = nVuln + 1
        rs.MoveNext
    Loop
    rs.Close

    ' ── 6. Verwijder orphan-kolommen (bestonden maar niet meer in database) ───
    ' Verzamel orphan-kolommen (in existingVulnCols maar NIET in matchedCols)
    Dim orphanCols() As Long
    Dim nOrphan As Integer: nOrphan = 0
    Dim exKey2 As Variant
    For Each exKey2 In existingVulnCols.Keys
        Dim exColNr As Long: exColNr = CLng(existingVulnCols.Item(exKey2))
        If Not matchedCols.Exists(exColNr) Then
            ReDim Preserve orphanCols(0 To nOrphan)
            orphanCols(nOrphan) = exColNr
            nOrphan = nOrphan + 1
        End If
    Next exKey2
    ' Sorteer descenderend (bubble sort) zodat rechts-naar-links wordt gewist
    Dim si As Integer, sj As Integer, tmpCol As Long
    For si = 0 To nOrphan - 2
        For sj = 0 To nOrphan - 2 - si
            If orphanCols(sj) < orphanCols(sj + 1) Then
                tmpCol = orphanCols(sj): orphanCols(sj) = orphanCols(sj + 1): orphanCols(sj + 1) = tmpCol
            End If
        Next sj
    Next si
    ' Verwijder orphan-kolommen en pas vulnCols aan voor kolom-verschuiving
    Dim oi As Integer, adj As Integer
    For oi = 0 To nOrphan - 1
        Dim delCol As Long: delCol = orphanCols(oi)
        ws.Columns(delCol).Delete
        For adj = 0 To nVuln - 1
            If vulnCols(adj) > delCol Then vulnCols(adj) = vulnCols(adj) - 1
        Next adj
    Next oi

    ' ── 7. Titelrij samenvoegen tot laatste kwetsbaarheid-kolom ──────────────
    Dim lastVulnCol As Long: lastVulnCol = COL_V_START - 1
    For v = 0 To nVuln - 1
        If vulnCols(v) > lastVulnCol Then lastVulnCol = vulnCols(v)
    Next v
    On Error Resume Next
    ws.Cells(ROW_TITLE, 1).MergeArea.UnMerge
    On Error GoTo 0
    If lastVulnCol >= COL_V_START Then
        ws.Range(ws.Cells(ROW_TITLE, 1), ws.Cells(ROW_TITLE, lastVulnCol)).Merge
        ws.Cells(ROW_TITLE, 1).HorizontalAlignment = xlLeft
    End If

    ' ══════════════════════════════════════════════════════════════════════════
    ' FASE 2 — Alle kolommen overlopen: controls per kwetsbaarheid opzoeken en markeren
    ' ══════════════════════════════════════════════════════════════════════════

    ' ── 8. LT-tabel in geheugen laden: vulnID → set van ctrlRefs ─────────────
    Set afwList = CreateObject("Scripting.Dictionary")
    nMatched = 0

    Dim vulnLTMap As Object
    Set vulnLTMap = CreateObject("Scripting.Dictionary")

    Set rsLT = CreateObject("ADODB.Recordset")
    On Error Resume Next
    rsLT.Open "SELECT Vulnerability, CyFunControl " & _
              "FROM [LT -  Vulnerability to control - fixed]", conn
    If Err.Number <> 0 Then
        MsgBox "Fout bij openen LT-tabel: " & Err.Description, vbExclamation, "GRC Import"
        On Error GoTo 0
        Exit Sub
    End If
    On Error GoTo 0

    Do While Not rsLT.EOF
        Dim ltVId As Long, ltCRef As Long
        ltVId  = CLng(rsLT.Fields("Vulnerability").Value)
        ltCRef = CLng(rsLT.Fields("CyFunControl").Value)
        If Not vulnLTMap.Exists(ltVId) Then
            vulnLTMap.Add ltVId, CreateObject("Scripting.Dictionary")
        End If
        Dim ltDict As Object
        Set ltDict = vulnLTMap.Item(ltVId)
        If Not ltDict.Exists(ltCRef) Then ltDict.Add ltCRef, True
        rsLT.MoveNext
    Loop
    rsLT.Close

    ' ── 9. Per kwetsbaarheid: popup tonen en controls markeren ───────────────
    For v = 0 To nVuln - 1
        Application.StatusBar = "Kwetsbaarheden importeren... (" & (v + 1) & "/" & nVuln & "): " & vulnNames(v)
        DoEvents

        If vulnLTMap.Exists(vulnIDs(v)) Then
            Dim vLTDict As Object
            Set vLTDict = vulnLTMap.Item(vulnIDs(v))
            Dim ltRef As Variant
            For Each ltRef In vLTDict.Keys
                Dim ltCtrlRef As Long: ltCtrlRef = CLng(ltRef)
                If refNrMap.Exists(ltCtrlRef) Then
                    Dim id23v As String
                    id23v = NormId23(CStr(refNrMap.Item(ltCtrlRef)))

                    Dim id25v As String
                    If rev23to25.Exists(id23v) Then
                        id25v = CStr(rev23to25.Item(id23v))
                    Else
                        id25v = id23v
                    End If

                    Dim ctrlRow As Long: ctrlRow = 0
                    If ctrlRowMap.Exists(id25v) Then
                        ctrlRow = CLng(ctrlRowMap.Item(id25v))
                        nMatched = nMatched + 1
                        With ws.Cells(ctrlRow, vulnCols(v))
                            .Value = ChrW(10004)
                            .Interior.Color = RGB(198, 239, 206)
                            .HorizontalAlignment = xlCenter
                            .VerticalAlignment = xlCenter
                            .Font.Size = 10
                        End With
                    Else
                        Dim afwK As String
                        afwK = vulnNames(v) & "|" & id23v
                        If Not afwList.Exists(afwK) Then
                            Dim cs As String: cs = ""
                            If vulnC(v) Then cs = cs & "C"
                            If vulnI(v) Then cs = cs & "I"
                            If vulnA(v) Then cs = cs & "A"
                            Dim reden As String
                            If enkel23Set.Exists(id23v) Then
                                reden = "Enkel in 2023"
                            Else
                                reden = "Niet gevonden"
                            End If
                            afwList.Add afwK, Array(vulnNames(v), cs, id23v, reden)
                        End If
                    End If
                End If
            Next ltRef
        End If
    Next v
    Application.StatusBar = False

    ' ── 9. Afwijkingen-sheet bijwerken ────────────────────────────────────────
    On Error Resume Next
    Set wsAfw = ThisWorkbook.Sheets("Afwijkingen")
    On Error GoTo 0
    If wsAfw Is Nothing Then
        Set wsAfw = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        wsAfw.Name = "Afwijkingen"
    End If

    Dim aLastRow As Long
    aLastRow = wsAfw.Cells(wsAfw.Rows.Count, 1).End(xlUp).Row
    If aLastRow >= 3 Then wsAfw.Rows("3:" & aLastRow).ClearContents

    Dim aRow As Long
    aRow = 3
    Dim afwKey2 As Variant
    For Each afwKey2 In afwList.Keys
        Dim info As Variant
        info = afwList.Item(afwKey2)
        wsAfw.Cells(aRow, 1).Value = info(0)
        wsAfw.Cells(aRow, 2).Value = info(1)
        wsAfw.Cells(aRow, 3).Value = info(2)
        wsAfw.Cells(aRow, 4).Value = info(3)
        aRow = aRow + 1
    Next afwKey2

    MsgBox "Geladen: " & nVuln & " kwetsbaarheden, " & nMatched & " matches, " & _
           afwList.Count & " afwijkingen.", vbInformation, "GRC Import"
End Sub

' ── ImportKwetsbaarheden ─────────────────────────────────────────────────────
' Standalone wrapper die de gebruiker vraagt een Access-database te kiezen
' en vervolgens CoreImportKwetsbaarheden aanroept.
Sub ImportKwetsbaarheden()
    Dim fd As FileDialog
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecteer de Access-database (kwetsbaarheden)"
    fd.Filters.Clear
    fd.Filters.Add "Access-bestanden", "*.accdb; *.mdb"
    If fd.Show = False Then Exit Sub
    Application.ScreenUpdating = False
    Dim conn As Object
    Set conn = OpenAccess(fd.SelectedItems(1))
    If conn Is Nothing Then Application.ScreenUpdating = True: Exit Sub
    CoreImportKwetsbaarheden conn
    conn.Close
    Application.ScreenUpdating = True
End Sub

' ── NormId23 ─────────────────────────────────────────────────────────────────
' Normaliseert een CyFun 2023 ID-string naar een kort, lowercase, canoniek ID.
' Dit is nodig omdat de database CyFun 2023 IDs opslaat in uiteenlopende formaten
' afhankelijk van het assurance-niveau en de bron.
'
' Voorbeelden:
'   "IMPORTANT_RS.CO-3.2: The organization..." → "rs.co-3.2"
'   "RS.CO-3.2: The organization..."           → "rs.co-3.2"
'   "RS.CO-3.2.a"                              → "rs.co-3.2"  (sub-letter gestript)
'   "rs.co-3-2"                                → "rs.co-3.2"  (koppelteken → punt)
'   "RS.CO-3.2"                                → "rs.co-3.2"
'
' Stap 1 : strip niveau-prefix (IMPORTANT_, BASIC_, ESSENTIAL_)
' Stap 2 : strip beschrijvingstekst na dubbele punt ("ID: tekst" → "ID")
' Stap 3 : lowercase
' Stap 4 : strip trailing letter-extensie (.a t/m .z voor sub-controls)
' Stap 5 : converteer trailing -N → .N  (bv. "rs.co-3-2" → "rs.co-3.2")
'          zodat de ID consistent is met de 2025-notatie
Function NormId23(s As String) As String
    Dim r As String
    r = Trim(s)

    ' Stap 1: strip niveau-prefix (IMPORTANT_, BASIC_, ESSENTIAL_)
    If Left(LCase(r), 10) = "important_" Then
        r = Mid(r, 11)
    ElseIf Left(LCase(r), 6) = "basic_" Then
        r = Mid(r, 7)
    ElseIf Left(LCase(r), 10) = "essential_" Then
        r = Mid(r, 11)
    End If

    ' Stap 2: strip beschrijving na dubbele punt ("RS.CO-3.2: tekst" → "RS.CO-3.2")
    Dim colon As Integer
    colon = InStr(r, ":")
    If colon > 0 Then r = Trim(Left(r, colon - 1))

    r = LCase(r)

    ' Stap 3: strip trailing letter-extensie (.a tot .z)
    Dim lastDot As Integer
    lastDot = InStrRev(r, ".")
    If lastDot > 0 Then
        Dim suf As String
        suf = Mid(r, lastDot + 1)
        If Len(suf) = 1 And suf >= "a" And suf <= "z" Then
            r = Left(r, lastDot - 1)
        End If
    End If

    ' Stap 4: trailing -N → .N  (bv. "rs.co-3-2" → "rs.co-3.2")
    Dim i As Integer, lastHyph As Integer
    lastHyph = 0
    For i = Len(r) To 1 Step -1
        Dim ch As String
        ch = Mid(r, i, 1)
        If ch >= "0" And ch <= "9" Then
            ' nog in cijfer-zone, blijf zoeken
        ElseIf ch = "-" Then
            lastHyph = i
            Exit For
        Else
            Exit For
        End If
    Next i
    If lastHyph > 0 Then
        Dim tail As String
        tail = Mid(r, lastHyph + 1)
        If Len(tail) > 0 Then
            r = Left(r, lastHyph - 1) & "." & tail
        End If
    End If

    NormId23 = r
End Function
'''

# ══════════════════════════════════════════════════════════════════════════════
# WERKBOEK OPBOUWEN
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# _Lang  (verborgen vertaalblad + benoemde bereiken)
# Bevat in kolommen A-D alle vertalingssleutels + NL/FR/EN-teksten (rijen 2+).
# Kolommen F-H (rijen 250+) bevatten de dropdown-lijsten CLS/ROLES/ASSET_TYPES.
# Het blad wordt verborgen zodat eindgebruikers het niet per ongeluk aanpassen.
# ─────────────────────────────────────────────────────────────────────────────
ws_lang = wb.active
ws_lang.title = "_Lang"

# Kolomkoppen voor de vertalingstabel
for col, h in enumerate(["Sleutel", "NL", "FR", "EN"], 1):
    c = ws_lang.cell(row=1, column=col, value=h)
    c.font = Font(name="Calibri", bold=True, color=C["navy"])
    c.fill = PatternFill("solid", fgColor=C["grey_mid"])

# Alle TRANSLATIONS tuples wegschrijven (sleutel + NL + FR + EN)
for r, row_data in enumerate(TRANSLATIONS, 2):
    for col, val in enumerate(row_data, 1):
        ws_lang.cell(row=r, column=col, value=val)

ws_lang.column_dimensions["A"].width = 28
for letter in ["B", "C", "D"]:
    ws_lang.column_dimensions[letter].width = 72

# Classificatiewaarden F=NL, G=FR, H=EN  — 5 niveaus
# Startrijnen zijn bewust ver van rij 1 geplaatst (250+) om conflicten
# met de vertalingstabel te vermijden zonder een apart blad te hoeven aanmaken.
CLS_START   = 250   # rijen 250-254: classificatieniveau-labels
ROLES_START = 258   # rijen 258-265: GRC-rollen (8 waarden)
TYPES_START = 270   # rijen 270-279: asset-types (10 waarden)

for col_off, lang in enumerate(["NL", "FR", "EN"]):
    col = 6 + col_off   # kolom F (NL), G (FR), H (EN)
    for row_off, val in enumerate(CLS[lang]):
        ws_lang.cell(row=CLS_START + row_off, column=col, value=val)
    for row_off, val in enumerate(ROLES[lang]):
        ws_lang.cell(row=ROLES_START + row_off, column=col, value=val)
    for row_off, val in enumerate(ASSET_TYPES[lang]):
        ws_lang.cell(row=TYPES_START + row_off, column=col, value=val)

ws_lang.sheet_state = "hidden"

# Benoemde bereiken: cls_NL, cls_FR, cls_EN, roles_NL, roles_FR, roles_EN,
# types_NL, types_FR, types_EN.
# Dropdowns in de gegevensbladen gebruiken =INDIRECT("cls_"&Config!$D$9)
# zodat ze automatisch wisselen bij taalwijziging.
n_cls   = len(CLS["NL"])          # 5 classificatieniveaus
n_roles = len(ROLES["NL"])        # 8 rollen
n_types = len(ASSET_TYPES["NL"])  # 10 asset-types

for col_off, lang in enumerate(["NL", "FR", "EN"]):
    col = 6 + col_off
    cl = get_column_letter(col)
    wb.defined_names[f"cls_{lang}"]   = DefinedName(f"cls_{lang}",   attr_text=f"_Lang!${cl}${CLS_START}:${cl}${CLS_START+n_cls-1}")
    wb.defined_names[f"roles_{lang}"] = DefinedName(f"roles_{lang}", attr_text=f"_Lang!${cl}${ROLES_START}:${cl}${ROLES_START+n_roles-1}")
    wb.defined_names[f"types_{lang}"] = DefinedName(f"types_{lang}", attr_text=f"_Lang!${cl}${TYPES_START}:${cl}${TYPES_START+n_types-1}")

# ── Layout-hulpfuncties ───────────────────────────────────────────────────────
def sheet_title_bar(ws, text, n_cols):
    """
    Plaatst een tweeregelige titelbalk bovenaan een werkblad.
    Rij 1 : grote marineblauwe balk met de hoofdtitel (tekst of t()-formule).
    Rij 2 : kleinere blauwe balk met de tagline 'Governance · Risico · Compliance'
             (via t("app_tagline"), taalvolgend).

    Parameters
    ----------
    ws     : openpyxl.Worksheet
    text   : str — celwaarde of formule voor de hoofdtitel
    n_cols : int — aantal kolommen dat de balk overspant
    """
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1");  ws.row_dimensions[1].height = 52
    c = ws["A1"]; c.value = text
    c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=20, color=C["white"])
    c.alignment = align("center", "center")
    ws.merge_cells(f"A2:{last}2"); ws.row_dimensions[2].height = 22
    c = ws["A2"]; c.value = t("app_tagline")
    c.fill = fill("blue_mid"); c.font = Font(name="Calibri", size=11, color=C["blue_xlight"], italic=True)
    c.alignment = align("center", "center")

def section_header(ws, row, key, n_cols, col_start=1):
    """
    Plaatst een donkerblauwe sectiekop (gemerged over n_cols kolommen).
    Wordt gebruikt voor blokken zoals 'ALGEMENE INSTELLINGEN' of 'IMPORTEREN'.

    Parameters
    ----------
    ws        : openpyxl.Worksheet
    row       : int — rijnummer
    key       : str — vertaalsleutel (t()-formule)
    n_cols    : int — breedte in kolommen
    col_start : int — startkolom (standaard 1 = A)
    """
    ws.row_dimensions[row].height = 20
    lc = get_column_letter(col_start + n_cols - 1)
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{lc}{row}")
    c = ws.cell(row=row, column=col_start, value=t(key))
    c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=9, color=C["blue_light"])
    c.alignment = align("left", "center", indent=1)

def label_value_row(ws, row, col_lbl, col_val, label, value,
                    val_bold=False, val_color="text", val_size=10, dropdown=None, editable=True):
    """
    Plaatst een label-waarde paar op één rij: grijs label links, invoercel rechts.
    Editeerbare cellen krijgen een witte achtergrond + blauwe rand;
    niet-editeerbare cellen (metadata) krijgen grijs + grijze rand.
    Optioneel kan een dropdown-validatie worden meegegeven.

    Parameters
    ----------
    ws        : openpyxl.Worksheet
    row       : int — rijnummer
    col_lbl   : int — kolomindex van de labelcel
    col_val   : int — kolomindex van de waarde-/invoercel
    label     : str — labeltekst (wordt niet via t() vertaald — al een formule of vaste tekst)
    value     : str — beginwaarde van de invoercel
    val_bold  : bool — vetgedrukt voor de waarde
    val_color : str — kleursleutel voor de waarde
    val_size  : int — lettergrootte voor de waarde
    dropdown  : str | None — Excel-formule voor dropdownvalidatie (bv. '"NL,FR,EN"')
    editable  : bool — True = witte cel met blauwe rand, False = grijs (read-only)

    Geeft terug
    -----------
    openpyxl.Cell — de waarde-/invoercel (voor verdere stijlinstelling)
    """
    ws.row_dimensions[row].height = 26
    lc = ws.cell(row=row, column=col_lbl, value=label)
    lc.font = font(10, bold=True, color="grey_dark"); lc.alignment = align("right", "center")
    lc.fill = fill("grey_light"); lc.border = border_all()
    vc = ws.cell(row=row, column=col_val, value=value)
    vc.font = Font(name="Calibri", size=val_size, bold=val_bold, color=C[val_color])
    vc.alignment = align("left", "center", indent=1)
    vc.fill = fill("white") if editable else fill("grey_light")
    vc.border = border_all("blue_mid" if editable else "grey_border")
    if dropdown:
        dv = DataValidation(type="list", formula1=dropdown, showDropDown=False, allow_blank=False)
        dv.sqref = f"{get_column_letter(col_val)}{row}"
        ws.add_data_validation(dv)
    return vc

def hint_row(ws, row, n_cols, key, col_start=1):
    """
    Plaatst een cursieve hinttekst (grijs, klein) over n_cols kolommen.
    Wordt gebruikt voor aanwijzingen onder invoervelden (bv. taalwisseltip).

    Parameters
    ----------
    ws        : openpyxl.Worksheet
    row       : int — rijnummer
    n_cols    : int — breedte in kolommen
    key       : str — vertaalsleutel (t()-formule)
    col_start : int — startkolom (standaard 1 = A)
    """
    ws.row_dimensions[row].height = 18
    lc = get_column_letter(col_start + n_cols - 1)
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{lc}{row}")
    c = ws.cell(row=row, column=col_start, value=t(key))
    c.font = Font(name="Calibri", size=9, color=C["subtext"], italic=True)
    c.alignment = align("left", "center", indent=1)

def quicklinks_block(ws, row_start, links, n_cols, col_start=1):
    """Renders a SNELKOPPELINGEN block starting at row_start."""
    ws.row_dimensions[row_start].height = 10
    r = row_start + 1
    ws.row_dimensions[r].height = 20
    lc = get_column_letter(col_start + n_cols - 1)
    ws.merge_cells(f"{get_column_letter(col_start)}{r}:{lc}{r}")
    c = ws.cell(row=r, column=col_start, value=t("ui_quicklinks"))
    c.fill = fill("blue_xlight"); c.font = Font(name="Calibri", bold=True, size=9, color=C["blue_mid"])
    c.alignment = align("left", "center", indent=1)
    for key, target in links:
        r += 1
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"{get_column_letter(col_start)}{r}:{lc}{r}")
        c = ws.cell(row=r, column=col_start, value=t(key))
        c.hyperlink = f"#{target}"; c.fill = fill("blue_xlight")
        c.font = Font(name="Calibri", size=10, color=C["blue_mid"], underline="single")
        c.alignment = align("left", "center", indent=1)

def build_cls_sheet(ws, title_key, subtitle_key, id_prefix,
                    col_headers_before, col_headers_after,
                    widths, data_start=6, data_end=105):
    """
    Bouwt de basisstructuur van een classificatiegegevensblad (titelbalk,
    subtitelrij, kolomkoppen). Verwerkt de enkelvoudige kolomkoppen vóór en
    na de classificatiekolommen; de classificatiekolom-paren worden nadien
    afzonderlijk toegevoegd via add_cls_pairs().

    Parameters
    ----------
    ws                 : openpyxl.Worksheet
    title_key          : str — vertaalsleutel voor de paginatitel (rij 1)
    subtitle_key       : str — vertaalsleutel voor de ondertitel (rij 3)
    id_prefix          : str — voorvoegsel voor automatische ID-formule ("P", "A", "D")
    col_headers_before : list[(key, col_idx)] — kolomkoppen vóór de classificaties
    col_headers_after  : list[(key, col_idx)] — kolomkoppen na de classificaties
    widths             : list[float] — kolombreedtes (kolom 1 t/m N)
    data_start         : int — eerste datarij (standaard 6)
    data_end           : int — laatste datarij (standaard 105 = max. 100 records)

    Geeft terug
    -----------
    tuple (HDR_ROW, data_start, data_end, N) — benodigd door add_cls_pairs()
    en de data-rijlus in de aanroepende code
    """
    N = len(widths)
    set_col_widths(ws, widths)
    no_gridlines(ws)
    sheet_title_bar(ws, t(title_key), N)
    ws.row_dimensions[3].height = 22
    ws.merge_cells(f"A3:{get_column_letter(N)}3")
    c = ws["A3"]; c.value = t(subtitle_key)
    c.fill = fill("blue_xlight"); c.font = font(10, italic=True, color="grey_dark")
    c.alignment = align("center", "center")
    ws.row_dimensions[4].height = 8

    HDR_ROW = 5
    ws.row_dimensions[HDR_ROW].height = 44

    # Enkelvoudige koppen
    for key, col_idx in col_headers_before + col_headers_after:
        c = ws.cell(row=HDR_ROW, column=col_idx, value=t(key))
        c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

    return HDR_ROW, data_start, data_end, N

# ── Gepaarde classificatiekolommen ───────────────────────────────────────────
def add_cls_pairs(ws, cls_pairs, HDR_ROW, DATA_START, DATA_END, max_level=5):
    """
    Voegt voor elke classificatie-paar (code-kolom + label-kolom) het volgende toe:
      - Gemerged kolomkop in HDR_ROW (marineblauwe achtergrond)
      - Per datarij: code-cel (dropdown 1-5) + label-cel (=INDEX-formule op Reference Values)
      - DataValidation: dropdown beperkt tot 1..max_level (standaard 5)
      - Voorwaardelijke opmaak (CF): 5 kleurrings op basis van code-celwaarde

    De label-cel bevat een formule die het niveau-label ophaalt uit het
    'Reference Values'-blad (B8:B12) zodat de tekst automatisch meebeweegt
    als de gebruiker de taal wijzigt (niet via t(), maar via directe celverwijzing).

    Parameters
    ----------
    ws        : openpyxl.Worksheet
    cls_pairs : list[(key, code_col, label_col)] — tripels per classificatiedimensie
    HDR_ROW   : int — rijnummer van de kolomkop
    DATA_START: int — eerste datarij
    DATA_END  : int — laatste datarij
    max_level : int — maximum classificatieniveau (standaard 5; gebruik 4 voor niet-Classified)
    """
    for key, code_col, label_col in cls_pairs:
        cl = get_column_letter(code_col); ll = get_column_letter(label_col)
        ws.merge_cells(f"{cl}{HDR_ROW}:{ll}{HDR_ROW}")
        c = ws.cell(row=HDR_ROW, column=code_col, value=t(key))
        c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

    for r in range(DATA_START, DATA_END + 1):
        alt = "blue_xlight" if r % 2 == 0 else "white"
        for _, code_col, label_col in cls_pairs:
            cl = get_column_letter(code_col); ll = get_column_letter(label_col)
            cc = ws.cell(row=r, column=code_col)
            cc.alignment = align("center", "center")
            cc.font = Font(name="Calibri", bold=True, size=11, color=C["subtext"])
            lc = ws.cell(row=r, column=label_col)
            # Label-formule: haalt niveau-naam op uit 'Reference Values'!B8:B12
            # (rij 8 = niveau 1 / Laag, rij 12 = niveau 5 / Classified).
            # IFERROR valt terug op "?" bij ongeldige invoer.
            # De cel is grijs en niet editeerbaar — VBA-imports schrijven hier nooit naar.
            lc.value = (
                f'=IF({cl}{r}="","",IFERROR('
                f'INDEX(\'Reference Values\'!$B$8:$B$12,{cl}{r}),"?"))'
            )
            lc.alignment = align("center", "center")
            lc.font = Font(name="Calibri", size=10, italic=True)
            lc.fill = fill("grey_light")

    # Dropdown beperkt tot max_level: genereert bv. "1,2,3,4,5" als formula1-waarde.
    # showDropDown=False = dropdown-pijltje altijd zichtbaar (counter-intuitieve Excel API).
    levels_str = ",".join(str(i) for i in range(1, max_level + 1))
    dv = DataValidation(type="list", formula1=f'"{levels_str}"', showDropDown=False, allow_blank=True)
    dv.sqref = " ".join(
        f"{get_column_letter(cc)}{DATA_START}:{get_column_letter(cc)}{DATA_END}"
        for _, cc, _ in cls_pairs
    )
    ws.add_data_validation(dv)

    # Voorwaardelijke opmaak (numeriek, beperkt tot max_level)
    LEVEL_CF = [
        (1, C["green_light"],  C["green"]),
        (2, C["yellow_light"], C["yellow"]),
        (3, C["orange_light"], C["orange"]),
        (4, C["red_light"],    C["red"]),
        (5, C["purple_light"], C["purple"]),
    ]
    for _, code_col, label_col in cls_pairs:
        cl = get_column_letter(code_col); ll = get_column_letter(label_col)
        for lvl, bg, fg in [x for x in LEVEL_CF if x[0] <= max_level]:
            formula = f"{cl}{DATA_START}={lvl}"
            for col_ltr in [cl, ll]:
                rng = f"{col_ltr}{DATA_START}:{col_ltr}{DATA_END}"
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[formula],
                    fill=PatternFill("solid", fgColor=bg),
                    font=Font(name="Calibri", bold=True, size=10, color=fg),
                ))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Config
# Bevat de instellingen die alle andere bladen beïnvloeden:
#   - Organisatienaam (D5) en dienst/entiteit (D6)
#   - Taalinstelling (D9): NL / FR / EN — LANG_CELL verwijst hier naartoe
#   - Versie-metadata (read-only): versie, laatste update, gebruiker
# Kolom D is de invoerkolom; kolom B bevat de labels.
# ══════════════════════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("Config")
no_gridlines(ws_cfg); set_col_widths(ws_cfg, [2, 32, 3, 30, 3, 26, 26])
sheet_title_bar(ws_cfg, t("cfg_title_main"), 7)
ws_cfg.row_dimensions[3].height = 12

section_header(ws_cfg, 4, "cfg_section_general", 5, col_start=2)
label_value_row(ws_cfg, 5, 2, 4, t("cfg_org_label"), "Mijn Overheidsinstelling")
label_value_row(ws_cfg, 6, 2, 4, t("cfg_dept_label"), "")
ws_cfg.row_dimensions[7].height = 8

section_header(ws_cfg, 8, "cfg_section_display", 5, col_start=2)
lv = label_value_row(ws_cfg, 9, 2, 4, t("cfg_lang_label"), "NL",
                     val_bold=True, val_color="blue_mid", val_size=14, dropdown='"NL,FR,EN"')
lv.alignment = align("center", "center")
hint_row(ws_cfg, 10, 5, "cfg_lang_hint",        col_start=2)
hint_row(ws_cfg, 11, 5, "cfg_lang_change_hint", col_start=2)
ws_cfg.row_dimensions[12].height = 8

section_header(ws_cfg, 13, "cfg_section_meta", 5, col_start=2)
label_value_row(ws_cfg, 14, 2, 4, t("cfg_version_label"),    "0.3", editable=False)
label_value_row(ws_cfg, 15, 2, 4, t("cfg_updated_label"),    NOW_STR, editable=False)
label_value_row(ws_cfg, 16, 2, 4, t("cfg_updated_by_label"), USERNAME)
ws_cfg.row_dimensions[17].height = 8

# ── Assurance niveau entiteit (rijen 18–21) ───────────────────────────────────
# Config!D19 bevat het gekozen assurance niveau (Basic / Important / Essential).
# De Streamlit-app leest deze cel om te bepalen welke maatregelen getoond worden.
ws_cfg.row_dimensions[18].height = 20
ws_cfg.merge_cells("B18:F18")
c = ws_cfg.cell(row=18, column=2, value="Assurance niveau entiteit")
c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=9, color=C["blue_light"])
c.alignment = align("left", "center", indent=1)

label_value_row(ws_cfg, 19, 2, 4, "Assurance niveau", "",
                val_bold=True, val_color="blue_mid", val_size=12,
                dropdown='"Basic,Important,Essential"')

ws_cfg.row_dimensions[20].height = 18
ws_cfg.merge_cells("B20:F20")
c = ws_cfg.cell(row=20, column=2,
    value="Kies het assurance niveau dat de entiteit wenst te behalen. "
          "Dit bepaalt welke maatregelen in het risicobeheerplan getoond worden.")
c.font = Font(name="Calibri", size=9, color=C["subtext"], italic=True)
c.alignment = align("left", "center", indent=1)

ws_cfg.row_dimensions[21].height = 8

# Tips (verschoven van rijen 18–20 naar rijen 22–24)
ws_cfg.row_dimensions[22].height = 20; ws_cfg.merge_cells("B22:F22")
c = ws_cfg.cell(row=22, column=2, value=t("ui_tips"))
c.fill = fill("accent_light"); c.font = Font(name="Calibri", bold=True, size=9, color=C["accent"])
c.alignment = align("left", "center", indent=1)
for tip_r, tip_k in [(23, "cfg_save_hint"), (24, "cfg_lang_change_hint")]:
    ws_cfg.row_dimensions[tip_r].height = 20; ws_cfg.merge_cells(f"B{tip_r}:F{tip_r}")
    c = ws_cfg.cell(row=tip_r, column=2, value=t(tip_k))
    c.fill = fill("accent_light"); c.font = Font(name="Calibri", size=9, color=C["grey_dark"])
    c.alignment = align("left", "center", indent=1)

quicklinks_block(ws_cfg, 25,
    [("ui_goto_info",    "Info!A1"),
     ("ui_goto_proc",   "Processes!A1"),
     ("ui_goto_assets", "Information Assets!A1"),
     ("ui_goto_dep",    "Dependent Assets!A1"),
     ("ui_goto_verant", "Responsible Persons!A1"),
     ("ui_goto_import", "Import & Export!A1"),
     ("ui_goto_ref",    "Reference Values!A1")], 5, col_start=2)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Info
# ══════════════════════════════════════════════════════════════════════════════
ws_info = wb.create_sheet("Info")
no_gridlines(ws_info); set_col_widths(ws_info, [2, 22, 4, 58, 8])
sheet_title_bar(ws_info, t("info_title"), 5)
ws_info.row_dimensions[3].height = 12

def info_section(ws, row, key): section_header(ws, row, key, 4, col_start=2)
def info_text(ws, row, key, height=52, bg="white"):
    ws.row_dimensions[row].height = height; ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value=t(key))
    c.fill = fill(bg); c.font = font(10, color="text")
    c.alignment = align("left", "top", wrap=True, indent=1); c.border = border_all()

info_section(ws_info, 4, "info_section_doel");   info_text(ws_info, 5, "info_doel_txt", 60)
ws_info.row_dimensions[6].height = 8
info_section(ws_info, 7, "info_section_scope");  info_text(ws_info, 8, "info_scope_txt", 44)
ws_info.row_dimensions[9].height = 8
info_section(ws_info, 10, "info_section_modules")
for i, key in enumerate(["info_mod_proc", "info_mod_assets", "info_mod_verant", "info_mod_import"], 11):
    info_text(ws_info, i, key, 28, "blue_xlight")
ws_info.row_dimensions[15].height = 8
info_section(ws_info, 16, "info_section_guide")
for i, key in enumerate(["info_guide_1","info_guide_2","info_guide_3","info_guide_4"], 17):
    info_text(ws_info, i, key, 28, "grey_light")
ws_info.row_dimensions[21].height = 8
info_section(ws_info, 22, "info_section_version")
for i, key in enumerate(["info_v_tool","info_v_status","info_v_frameworks"], 23):
    ws_info.row_dimensions[i].height = 22; ws_info.merge_cells(f"B{i}:D{i}")
    c = ws_info.cell(row=i, column=2, value=t(key))
    c.fill = fill("grey_light"); c.font = font(10, color="subtext")
    c.alignment = align("left", "center", indent=1); c.border = border_all()

quicklinks_block(ws_info, 26,
    [("ui_goto_config",  "Config!A1"),
     ("ui_goto_proc",   "Processes!A1"),
     ("ui_goto_assets", "Information Assets!A1"),
     ("ui_goto_dep",    "Dependent Assets!A1"),
     ("ui_goto_verant", "Responsible Persons!A1"),
     ("ui_goto_import", "Import & Export!A1"),
     ("ui_goto_ref",    "Reference Values!A1")], 4, col_start=2)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Processen
# Kolomindeling (12 kolommen):
#   A(1) = ID (formule P001-P100)    B(2)  = Naam
#   C(3) = Omschrijving              D(4)  = Eigenaar
#   E(5) = Dienst                    F(6)  = Integriteit (code 1-5)
#   G(7) = Integriteit (label)       H(8)  = Beschikbaarheid (code 1-5)
#   I(9) = Beschikbaarheid (label)   J(10) = Opmerkingen
#   K(11)= Informatieassets (VBA)    L(12) = Afhankelijke assets (VBA)
# Kolom K en L worden ingevuld via de AssetPicker-popup (PROC_SHEET_CODE).
# ══════════════════════════════════════════════════════════════════════════════
ws_proc = wb.create_sheet("Processes")

PROC_WIDTHS = [5, 20, 36, 18, 16,  4, 15,  4, 15,  26,  35,  35]
# cls_pairs: (vertaalsleutel, code-kolom, label-kolom)
# Processen hebben I en A (geen C — vertrouwelijkheid zit op informatieasset-niveau)
PROC_CLS_PAIRS = [
    ("proc_col_integriteit",  6,  7),
    ("proc_col_beschikbaar",  8,  9),
]
HDR_PROC, DS_PROC, DE_PROC, N_PROC = build_cls_sheet(
    ws_proc, "proc_title", "proc_subtitle", "P",
    col_headers_before=[(k, i) for i, k in enumerate(
        ["proc_col_id","proc_col_naam","proc_col_omschrijving","proc_col_eigenaar","proc_col_dienst"], 1)],
    col_headers_after=[("proc_col_opmerkingen", 10), ("proc_col_assets", 11), ("proc_col_dep_assets", 12)],
    widths=PROC_WIDTHS)

# Datarijen: basis stijl
for r in range(DS_PROC, DE_PROC + 1):
    ws_proc.row_dimensions[r].height = 24
    alt = "blue_xlight" if r % 2 == 0 else "white"
    for col_idx in range(1, N_PROC + 1):
        c = ws_proc.cell(row=r, column=col_idx)
        c.fill = fill(alt); c.font = font(10, color="text")
        c.alignment = align("left", "center", wrap=True); c.border = border_all()
    id_c = ws_proc.cell(row=r, column=1)
    # ID-formule: toont "P001" t/m "P100" als kolom B niet leeg is.
    # TEXT(...,"P000") formatteert het volgnummer als 3-cijferig met voorloopnullen.
    id_c.value = f'=IF(B{r}="","",TEXT({r - DS_PROC + 1},"P000"))'
    id_c.font = Font(name="Calibri", size=9, bold=True, color=C["subtext"])
    id_c.alignment = align("center", "center")

add_cls_pairs(ws_proc, PROC_CLS_PAIRS, HDR_PROC, DS_PROC, DE_PROC)
ws_proc.freeze_panes = f"A{DS_PROC}"
ws_proc.auto_filter.ref = f"A{HDR_PROC}:{get_column_letter(N_PROC)}{HDR_PROC}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Informatieassets
# Kolomindeling (9 kolommen):
#   A(1) = ID (formule A001-A100)    B(2)  = Naam
#   C(3) = Omschrijving              D(4)  = Eigenaar
#   E(5) = Dienst                    F(6)  = Confidentialiteit (code 1-5)
#   G(7) = Confidentialiteit (label) H(8)  = Opmerkingen
#   I(9) = Gebruikt in processen (VBA, read-only grijs)
# Informatieassets dragen alleen C-classificatie (niet I/A — die zit op procesniveau).
# Kolom I wordt berekend door INFO_ASSET_SHEET_CODE (Worksheet_Activate).
# ══════════════════════════════════════════════════════════════════════════════
ws_asset = wb.create_sheet("Information Assets")

ASSET_WIDTHS = [5, 20, 30, 16, 16,  4, 15,  24,  35]
ASSET_CLS_PAIRS = [
    ("asset_col_conf",  6,  7),  # alleen Confidentialiteit voor informatieassets
]
HDR_ASSET, DS_ASSET, DE_ASSET, N_ASSET = build_cls_sheet(
    ws_asset, "asset_title", "asset_subtitle", "A",
    col_headers_before=[(k, i) for i, k in enumerate(
        ["asset_col_id","asset_col_naam","asset_col_omschrijving",
         "asset_col_eigenaar","asset_col_dienst"], 1)],
    col_headers_after=[("asset_col_opmerkingen", 8), ("asset_col_processes", 9)],
    widths=ASSET_WIDTHS)



for r in range(DS_ASSET, DE_ASSET + 1):
    ws_asset.row_dimensions[r].height = 24
    alt = "blue_xlight" if r % 2 == 0 else "white"
    for col_idx in range(1, N_ASSET + 1):
        c = ws_asset.cell(row=r, column=col_idx)
        c.fill = fill(alt); c.font = font(10, color="text")
        c.alignment = align("left", "center", wrap=True); c.border = border_all()
    id_c = ws_asset.cell(row=r, column=1)
    id_c.value = f'=IF(B{r}="","",TEXT({r - DS_ASSET + 1},"A000"))'
    id_c.font = Font(name="Calibri", size=9, bold=True, color=C["subtext"])
    id_c.alignment = align("center", "center")

add_cls_pairs(ws_asset, ASSET_CLS_PAIRS, HDR_ASSET, DS_ASSET, DE_ASSET)
# Kolom I (9) = "Gebruikt in processen" — wordt berekend door VBA bij activering.
# Grijs + cursief signaleert aan de gebruiker dat de cel read-only is.
for r in range(DS_ASSET, DE_ASSET + 1):
    c9 = ws_asset.cell(row=r, column=9)
    c9.fill = fill("grey_light")
    c9.font = Font(name="Calibri", size=9, italic=True, color=C["subtext"])
    c9.alignment = align("left", "top", wrap=True)
ws_asset.freeze_panes = f"A{DS_ASSET}"
ws_asset.auto_filter.ref = f"A{HDR_ASSET}:{get_column_letter(N_ASSET)}{HDR_ASSET}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Afhankelijke assets
# 25 kolommen, gegroepeerd in 5 zones:
#   A-I   (1-9) : basisgegevens (ID, Naam, Beschrijving, Eigenaar, Dienst,
#                  Overarching, Opmerkingen, Processen*, Info-assets*)
#   J-O (10-15) : Security Requirements (VBA-berekend, read-only):
#                  C-req(10+11), I-req(12+13), A-req(14+15)
#   P-U (16-21) : Security Objectives (editeerbaar door gebruiker):
#                  C-obj(16+17), I-obj(18+19), A-obj(20+21)
#   V-X (22-24) : Gap Analyse (VBA-berekend): C-gap, I-gap, A-gap
#   Y   (25)    : Commentaar (vrij tekstveld)
# * Kolommen H(8) en I(9) = VBA-berekend (DEP_ASSET_SHEET_CODE)
# Col F (6) = Overarching: niet-leeg (✔) = deze DA is van toepassing op alle processen.
# ══════════════════════════════════════════════════════════════════════════════
ws_dep = wb.create_sheet("Dependent Assets")

# Kolombreedtes voor alle 25 kolommen
DEP_WIDTHS = [5, 20, 30, 16, 16, 10, 24, 35, 35, 4, 15, 4, 15, 4, 15, 4, 15, 4, 15, 4, 15, 35, 35, 35, 35]
DEP_CLS_PAIRS = [
    ("dep_col_conf",  10, 11),
    ("dep_col_integ", 12, 13),
    ("dep_col_avail", 14, 15),
]
HDR_DEP, DS_DEP, DE_DEP, N_DEP = build_cls_sheet(
    ws_dep, "dep_title", "dep_subtitle", "D",
    col_headers_before=[(k, i) for i, k in enumerate(
        ["dep_col_id","dep_col_naam","dep_col_omschrijving",
         "dep_col_eigenaar","dep_col_dienst","dep_col_overarching",
         "dep_col_opmerkingen","dep_col_processes","dep_col_linked_info"], 1)],
    col_headers_after=[],
    widths=DEP_WIDTHS)

# ── Groepskoppen rij 4: kleurgecodeerde zone-aanduiding ─────────────────────
# Rij 4 staat boven de kolomkoppen (rij 5) en geeft per zone aan wat de functie is.
# Paars = Security Requirements (VBA-berekend, niet editeerbaar)
# Navy  = Security Objectives (editeerbaar door gebruiker)
# Rood  = Gap Analyse (VBA-berekend)
ws_dep.row_dimensions[4].height = 22
ws_dep.merge_cells("J4:O4")
c = ws_dep.cell(row=4, column=10, value=t("dep_sec_req_header"))
c.fill = fill("purple"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
c.alignment = align("center", "center"); c.border = border_all("purple")

ws_dep.merge_cells("P4:U4")
c = ws_dep.cell(row=4, column=16, value=t("dep_sec_obj_header"))
c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
c.alignment = align("center", "center"); c.border = border_all("navy")

ws_dep.merge_cells("V4:X4")
c = ws_dep.cell(row=4, column=22, value=t("dep_gap_header"))
c.fill = fill("red"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
c.alignment = align("center", "center"); c.border = border_all("red")

for r in range(DS_DEP, DE_DEP + 1):
    ws_dep.row_dimensions[r].height = 24
    alt = "blue_xlight" if r % 2 == 0 else "white"
    for col_idx in range(1, N_DEP + 1):
        c = ws_dep.cell(row=r, column=col_idx)
        c.fill = fill(alt); c.font = font(10, color="text")
        c.alignment = align("left", "center", wrap=True); c.border = border_all()
    id_c = ws_dep.cell(row=r, column=1)
    id_c.value = f'=IF(B{r}="","","D"&TEXT({r - DS_DEP + 1},"000"))'
    id_c.font = Font(name="Calibri", size=9, bold=True, color=C["subtext"])
    id_c.alignment = align("center", "center")

# Col 6 = Overarching — centre, groen lettertype als hint
for r in range(DS_DEP, DE_DEP + 1):
    c6 = ws_dep.cell(row=r, column=6)
    c6.alignment = align("center", "center")
    c6.font = Font(name="Calibri", size=11, color=C["green"])

# Cols 8-9 = computed door VBA — grijs/italic
for r in range(DS_DEP, DE_DEP + 1):
    for col_idx in [8, 9]:
        cd = ws_dep.cell(row=r, column=col_idx)
        cd.fill = fill("grey_light")
        cd.font = Font(name="Calibri", size=9, italic=True, color=C["subtext"])
        cd.alignment = align("left", "top", wrap=True)

# Security Requirements: code+label paren via add_cls_pairs (CF + formules)
add_cls_pairs(ws_dep, [("dep_col_conf",  10, 11)], HDR_DEP, DS_DEP, DE_DEP)
add_cls_pairs(ws_dep, [("dep_col_integ", 12, 13)], HDR_DEP, DS_DEP, DE_DEP)
add_cls_pairs(ws_dep, [("dep_col_avail", 14, 15)], HDR_DEP, DS_DEP, DE_DEP)

# Override kolomkoppen rij 5 naar paarse stijl (add_cls_pairs zette navy)
for code_col in [10, 12, 14]:
    c = ws_dep.cell(row=HDR_DEP, column=code_col)
    c.fill = fill("purple"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
    c.border = border_all("purple")

# Code-cols Security Requirements → grijs (VBA vult automatisch in, niet bedoeld voor handmatige invoer)
for r in range(DS_DEP, DE_DEP + 1):
    for col_idx in [10, 12, 14]:
        c = ws_dep.cell(row=r, column=col_idx)
        c.fill = fill("grey_light")
        c.font = Font(name="Calibri", size=9, bold=True, italic=True, color=C["subtext"])
        c.alignment = align("center", "center")

# Security Objectives: 3×code+label paren (editeerbaar, navy — zelfde layout als requirements)
add_cls_pairs(ws_dep, [("dep_col_conf",  16, 17)], HDR_DEP, DS_DEP, DE_DEP)
add_cls_pairs(ws_dep, [("dep_col_integ", 18, 19)], HDR_DEP, DS_DEP, DE_DEP)
add_cls_pairs(ws_dep, [("dep_col_avail", 20, 21)], HDR_DEP, DS_DEP, DE_DEP)

# Gap Analyse-kolommen (22-24) — niet editeerbaar, VBA-berekend
for col_idx, key in [(22, "dep_col_gap_conf"), (23, "dep_col_gap_integ"), (24, "dep_col_gap_avail")]:
    c = ws_dep.cell(row=HDR_DEP, column=col_idx, value=t(key))
    c.fill = fill("red"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
    c.alignment = align("center", "center", wrap=True); c.border = border_all("red")
for r in range(DS_DEP, DE_DEP + 1):
    for col_idx in [22, 23, 24]:
        c = ws_dep.cell(row=r, column=col_idx)
        c.fill = fill("grey_light")
        c.font = Font(name="Calibri", size=9, italic=True, color=C["subtext"])
        c.alignment = align("left", "top", wrap=True)
        c.border = border_all()

# CF: gap-cel niet leeg → rood (er is een tekort)
for gap_col in [22, 23, 24]:
    cl = get_column_letter(gap_col)
    rng = f"{cl}{DS_DEP}:{cl}{DE_DEP}"
    ws_dep.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{cl}{DS_DEP}<>""'],
        fill=PatternFill("solid", fgColor=C["red_light"]),
        font=Font(name="Calibri", bold=True, size=9, color=C["red"]),
    ))

# Commentaar-kolom (25) — vrij tekstveld
c = ws_dep.cell(row=HDR_DEP, column=25, value=t("dep_col_obj_commentaar"))
c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")
for r in range(DS_DEP, DE_DEP + 1):
    c = ws_dep.cell(row=r, column=25)
    c.alignment = align("left", "top", wrap=True)

# Celbeveiliging: het blad is beveiligd (ws_dep.protection.sheet = True).
# Standaard zijn alle cellen vergrendeld; editeerbare cellen krijgen
# Protection(locked=False) zodat de gebruiker ze wél kan aanpassen.
# Editeerbaar: B-G (2-7) = basisgegevens incl. Overarching + Opmerkingen
#              P-U (16-21) = Security Objectives (door gebruiker in te vullen)
#              Y (25)      = vrij commentaarveld
# Vergrendeld (VBA-berekend of read-only): H-I (8-9) koppelingen, J-O (10-15) requirements, V-X (22-24) gap
EDITABLE_COLS = list(range(2, 8)) + list(range(16, 22)) + [25]
for r in range(DS_DEP, DE_DEP + 1):
    for col_idx in EDITABLE_COLS:
        ws_dep.cell(row=r, column=col_idx).protection = Protection(locked=False)

ws_dep.protection.sheet = True
ws_dep.protection.autoFilter = False   # filteren toegestaan
ws_dep.protection.sort = False         # sorteren toegestaan

ws_dep.freeze_panes = f"A{DS_DEP}"
ws_dep.auto_filter.ref = f"A{HDR_DEP}:{get_column_letter(N_DEP)}{HDR_DEP}"


def build_rarm(ws):
    """
    Bouwt de statische structuur van de RARM-sheet (Risk Assessment & Remediation Matrix).

    Opbouw:
      Rij 1  : Titelbalk (gemerged over alle kolommen)
      Rij 2  : Kolomkoppen — A=Control ID, B=Richtlijn, C=Assurance, D=Sleutelmaatregel,
               E=# Aangevinkt, F+ = placeholder DA-kolommen
      Rij 3  : Kwetsbaarheden-rij
      Rij 4  : C Objectief per DA (confidentialiteit)
      Rij 5  : I Objectief per DA (integriteit)
      Rij 6  : A Objectief per DA (beschikbaarheid)
      Rij 7  : Aangevinkt-teller per DA (COUNTIF ✔ in controlrijen)
      Rij 8+ : CyFun 2025 ESSENTIAL controls
    """
    ROW_DA       = 2
    ROW_VULN     = 3
    ROW_C_OBJ    = 4
    ROW_I_OBJ    = 5
    ROW_A_OBJ    = 6
    ROW_COUNT    = 7
    ROW_DATA     = 8
    COL_ID       = 1
    COL_TITLE    = 2
    COL_ASS      = 3
    COL_KM       = 4   # Sleutelmaatregel (key measure)
    COL_COUNT    = 5   # # Aangevinkt (COUNTIF per control row)
    COL_DA_START = 6
    N_TEMPLATE   = 10

    CYFUN_TABS = ["GOVERN", "IDENTIFY", "PROTECT", "DETECT", "RESPOND", "RECOVER"]
    ASS_STYLE  = {"Basic": ("green_light","green"), "Important": ("yellow_light","yellow"),
                  "Essential": ("red_light","red")}
    LAST_COL = get_column_letter(COL_DA_START + N_TEMPLATE - 1)

    no_gridlines(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 18   # Sleutelmaatregel
    ws.column_dimensions["E"].width = 14   # # Aangevinkt
    for k in range(N_TEMPLATE):
        ws.column_dimensions[get_column_letter(COL_DA_START + k)].width = 22

    # ── Controls laden uit CYFUN_SRC ──────────────────────────────────────────
    controls = []
    if CYFUN_SRC.exists():
        src_wb = load_workbook(str(CYFUN_SRC), read_only=True, data_only=True)
        for tab in CYFUN_TABS:
            if tab not in src_wb.sheetnames:
                continue
            for row in src_wb[tab].iter_rows(min_row=3, values_only=True):
                if row[5] is None:
                    continue
                assurance = str(row[4]).strip() if row[4] is not None else ""
                req_text  = str(row[5]).strip().replace("\n", " ").replace("\t", "")
                parts     = req_text.split(":", 1)
                req_id    = parts[0].strip()
                req_title = parts[1].strip() if len(parts) > 1 else req_text
                is_km     = row[2] is not None and str(row[2]).strip() != ""
                controls.append((req_id, req_title, assurance, is_km))
        src_wb.close()

    # ── Rij 1: titelrij ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    c.value = "RARM — Risk Assessment & Remediation Matrix"
    c.fill = fill("navy"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[1].height = 34

    # ── Rij 2: kolomkoppen ────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 36
    for col_idx, label in [(COL_ID, "Control ID"), (COL_TITLE, "Richtlijn"),
                           (COL_ASS, "Assurance"), (COL_KM, "Sleutelmaatregel"),
                           (COL_COUNT, "# Aangevinkt")]:
        c = ws.cell(row=ROW_DA, column=col_idx, value=label)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")
    for k in range(N_TEMPLATE):
        c = ws.cell(row=ROW_DA, column=COL_DA_START + k)
        c.fill = fill("yellow_light"); c.font = font(10, bold=True)
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

    # ── Rij 3: kwetsbaarheden-rij ─────────────────────────────────────────────
    ws.row_dimensions[ROW_VULN].height = 28
    for col_idx in [COL_ID, COL_TITLE, COL_ASS, COL_KM, COL_COUNT]:
        c = ws.cell(row=ROW_VULN, column=col_idx)
        c.fill = fill("grey_light"); c.border = border_all("grey_border")
    for k in range(N_TEMPLATE):
        c = ws.cell(row=ROW_VULN, column=COL_DA_START + k)
        c.fill = fill("blue_xlight"); c.font = font(9, italic=True, color="subtext")
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

    # ── Rijen 4-7: CIA objectief + aangevinkt teller ──────────────────────────
    cia_rows = [
        (ROW_C_OBJ, "C Objectief", "blue_xlight"),
        (ROW_I_OBJ, "I Objectief", "blue_xlight"),
        (ROW_A_OBJ, "A Objectief", "blue_xlight"),
        (ROW_COUNT, "Aangevinkt",  "green_light"),
    ]
    for row_idx, label, bg in cia_rows:
        ws.row_dimensions[row_idx].height = 22
        c = ws.cell(row=row_idx, column=COL_ID, value=label)
        c.fill = fill(bg); c.font = font(9, bold=True)
        c.alignment = align("left", "center", indent=1); c.border = border_all("grey_border")
        for col_idx in [COL_TITLE, COL_ASS, COL_KM, COL_COUNT]:
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill(bg); c.border = border_all("grey_border")
        for k in range(N_TEMPLATE):
            col_da = COL_DA_START + k
            c = ws.cell(row=row_idx, column=col_da)
            c.fill = fill(bg); c.border = border_all("grey_border")
            c.alignment = align("center", "center")
            # Rij 7 (ROW_COUNT) krijgt een COUNTIF-formule per DA-kolom
            if row_idx == ROW_COUNT:
                col_letter = get_column_letter(col_da)
                c.value = f'=COUNTA({col_letter}{ROW_DATA}:{col_letter}1048576)'
                c.font = font(9, bold=True)

    # ── Thick border: scheidt vaste meta-kolommen (A-D) van rest (E+) ─────────
    thick = Side(border_style="medium", color=C["navy"])
    for row_idx in range(1, ROW_DATA + len(controls) + 1):
        cell = ws.cell(row=row_idx, column=COL_COUNT)
        b = cell.border
        cell.border = Border(left=thick, right=b.right, top=b.top, bottom=b.bottom)
        cell2 = ws.cell(row=row_idx, column=COL_DA_START)
        b2 = cell2.border
        cell2.border = Border(left=thick, right=b2.right, top=b2.top, bottom=b2.bottom)

    ws.freeze_panes = "G8"

    # ── Rij 8+: controlrijen ─────────────────────────────────────────────────
    for i, (req_id, req_title, assurance, is_km) in enumerate(controls):
        row_idx = ROW_DATA + i
        bg_key, fg_key = ASS_STYLE.get(assurance, ("white", "text"))
        row_bg = "grey_light" if i % 2 == 0 else "white"

        c = ws.cell(row=row_idx, column=COL_ID, value=req_id)
        c.fill = fill(bg_key); c.font = font(9, bold=True, color=fg_key)
        c.alignment = align("center", "center"); c.border = border_all("grey_border")

        c = ws.cell(row=row_idx, column=COL_TITLE, value=req_title)
        c.fill = fill(row_bg); c.font = font(9)
        c.alignment = align("left", "top", wrap=True); c.border = border_all("grey_border")
        ws.row_dimensions[row_idx].height = 30

        c = ws.cell(row=row_idx, column=COL_ASS, value=assurance)
        c.fill = fill(bg_key); c.font = font(9, bold=True, color=fg_key)
        c.alignment = align("center", "center"); c.border = border_all("grey_border")

        c = ws.cell(row=row_idx, column=COL_KM)
        c.fill = fill(row_bg)
        c.value = "✔" if is_km else ""
        c.font = font(10, bold=True, color="green")
        c.alignment = align("center", "center"); c.border = border_all("grey_border")

        # # Aangevinkt: COUNTIF over alle DA-kolommen in deze rij
        last_possible = get_column_letter(COL_DA_START + N_TEMPLATE + 90)
        c = ws.cell(row=row_idx, column=COL_COUNT)
        c.value = f'=COUNTA({get_column_letter(COL_DA_START)}{row_idx}:{last_possible}{row_idx})'
        c.fill = fill(row_bg); c.font = font(9, bold=True)
        c.alignment = align("center", "center"); c.border = border_all("grey_border")

        for k in range(N_TEMPLATE):
            c = ws.cell(row=row_idx, column=COL_DA_START + k)
            c.fill = fill(row_bg); c.border = border_all("grey_border")
            c.alignment = align("center", "center")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: RARM
# ══════════════════════════════════════════════════════════════════════════════
ws_rarm = wb.create_sheet("RARM")
build_rarm(ws_rarm)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Verantwoordelijken
# ══════════════════════════════════════════════════════════════════════════════
ws_ver = wb.create_sheet("Responsible Persons")
no_gridlines(ws_ver)
VER_COLS = [("verant_col_id",6),("verant_col_naam",20),("verant_col_voornaam",18),
            ("verant_col_functie",24),("verant_col_dienst",22),("verant_col_email",28),
            ("verant_col_tel",16),("verant_col_rol",22),("verant_col_opmerkingen",28)]
N_VER = len(VER_COLS)
set_col_widths(ws_ver, [w for _, w in VER_COLS])
sheet_title_bar(ws_ver, t("verant_title"), N_VER)
ws_ver.row_dimensions[3].height = 22; ws_ver.merge_cells(f"A3:{get_column_letter(N_VER)}3")
c = ws_ver["A3"]; c.value = t("verant_subtitle")
c.fill = fill("blue_xlight"); c.font = font(10, italic=True, color="grey_dark")
c.alignment = align("center", "center")
ws_ver.row_dimensions[4].height = 8

VER_HDR = 5; ws_ver.row_dimensions[VER_HDR].height = 44
for col_idx, (key, _) in enumerate(VER_COLS, 1):
    c = ws_ver.cell(row=VER_HDR, column=col_idx, value=t(key))
    c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
    c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

VER_START, VER_END = 6, 55
dv_rol = DataValidation(type="list", formula1=f'=INDIRECT("roles_"&{LANG_CELL})',
                        showDropDown=False, allow_blank=True)
dv_rol.sqref = f"H{VER_START}:H{VER_END}"; ws_ver.add_data_validation(dv_rol)

for r in range(VER_START, VER_END + 1):
    ws_ver.row_dimensions[r].height = 24
    alt = "blue_xlight" if r % 2 == 0 else "white"
    for col_idx in range(1, N_VER + 1):
        c = ws_ver.cell(row=r, column=col_idx)
        c.fill = fill(alt); c.font = font(10, color="text")
        c.alignment = align("left", "center", wrap=True); c.border = border_all()
    id_c = ws_ver.cell(row=r, column=1)
    id_c.value = f'=IF(B{r}="","",TEXT({r - VER_START + 1},"V00"))'
    id_c.font = Font(name="Calibri", size=9, bold=True, color=C["subtext"])
    id_c.alignment = align("center", "center")

ws_ver.freeze_panes = f"A{VER_START}"
ws_ver.auto_filter.ref = f"A{VER_HDR}:{get_column_letter(N_VER)}{VER_HDR}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Import & Export
# ══════════════════════════════════════════════════════════════════════════════
ws_ie = wb.create_sheet("Import & Export")
no_gridlines(ws_ie); set_col_widths(ws_ie, [2, 26, 2, 56, 8])
sheet_title_bar(ws_ie, t("ie_title"), 5)
ws_ie.row_dimensions[3].height = 22; ws_ie.merge_cells("A3:E3")
c = ws_ie["A3"]; c.value = t("ie_subtitle")
c.fill = fill("blue_xlight"); c.font = font(10, italic=True, color="grey_dark")
c.alignment = align("center", "center")
ws_ie.row_dimensions[4].height = 10

def ie_action_block(ws, row, btn_key, desc_key, btn_color="blue_mid"):
    """
    Maakt een actieblok voor de Import & Export-sheet bestaande uit:
      - Rij (row)   : gekleurde knop-placeholder (tekstcel met knoopuiterlijk)
      - Rij (row+1) : beschrijvingstekst (gemerged over B-D)
      - Rij (row+2) : lege spacer-rij

    De knop-placeholder wordt na VBA-injectie vervangen door een echte
    Shape-knop via add_button() in de win32com-sectie.

    Parameters
    ----------
    ws        : openpyxl.Worksheet
    row       : int — startrij van het actieblok
    btn_key   : str — vertaalsleutel voor de knoopkapsel
    desc_key  : str — vertaalsleutel voor de beschrijvingstekst
    btn_color : str — kleursleutel voor de knop (standaard blue_mid)

    Geeft terug
    -----------
    int — volgende beschikbare startrij (row + 3)
    """
    ws.row_dimensions[row].height = 34
    btn_cell = ws.cell(row=row, column=2)
    btn_cell.value = t(btn_key)
    btn_cell.fill = fill(btn_color); btn_cell.font = Font(name="Calibri", bold=True, size=11, color=C["white"])
    btn_cell.alignment = align("center", "center"); btn_cell.border = border_all(btn_color)

    ws.row_dimensions[row + 1].height = 44
    ws.merge_cells(f"B{row+1}:D{row+1}")
    desc = ws.cell(row=row + 1, column=2, value=t(desc_key))
    desc.fill = fill("grey_light"); desc.font = font(10, color="subtext", italic=True)
    desc.alignment = align("left", "top", wrap=True, indent=1); desc.border = border_all()
    ws.row_dimensions[row + 2].height = 10
    return row + 3

section_header(ws_ie, 5, "ie_section_import", 4, col_start=2)
cur = 6
cur = ie_action_block(ws_ie, cur, "ie_import_all_btn",   "ie_import_all_desc")
cur = ie_action_block(ws_ie, cur, "ie_import_ia_btn",    "ie_import_ia_desc")
cur = ie_action_block(ws_ie, cur, "ie_import_da_btn",    "ie_import_da_desc")
cur = ie_action_block(ws_ie, cur, "ie_import_proc_btn",  "ie_import_proc_desc")
cur = ie_action_block(ws_ie, cur, "ie_import_kwets_btn", "ie_import_kwets_desc")
cur = ie_action_block(ws_ie, cur, "ie_import_links_btn", "ie_import_links_desc", btn_color="orange")

section_header(ws_ie, cur, "ie_section_export", 4, col_start=2)
cur += 1

section_header(ws_ie, cur, "ie_section_info", 4, col_start=2)
cur += 1

# Tabelstructuur bronbestand tonen als referentie
def ie_table_info(ws, row, tbl_key, cols_str):
    ws.row_dimensions[row].height = 22; ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value=t(tbl_key))
    c.fill = fill("navy"); c.font = Font(name="Calibri", bold=True, size=10, color=C["white"])
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row+1].height = 18; ws.merge_cells(f"B{row+1}:D{row+1}")
    c2 = ws.cell(row=row+1, column=2, value=t("ie_col_hint"))
    c2.fill = fill("grey_light"); c2.font = font(9, italic=True, color="subtext")
    c2.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row+2].height = 52; ws.merge_cells(f"B{row+2}:D{row+2}")
    c3 = ws.cell(row=row+2, column=2, value=cols_str)
    c3.fill = fill("grey_light"); c3.font = Font(name="Courier New", size=9, color=C["text"])
    c3.alignment = align("left", "top", wrap=True, indent=1); c3.border = border_all()
    return row + 4

proc_cols = ("naam / name / nom\nomschrijving / description\neigenaar / owner / propriétaire\n"
             "dienst / department / service\nintegriteit / integrity / intégrité\n"
             "beschikbaarheid / availability / disponibilité\nconfidentialiteit / confidentiality / confidentialité")
asset_cols = ("naam / name / nom\ntype / asset type\nomschrijving / description\n"
              "eigenaar / owner / propriétaire\ndienst / department / service\n"
              "integriteit / integrity / intégrité\nbeschikbaarheid / availability / disponibilité\n"
              "confidentialiteit / confidentiality / confidentialité")

cur = ie_table_info(ws_ie, cur, "ie_table_proc",  proc_cols)
cur = ie_table_info(ws_ie, cur, "ie_table_asset", asset_cols)

ws_ie.row_dimensions[cur].height = 22; ws_ie.merge_cells(f"B{cur}:D{cur}")
c = ws_ie.cell(row=cur, column=2, value=t("ie_macro_hint"))
c.fill = fill("accent_light"); c.font = Font(name="Calibri", size=9, color=C["grey_dark"], italic=True)
c.alignment = align("left", "center", indent=1)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Referentiewaarden  (naar achter geschoven)
# ══════════════════════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("Reference Values")
no_gridlines(ws_ref); set_col_widths(ws_ref, [2, 18, 52, 8, 8, 8]); N_REF = 6
sheet_title_bar(ws_ref, t("ref_title"), N_REF)
ws_ref.merge_cells("A3:F3"); ws_ref.row_dimensions[3].height = 30
c = ws_ref["A3"]; c.value = t("ref_subtitle")
c.fill = fill("blue_xlight"); c.font = font(10, italic=True, color="grey_dark")
c.alignment = align("center", "center", wrap=True)
ws_ref.row_dimensions[4].height = 10

def ref_dim_table(ws, start_row, section_key, subtitle_key, desc_keys, n_cols=6):
    """
    Bouwt een classificatieschaalblok voor één CIA-dimensie in de Reference Values-sheet.
    Structuur per blok:
      - Sectiekop (marineblauwe balk)
      - Subtitelrij (grijs)
      - Kopregel: Niveau | Benaming | Omschrijving
      - 5 datarijen (1 per classificatieniveau, gekleurd volgens LEVEL_FILLS/LEVEL_FONTS)

    Parameters
    ----------
    ws           : openpyxl.Worksheet
    start_row    : int — startrij van het blok
    section_key  : str — vertaalsleutel voor de sectietitel
    subtitle_key : str — vertaalsleutel voor de subtitelrij
    desc_keys    : list[str] — lijst van 5 vertaalsleutels (niveau 1 t/m 5)
    n_cols       : int — breedte van het blok in kolommen

    Geeft terug
    -----------
    int — startrij voor het volgende blok (= row na het blok + 2 spacer)
    """
    section_header(ws, start_row, section_key, n_cols)
    row = start_row + 1
    ws.row_dimensions[row].height = 22; ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=t(subtitle_key))
    c.fill = fill("grey_light"); c.font = font(10, italic=True, color="subtext")
    c.alignment = align("left", "center", indent=1)
    row += 1
    ws.row_dimensions[row].height = 26
    for col, key in enumerate(["ref_level", "ref_label", "ref_description"], 1):
        c = ws.cell(row=row, column=col, value=t(key))
        c.fill = fill("blue_mid"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")
    lvl_keys = ["ref_lvl1_label","ref_lvl2_label","ref_lvl3_label","ref_lvl4_label","ref_lvl5_label"]
    for i, (lvl_key, desc_key, fill_k, font_k) in enumerate(
            zip(lvl_keys, desc_keys, LEVEL_FILLS, LEVEL_FONTS)):
        row += 1; ws.row_dimensions[row].height = 52
        # Nummer
        c = ws.cell(row=row, column=1, value=i + 1)
        c.fill = fill(fill_k); c.font = Font(name="Calibri", bold=True, size=14, color=C[font_k])
        c.alignment = align("center", "center"); c.border = border_all()
        # Benaming
        c = ws.cell(row=row, column=2, value=t(lvl_key))
        c.fill = fill(fill_k); c.font = Font(name="Calibri", bold=True, size=10, color=C[font_k])
        c.alignment = align("center", "center", wrap=True); c.border = border_all()
        # Beschrijving
        ws.merge_cells(f"C{row}:F{row}")
        c = ws.cell(row=row, column=3, value=t(desc_key))
        c.fill = fill(fill_k); c.font = font(10, color=font_k)
        c.alignment = align("left", "center", wrap=True, indent=1); c.border = border_all()
    return row + 2

cur_ref = 5
cur_ref = ref_dim_table(ws_ref, cur_ref, "ref_section_int",   "ref_int_subtitle",
                         ["ref_int_1","ref_int_2","ref_int_3","ref_int_4","ref_int_5"])
cur_ref = ref_dim_table(ws_ref, cur_ref, "ref_section_avail", "ref_avail_subtitle",
                         ["ref_avail_1","ref_avail_2","ref_avail_3","ref_avail_4","ref_avail_5"])
cur_ref = ref_dim_table(ws_ref, cur_ref, "ref_section_conf",  "ref_conf_subtitle",
                         ["ref_conf_1","ref_conf_2","ref_conf_3","ref_conf_4","ref_conf_5"])


# ── Kwetsbaarheden sheet ─────────────────────────────────────────────────────
def _load_kwets_db():
    """
    Laadt kwetsbaarheden en control-links uit de Access-database ten tijde
    van de Python build. Wordt alleen aangeroepen als pyodbc beschikbaar is
    en de Access DB bestaat in data/Import/.

    Geeft terug
    -----------
    tuple (vulns, ctrl_links_2025, unmatched_23ids) waarbij:
      vulns            : list[(id, name, C_bool, I_bool, A_bool)] — kwetsbaarheden
      ctrl_links_2025  : dict{vuln_id: set(norm_2025_req_ids)} — controls per kwetsbaarheid
                         IDs zijn genormaliseerd via _norm_id() voor matching met sheet-koppen
      unmatched_23ids  : set(norm_2023_id) — 2023-controls die niet gemapped zijn naar 2025

    Werking
    -------
    1. Laad kwetsbaarheden uit T - Vulnerabilities
    2. Laad RefNr → 2023 ID mapping uit T - CyFunEssentiel
    3. Laad koppelingstabel LT - Vulnerability to control - fixed
    4. Bouw 2023→2025 reverse mapping via _load_all_mappings()
    5. Zet elke LT-link om: vuln_id → set(norm_2025_req_ids)
       via: ctrl_ref → refnr_map[ctrl_ref] → norm_id → rev_map → norm_2025_id

    Koppelingssleutel: T-Vulnerabilities.Reference (NIET .ID)
    omdat LT.Vulnerability verwijst naar T-Vulnerabilities.Reference.
    """
    DB_PATH = Path(__file__).parent.parent / "data" / "Import" / "MNMTool - SocSec.accdb"
    if not DB_PATH.exists():
        print(f"[kwetsbaarheden] Access DB niet gevonden: {DB_PATH}")
        return [], {}, set()
    try:
        import pyodbc
        conn = pyodbc.connect(
            f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={DB_PATH};"
        )
        cur = conn.cursor()
        cur.execute(
            "SELECT ID, Reference, Vulnerability, Confidentialiteit, Integriteit, Beschikbaarheid "
            "FROM [T - Vulnerabilities] ORDER BY Reference"
        )
        vulns = [(int(r[0]), str(r[2]), bool(r[3]), bool(r[4]), bool(r[5]))
                 for r in cur.fetchall()]
        cur.execute("SELECT RefNr, Requirement FROM [T - CyFunEssentiel]")
        refnr_map = {int(r[0]): str(r[1]) for r in cur.fetchall()}
        cur.execute(
            "SELECT Vulnerability, CyFunControl "
            "FROM [LT -  Vulnerability to control - fixed]"
        )
        raw_links = [(r[0], r[1]) for r in cur.fetchall()
                     if r[0] is not None and r[1] is not None]
        conn.close()
    except Exception as e:
        print(f"[kwetsbaarheden] DB-fout: {e}")
        return [], {}, set()

    # ── Bouw 2023→2025 reverse mapping via MAPPING_SRC ────────────────────────
    # _load_all_mappings() geeft {norm_2025_id → "2023_id"} (both_map)
    # We draaien dit om naar {norm_2023_id → norm_2025_id} zodat we
    # vanuit een 2023-ID direct het 2025-ID kunnen opzoeken.
    # Normalisatie van 2023-IDs: trailing -N → .N (bv. "rs.co-3-2" → "rs.co-3.2")
    # via re.sub zodat de sleutels consistent zijn met NormId23() in VBA.
    rev_map = {}   # norm_2023_id → norm_2025_id
    try:
        both_map, _, _ = _load_all_mappings()
        for norm25, id23 in both_map.items():
            n23 = re.sub(r'-(\d+)$', r'.\1', str(id23).strip().lower())
            # norm25 is reeds genormaliseerd via _norm_id(); gebruik als target-ID
            # voor matching met sheet-koppen in de Kwetsbaarheden-sheet.
            rev_map[n23] = norm25
    except Exception as e:
        print(f"[kwetsbaarheden] mapping-fout: {e}")

    # ── Verwerk LT-links: 2023 ID → 2025 ID via rev_map ──────────────────────
    # Voor elke koppeling (vuln_id, ctrl_ref) in de LT-tabel:
    #   1. ctrl_ref → refnr_map → 2023-ID string
    #   2. normaliseer 2023-ID (re.sub trailing -N → .N)
    #   3. rev_map → 2025-ID (of sla 2023-ID op als fallback / niet gevonden)
    # Koppelingssleutel: LT.Vulnerability verwijst naar T-Vulnerabilities.Reference
    # (NIET .ID — dit is een andere kolom in de Access-tabel!)
    ctrl_links_2025 = {}
    unmatched_23ids = set()
    for vuln_id_raw, ctrl_ref_raw in raw_links:
        vid = int(vuln_id_raw)
        cid = int(ctrl_ref_raw)
        id23 = refnr_map.get(cid)
        if not id23:
            continue
        n23 = re.sub(r'-(\d+)$', r'.\1', str(id23).strip().lower())
        if n23 in rev_map:
            # Sla normalized 2025 ID op voor matching met sheet-headers
            ctrl_links_2025.setdefault(vid, set()).add(rev_map[n23])
        else:
            unmatched_23ids.add(id23)
            # Controleer ook of het ID al in 2025-formaat is (bijv. RefNr 219-227)
            # door het direct op te slaan als norm
            ctrl_links_2025.setdefault(vid, set()).add(n23)

    return vulns, ctrl_links_2025, unmatched_23ids


def build_kwetsbaarheden(ws):
    """
    Statische structuur — data wordt ENKEL via macro geladen (ImportKwetsbaarheden).

    Kolom A = Control ID, kolom B = Richtlijn  (statisch vanuit CYFUN_SRC)
    Kolom C+ = per kwetsbaarheid 1 kolom       (dynamisch, door macro)

    Rij 1  : Titel
    Rij 2  : "Control ID" | "Richtlijn" | [kwetsbaarheid-namen]
    Rij 3  : "C" | "Vertrouwelijkheid"  | [✔ als kwetsbaarheid C-impact heeft]
    Rij 4  : "I" | "Integriteit"        | [✔ als kwetsbaarheid I-impact heeft]
    Rij 5  : "A" | "Beschikbaarheid"    | [✔ als kwetsbaarheid A-impact heeft]
    Rij 6+ : ctrl_id | richtlijn        | [✔ als control remediëring biedt]
    """
    ROW_TITLE = 1
    ROW_VULN  = 2
    ROW_C     = 3
    ROW_I     = 4
    ROW_A     = 5
    ROW_DATA  = 6
    COL_ID    = 1   # A
    COL_TITLE = 2   # B
    CYFUN_TABS = ["GOVERN", "IDENTIFY", "PROTECT", "DETECT", "RESPOND", "RECOVER"]
    ASSURANCE_STYLE = {
        "Basic":     ("green_light",  "green"),
        "Important": ("yellow_light", "yellow"),
        "Essential": ("red_light",    "red"),
    }

    no_gridlines(ws)

    # ── Controls laden uit CYFUN_SRC (statische referentiedata) ──────────────
    controls = []   # [(req_id, req_title, assurance), ...]
    if CYFUN_SRC.exists():
        src_wb = load_workbook(str(CYFUN_SRC), read_only=True, data_only=True)
        for tab in CYFUN_TABS:
            if tab not in src_wb.sheetnames:
                continue
            for row in src_wb[tab].iter_rows(min_row=3, values_only=True):
                if row[5] is None:
                    continue
                assurance = str(row[4]).strip() if row[4] is not None else ""
                req_text  = str(row[5]).strip().replace("\n", " ").replace("\t", "")
                parts     = req_text.split(":", 1)
                req_id    = parts[0].strip()
                req_title = parts[1].strip() if len(parts) > 1 else req_text
                controls.append((req_id, req_title, assurance))
        src_wb.close()

    N_TEMPLATE    = 5   # lege invoerkolommen voor manuele input
    COL_INP_START = 3   # C
    LAST_HDR_COL  = get_column_letter(COL_INP_START + N_TEMPLATE - 1)

    # ── Kolombreedtes ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 65
    for k in range(N_TEMPLATE):
        ws.column_dimensions[get_column_letter(COL_INP_START + k)].width = 14

    # ── Rij 1: titel (gemerged over A t/m laatste invoerkolom) ───────────────
    ws.merge_cells(f"A1:{LAST_HDR_COL}1")
    c = ws["A1"]
    c.value = "Kwetsbaarheden — Remediëring via CyFun 2025 Controls  ·  dubbelklik op een cel om ✔ te plaatsen"
    c.fill = fill("navy"); c.font = font(11, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[ROW_TITLE].height = 34

    # ── Rij 2: "Control ID" / "Richtlijn" + lege invoerkoppen ────────────────
    ws.row_dimensions[ROW_VULN].height = 55
    for col, lbl in [(COL_ID, "Control ID"), (COL_TITLE, "Richtlijn")]:
        c = ws.cell(row=ROW_VULN, column=col, value=lbl)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True)
        c.border = border_all("navy")

    # ── Rijen 3-5: C/I/A rijlabels ───────────────────────────────────────────
    CIA_DEFS = [
        (ROW_C, "C", "Vertrouwelijkheid", "blue_light",   "blue_mid"),
        (ROW_I, "I", "Integriteit",       "green_light",  "green"),
        (ROW_A, "A", "Beschikbaarheid",   "orange_light", "orange"),
    ]
    for row, short, full, bg, fg in CIA_DEFS:
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=COL_ID, value=short)
        c.fill = fill(bg); c.font = Font(name="Calibri", size=9, bold=True, color=C[fg])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all("grey_border")
        c = ws.cell(row=row, column=COL_TITLE, value=full)
        c.fill = fill("white"); c.font = Font(name="Calibri", size=9, color=C["text"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border_all("grey_border")

    # Bevroren venster: rijen 1-5 EN kolommen A-B altijd zichtbaar
    ws.freeze_panes = "C6"

    if not controls:
        ws.cell(row=ROW_DATA, column=COL_ID,
                value="CYFUN_SRC niet gevonden — controleer het pad")
        return []

    # ── Rijen 6+: controls (A = ID, B = richtlijn) — ✔ marks door macro ──────
    _bdr = border_all("grey_border")
    for j, (ctrl_id, ctrl_title, assurance) in enumerate(controls):
        row    = ROW_DATA + j
        row_bg = "grey_light" if j % 2 == 0 else "white"
        bg_key, fg_key = ASSURANCE_STYLE.get(assurance, ("white", "text"))

        c = ws.cell(row=row, column=COL_ID, value=ctrl_id)
        c.fill = fill(bg_key)
        c.font = Font(name="Calibri", size=9, bold=True, color=C[fg_key])
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bdr

        c = ws.cell(row=row, column=COL_TITLE, value=ctrl_title)
        c.fill = fill(row_bg)
        c.font = Font(name="Calibri", size=9, color=C["text"])
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = _bdr

        ws.row_dimensions[row].height = 30

    # ── 5 lege invoerkolommen voor manuele input (C t/m G) ───────────────────
    # Dikke linker scheidingslijn op col C, dunne randen op D-G
    _thin   = Side(style="thin",   color=C["grey_border"])
    _medium = Side(style="medium", color=C["navy"])
    _align_inp = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CIA_COLORS = [("blue_light", ROW_C), ("green_light", ROW_I), ("orange_light", ROW_A)]

    for k in range(N_TEMPLATE):
        col  = COL_INP_START + k
        left = _medium if k == 0 else _thin

        def _inp_bdr(left_side):
            return Border(left=left_side, right=_thin, top=_thin, bottom=_thin)

        # Rij 2: kwetsbaarheid-naam invoerveld (geel = actief invoerveld)
        c = ws.cell(row=ROW_VULN, column=col)
        c.fill = fill("yellow_light")
        c.alignment = _align_inp
        c.border = _inp_bdr(left)

        # Rijen 3-5: CIA-impact invoervelden (gekleurde achtergrond)
        for bg, row in CIA_COLORS:
            c = ws.cell(row=row, column=col)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _inp_bdr(left)

        # Rijen 6+: control-koppeling invoervelden (wisselende achtergrond)
        for j in range(len(controls)):
            r      = ROW_DATA + j
            row_bg = "grey_light" if j % 2 == 0 else "white"
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _inp_bdr(left)

    ws.auto_filter.ref = "A2:B2"
    return []


def build_afwijkingen(ws, afw_rows):
    """
    Bouwt de Afwijkingen-rapportagesheet met controls die in de Access DB als
    remediëring zijn gelinkt aan kwetsbaarheden, maar NIET bestaan in de
    CyFun 2025 ESSENTIAL-set (zgn. "Enkel-2023" controls).

    Deze sheet dient als technische verantwoording: de CISO kan zien welke
    2023-controls niet meer van toepassing zijn in de 2025-versie en of
    er actie nodig is (handmatige herclassificatie, archivering, …).

    Parameters
    ----------
    ws       : openpyxl.Worksheet — het Afwijkingen-werkblad
    afw_rows : list[(vname, cia_str, ctrl_id)] — afwijkende control-koppelingen
               • vname   : naam van de kwetsbaarheid
               • cia_str : betrokken CIA-dimensies (bv. "CI" of "CIA")
               • ctrl_id : het 2023-control-ID dat niet gemapped kon worden
    """
    ROW_TITLE = 1
    ROW_HDR   = 2
    ROW_DATA  = 3

    no_gridlines(ws)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40

    # Titel
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Afwijkingen — CyFun 2023 controls zonder 2025 equivalent"
    c.fill = fill("orange"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[ROW_TITLE].height = 34

    # Kolomkoppen
    headers = ["Kwetsbaarheid", "C / I / A", "Control ID (DB)", "Reden"]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=ROW_HDR, column=col, value=hdr)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True)
        c.border = border_all("navy")
    ws.row_dimensions[ROW_HDR].height = 30

    if not afw_rows:
        c = ws.cell(row=ROW_DATA, column=1,
                    value="Geen afwijkingen — alle DB-controls bestaan in CyFun 2025.")
        c.font = font(9, color="subtext"); c.alignment = align("left", "center")
        return

    for idx, (vname, cia_str, ctrl_id) in enumerate(afw_rows):
        r = ROW_DATA + idx
        row_bg = "grey_light" if idx % 2 == 0 else "white"
        vals = [vname, cia_str, ctrl_id,
                "Control bestaat enkel in CyFun 2023 — geen 2025 equivalent"]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(row_bg)
            c.font = font(9, color="text")
            c.alignment = align("left", "center", wrap=(col == 1))
            c.border = border_all("grey_border")
        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = "A2:D2"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET-BOUWERS: Risicobeheer | Controles | Acties
# ══════════════════════════════════════════════════════════════════════════════

def build_risicobeheer(ws):
    """
    Bouwt de Risicobeheer-sheet: risicobeheerplan per control per dependent asset.

    De Streamlit-app vult deze sheet in via read/write met openpyxl.
    Per rij staat de status van één CyFun-control voor één DA.

    Structuur
    ---------
    Rij 1  : titelrij (gemerged A–H)
    Rij 2  : kolomkoppen
    Rij 3+ : data (één rij per control × DA-combinatie)

    Kolommen
    --------
    A  Control ID      | B  Richtlijn       | C  Assurance  | D  Dependent Asset
    E  Status          | F  Datum           | G  Verantwoordelijke | H  Opmerkingen
    """
    no_gridlines(ws)
    COLS = [
        ("Control ID",          14),
        ("Richtlijn",           60),
        ("Assurance",           13),
        ("Dependent Asset",     25),
        ("Status",              18),
        ("Datum",               14),
        ("Verantwoordelijke",   25),
        ("Opmerkingen",         40),
    ]
    N = len(COLS)
    LAST = get_column_letter(N)

    set_col_widths(ws, [w for _, w in COLS])

    # Titelrij
    ws.merge_cells(f"A1:{LAST}1")
    c = ws["A1"]
    c.value = "Risicobeheer — Status maatregelen per control per afhankelijk asset"
    c.fill = fill("navy"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[1].height = 34

    # Kolomkoppen
    for col, (hdr, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True)
        c.border = border_all("navy")
    ws.row_dimensions[2].height = 30

    # Dropdown: Status (kolom E = 5)
    dv_status = DataValidation(
        type="list",
        formula1='"Uitgevoerd,Gepland,Niet uitvoeren"',
        showDropDown=False, allow_blank=True
    )
    dv_status.sqref = f"E3:E2000"
    ws.add_data_validation(dv_status)

    # AutoFilter
    ws.auto_filter.ref = f"A2:{LAST}2"

    # Freeze panes: rij 1-2 altijd zichtbaar
    ws.freeze_panes = "A3"

    # Placeholder-rij (helpt zien dat de sheet leeg is bij verse build)
    c = ws.cell(row=3, column=1,
                value="(leeg — wordt ingevuld via Streamlit of via ImportGeselecteerdeControls)")
    c.font = font(9, color="subtext", italic=True)
    c.alignment = align("left", "center")


def build_controles(ws):
    """
    Bouwt de Controles-sheet: schema voor periodieke controle van maatregelen.

    Per rij staat één te controleren maatregel voor één DA, met frequentie,
    methode, artefactenlocatie en datumopvolging.

    Structuur
    ---------
    Rij 1  : titelrij
    Rij 2  : kolomkoppen
    Rij 3+ : data

    Kolommen
    --------
    A  DA  | B  Control ID  | C  Maatregel  | D  Frequentie  | E  Methode
    F  Artefacten locatie   | G  Laatste controle | H  Volgende controle | I  Status
    """
    no_gridlines(ws)
    COLS = [
        ("DA",                  25),
        ("Control ID",          14),
        ("Maatregel",           60),
        ("Frequentie",          15),
        ("Methode",             40),
        ("Artefacten locatie",  50),
        ("Laatste controle",    16),
        ("Volgende controle",   16),
        ("Status",              18),
    ]
    N = len(COLS)
    LAST = get_column_letter(N)

    set_col_widths(ws, [w for _, w in COLS])

    # Titelrij
    ws.merge_cells(f"A1:{LAST}1")
    c = ws["A1"]
    c.value = "Periodieke Controles — Schema voor opvolging van maatregelen"
    c.fill = fill("navy"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[1].height = 34

    # Kolomkoppen
    for col, (hdr, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True)
        c.border = border_all("navy")
    ws.row_dimensions[2].height = 30

    # Dropdown: Frequentie (kolom D = 4)
    dv_freq = DataValidation(
        type="list",
        formula1='"Maandelijks,Kwartaal,Halfjaarlijks,Jaarlijks,Ad hoc"',
        showDropDown=False, allow_blank=True
    )
    dv_freq.sqref = "D3:D2000"
    ws.add_data_validation(dv_freq)

    # Dropdown: Status (kolom I = 9)
    dv_status = DataValidation(
        type="list",
        formula1='"OK,Te controleren,Vervallen"',
        showDropDown=False, allow_blank=True
    )
    dv_status.sqref = "I3:I2000"
    ws.add_data_validation(dv_status)

    # AutoFilter + freeze
    ws.auto_filter.ref = f"A2:{LAST}2"
    ws.freeze_panes = "A3"

    # Datumopmaak voor kolommen G en H
    for col_letter in ("G", "H"):
        for r in range(3, 2001):
            ws.cell(row=r, column=ord(col_letter) - 64).number_format = "DD/MM/YYYY"

    # Placeholder
    c = ws.cell(row=3, column=1,
                value="(leeg — voeg controleregels toe via de Streamlit-app)")
    c.font = font(9, color="subtext", italic=True)
    c.alignment = align("left", "center")


def build_acties(ws):
    """
    Bouwt de Acties-sheet: actielijst voor opvolging van maatregelen en controles.

    Acties worden automatisch aangemaakt door de Streamlit-app (bv. bij het
    plannen van een maatregel of het uitvoeren van een periodieke controle).

    Structuur
    ---------
    Rij 1  : titelrij
    Rij 2  : kolomkoppen
    Rij 3+ : data

    Kolommen
    --------
    A  Actie ID  | B  Type  | C  DA  | D  Control ID  | E  Omschrijving
    F  Vervaldatum  | G  Verantwoordelijke  | H  Status  | I  Aangemaakt op
    """
    no_gridlines(ws)
    COLS = [
        ("Actie ID",            10),
        ("Type",                20),
        ("DA",                  25),
        ("Control ID",          14),
        ("Omschrijving",        60),
        ("Vervaldatum",         14),
        ("Verantwoordelijke",   25),
        ("Status",              18),
        ("Aangemaakt op",       16),
    ]
    N = len(COLS)
    LAST = get_column_letter(N)

    set_col_widths(ws, [w for _, w in COLS])

    # Titelrij
    ws.merge_cells(f"A1:{LAST}1")
    c = ws["A1"]
    c.value = "Acties — Opvolging van maatregelen en periodieke controles"
    c.fill = fill("navy"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[1].height = 34

    # Kolomkoppen
    for col, (hdr, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True)
        c.border = border_all("navy")
    ws.row_dimensions[2].height = 30

    # Dropdown: Type (kolom B = 2)
    dv_type = DataValidation(
        type="list",
        formula1='"Periodieke controle,Maatregel,Opvolging"',
        showDropDown=False, allow_blank=True
    )
    dv_type.sqref = "B3:B2000"
    ws.add_data_validation(dv_type)

    # Dropdown: Status (kolom H = 8)
    dv_status = DataValidation(
        type="list",
        formula1='"Open,In uitvoering,Gesloten"',
        showDropDown=False, allow_blank=True
    )
    dv_status.sqref = "H3:H2000"
    ws.add_data_validation(dv_status)

    # AutoFilter + freeze
    ws.auto_filter.ref = f"A2:{LAST}2"
    ws.freeze_panes = "A3"

    # Datumopmaak voor kolommen F en I
    for col_letter in ("F", "I"):
        for r in range(3, 2001):
            ws.cell(row=r, column=ord(col_letter) - 64).number_format = "DD/MM/YYYY"

    # Conditionele opmaak: Open = lichtgeel, In uitvoering = lichtblauw, Gesloten = lichtgroen
    from openpyxl.formatting.rule import CellIsRule
    status_col = "H"
    ws.conditional_formatting.add(
        f"{status_col}3:{status_col}2000",
        CellIsRule(operator="equal", formula=['"Open"'],
                   fill=PatternFill("solid", fgColor="FEF9C3"))
    )
    ws.conditional_formatting.add(
        f"{status_col}3:{status_col}2000",
        CellIsRule(operator="equal", formula=['"In uitvoering"'],
                   fill=PatternFill("solid", fgColor="DBEAFE"))
    )
    ws.conditional_formatting.add(
        f"{status_col}3:{status_col}2000",
        CellIsRule(operator="equal", formula=['"Gesloten"'],
                   fill=PatternFill("solid", fgColor="DCFCE7"))
    )

    # Placeholder
    c = ws.cell(row=3, column=1,
                value="(leeg — acties worden aangemaakt via de Streamlit-app)")
    c.font = font(9, color="subtext", italic=True)
    c.alignment = align("left", "center")


# ── CyFun Controls sheet ─────────────────────────────────────────────────────
# Kolomindeling:  A–F = 2025 data  |  G = Versie  |  H–M = 2023 data
COL_25_START = 1   # A
COL_VERSIE   = 7   # G
COL_23_START = 8   # H

def _norm_id(s):
    """
    Normaliseert een CyFun ID-string naar een vergelijkbaar lowercase formaat.
    Converteert trailing -N naar .N zodat 'ID.AM-03-3' → 'id.am-03.3'.
    Wordt gebruikt als sleutel in both_map/only25 voor consistente vergelijkingen.

    Parameters
    ----------
    s : str — raw ID-string (kan hoofdletters en verschillende delimiters bevatten)

    Geeft terug
    -----------
    str — genormaliseerde ID (lowercase, trailing koppelteken vervangen door punt)
    """
    return re.sub(r'-(\d+)$', r'.\1', str(s).strip().lower())

def _load_all_mappings():
    """
    Laadt de 2023↔2025 control-mapping vanuit het MAPPING_SRC-werkboek.
    Het werkboek bevat 3 sheets (CyFun 2023>25 BASIC, IMPORTANT, ESSENTIAL).

    BASIC-sheet heeft geen DELETED/NEW-kolommen; kolomindeling is anders:
      [0]=2023_ID, [2]=km23, [4]=2025_ID
    IMPORTANT/ESSENTIAL hebben wel DELETED/NEW-flags:
      [0]=2023_ID, [2]=km23, [3]=DELETED, [5]=2025_ID, [8]=New

    Geeft terug
    -----------
    tuple (both_map, only25, only23) waarbij:
      both_map : dict{norm_2025_id → "2023_id_raw"} — controls die in beide versies bestaan
      only25   : set(norm_2025_id) — nieuwe controls in 2025 (geen 2023-equivalent)
      only23   : list[dict{id, ctrl_linked, key_meas}] — verwijderde/enkel-2023 controls
    """
    both_map, only25, only23 = {}, set(), []
    if not MAPPING_SRC.exists():
        return both_map, only25, only23

    wb = load_workbook(str(MAPPING_SRC), read_only=True, data_only=True)

    for sheet in ["CyFun 2023>25 BASIC", "CyFun 2023>25 IMPORTANT", "CyFun 2023>25 ESSENTIAL"]:
        ws = wb[sheet]
        is_basic = sheet.endswith("BASIC")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_basic:
                # BASIC: [2023_ID, req23, km23, score, 2025_ID, req25, km25, score]
                id23, km23  = row[0], row[2]
                id25        = row[4]
                deleted     = False
                new_flag    = False
            else:
                # IMPORTANT/ESSENTIAL: [2023_ID, req23, km23, DELETED, score, 2025_ID, req25, km25, NEW, score]
                id23, km23  = row[0], row[2]
                deleted     = row[3] == "DELETED"
                id25        = row[5]
                new_flag    = row[8] == "New"

            # Lege rijen overslaan
            if id23 is None and id25 is None:
                continue

            km23_s = str(km23).strip() if km23 else ""

            if deleted:
                only23.append({
                    "id":          str(id23).strip(),
                    "ctrl_linked": km23_s if "Control" in km23_s else "",
                    "key_meas":    "Key Measure" if "Key" in km23_s else "",
                })
            elif new_flag:
                if id25:
                    only25.add(_norm_id(id25))
            else:
                if id23 and id25:
                    both_map[_norm_id(id25)] = str(id23).strip()

    wb.close()
    return both_map, only25, only23

def _load_2023_details(target_ids):
    """
    Zoekt details op voor een set van 2023-control IDs in de Details-sheets
    van het CyFun 2023 bronwerkboek (CYFUN23_SRC).

    De Details-sheets ESSENTIAL Details, IMPORTANT Details en BASIC Details
    worden in die volgorde doorzocht. Een control wordt niet dubbel geladen
    (req_id in found → overslaan).
    Assurance-niveau bepaling:
      - prefix "BASIC_"     → Basic
      - prefix "IMPORTANT_" → Important
      - geen prefix         → niveau van de sheet (Essential / Important / Basic)

    Parameters
    ----------
    target_ids : set(str) — ruwe 2023 requirement IDs (bv. "RS.CO-3.2")

    Geeft terug
    -----------
    dict{req_id → {category, subcategory, assurance, requirement, key_meas, ctrl_linked}}
    """
    found = {}
    if not CYFUN23_SRC.exists():
        return found

    wb = load_workbook(str(CYFUN23_SRC), read_only=True, data_only=True)

    # ESSENTIAL Details eerst: heeft expliciete BASIC_/IMPORTANT_ prefixes voor alle niveaus.
    # Daarna IMPORTANT Details en BASIC Details als fallback voor controls die enkel daar staan.
    # Als een control geen prefix heeft: niveau = wat de sheet aangeeft (Essential/Important/Basic).
    sheet_default = {
        "ESSENTIAL Details": "Essential",
        "IMPORTANT Details": "Important",
        "BASIC Details":     "Basic",
    }
    for sheet in ["ESSENTIAL Details", "IMPORTANT Details", "BASIC Details"]:
        ws = wb[sheet]
        cur_cat, cur_sub = "", ""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[1]: cur_cat = str(row[1]).strip()
            if row[3]: cur_sub = str(row[3]).strip()
            if not row[4]:
                continue
            req_str = str(row[4]).strip()
            clean   = re.sub(r'^(BASIC_|IMPORTANT_)', '', req_str)
            parts   = clean.split(":", 1)
            req_id  = parts[0].strip()
            if req_id not in target_ids or req_id in found:
                continue
            if req_str.startswith("BASIC_"):
                level = "Basic"
            elif req_str.startswith("IMPORTANT_"):
                level = "Important"
            else:
                level = sheet_default[sheet]   # geen prefix → niveau van de sheet
            req_text = f"{req_id}: {parts[1].strip()}" if len(parts) > 1 else req_str
            km_raw   = str(row[2]).strip() if row[2] else ""
            key_m    = "Key Measure" if "Key" in km_raw else ""
            ctrl_l   = km_raw if "Control" in km_raw else ""
            found[req_id] = {
                "category":    cur_cat,
                "subcategory": cur_sub,
                "assurance":   level,
                "requirement": req_text.replace("\n", " ").replace("\t", ""),
                "key_meas":    key_m,
                "ctrl_linked": ctrl_l,
            }

    wb.close()
    return found

# STATUS_STYLE: kleuring per versie-status in de CyFun Controls-sheet.
# "Beide"     : control bestaat in 2023 én 2025 → groen
# "Enkel 2025": nieuw in 2025 (geen 2023-equivalent) → blauw
# "Enkel 2023": verwijderd uit 2025 → oranje
STATUS_STYLE = {
    "Beide":      ("green_light",  "green"),
    "Enkel 2025": ("blue_xlight",  "blue_mid"),
    "Enkel 2023": ("orange_light", "orange"),
}

def _write_side(ws, data_row, col_start, vals6, assurance, row_bg, assurance_style):
    """
    Schrijft 6 datacellen voor één zijde van de CyFun Controls vergelijkingstabel
    (2025-zijde: kolommen A-F of 2023-zijde: kolommen H-M).

    Kolomvolgorde (1-gebaseerd binnen de 6 cellen):
      1 = Category        4 = Subcategory
      2 = Controls linked 5 = Assurance level (gekleurde badge)
      3 = Key Measure     6 = Requirement (volledige tekst, wrap)

    Parameters
    ----------
    ws              : openpyxl.Worksheet
    data_row        : int — rijnummer
    col_start       : int — startkolom (COL_25_START of COL_23_START)
    vals6           : list[str] | None — 6 celwaarden; None schrijft lege cellen
    assurance       : str — assurance-niveau voor kleurstijl ("Basic" / "Important" / "Essential")
    row_bg          : str — achtergrondkleursleutel voor wisselende strepen
    assurance_style : dict — mapping assurance → (bg_key, fg_key)
    """
    bg_key, fg_key = assurance_style.get(assurance, ("white", "text"))
    values = vals6 if vals6 is not None else [""] * 6
    for i, val in enumerate(values):
        col = col_start + i
        pos = i + 1                  # 1=Cat 2=Ctrl 3=KM 4=Sub 5=Ass 6=Req
        c = ws.cell(row=data_row, column=col, value=val if val else None)
        c.border = border_all("grey_border")
        if not val:
            c.fill = fill(row_bg)
            c.alignment = align("left", "top")
        elif pos == 5:               # Assurance level
            c.fill = fill(bg_key)
            c.font = font(9, bold=True, color=fg_key)
            c.alignment = align("center", "top")
        elif pos == 3:               # Key Measure badge
            c.fill = fill("accent_light")
            c.font = font(9, bold=True, color="accent")
            c.alignment = align("center", "top")
        else:
            c.fill = fill(row_bg)
            c.font = font(9, color="text")
            c.alignment = align("left", "top", wrap=True)

def _write_versie(ws, data_row, status):
    """
    Schrijft de Versie-cel (kolom G) voor één datarij in de CyFun Controls-sheet.
    De kleur wordt bepaald door STATUS_STYLE op basis van de status-string.

    Parameters
    ----------
    ws       : openpyxl.Worksheet
    data_row : int — rijnummer
    status   : str — "Beide", "Enkel 2025" of "Enkel 2023"
    """
    sbg, sfg = STATUS_STYLE.get(status, ("white", "text"))
    c = ws.cell(row=data_row, column=COL_VERSIE, value=status)
    c.border = border_all("grey_border")
    c.fill = fill(sbg)
    c.font = font(9, bold=True, color=sfg)
    c.alignment = align("center", "top")

def build_cyfun_controls(ws):
    """
    Bouwt de CyFun Controls vergelijkingssheet met een zij-aan-zij overzicht
    van CyFun 2025 ESSENTIAL controls (kolommen A-F) en hun 2023-equivalenten
    (kolommen H-M), gescheiden door een Versie-kolom (G).

    Structuur:
      Rij 1  : Titelregel
      Rij 2  : Groepskoppen "CyFun 2025" (A-F), "Versie" (G), "CyFun 2023" (H-M)
      Rij 3  : Kolomkoppen (13 kolommen)
      Rij 4+ : Data (2025-controls uit CYFUN_SRC, met bijhorende 2023-gegevens)
      Scheidingsregel + Enkel-2023 controls onderaan

    Parameters
    ----------
    ws : openpyxl.Worksheet — het CyFun Controls-werkblad
    """
    CYFUN_TABS = ["GOVERN", "IDENTIFY", "PROTECT", "DETECT", "RESPOND", "RECOVER"]
    # 6 kolomnamen per zijde
    HDR_25 = ["Category",
              "Controls linked to\nthe management aspects",
              "Key Measure",
              "Subcategory",
              "Assurance level",
              "Requirement"]
    HDR_23 = ["Category (2023)",
              "Controls linked (2023)",
              "Key Measure (2023)",
              "Subcategory (2023)",
              "Assurance level (2023)",
              "Requirement (2023)"]
    # Kolombreedtes: 6×(2025) + Versie + 6×(2023)
    COL_WIDTHS = [30, 24, 14, 36, 16, 90,   # A–F
                  13,                         # G
                  30, 24, 14, 36, 16, 90]     # H–M
    N        = len(COL_WIDTHS)               # 13
    LAST_COL = get_column_letter(N)
    ASSURANCE_STYLE = {
        "Basic":     ("green_light",  "green"),
        "Important": ("yellow_light", "yellow"),
        "Essential": ("red_light",    "red"),
    }

    # ── Mapping + 2023-details laden ─────────────────────────────────────────
    both_map, only25_set, only23_meta = _load_all_mappings()
    # Laad 2023-details voor ALLE gekoppelde IDs + only-2023
    all_2023_ids = set(both_map.values()) | {d["id"] for d in only23_meta}
    details_2023 = _load_2023_details(all_2023_ids)

    # ── Opmaak ───────────────────────────────────────────────────────────────
    no_gridlines(ws)
    set_col_widths(ws, COL_WIDTHS)

    # Rij 1: titelregel
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    c.value = "CyFun 2025 + 2023 — Controls & Requirements (ESSENTIAL) — Vergelijking"
    c.fill = fill("navy"); c.font = font(13, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[1].height = 34

    # Rij 2: groepskoppen
    GRP = 2
    ws.row_dimensions[GRP].height = 22
    lc25 = get_column_letter(COL_25_START + 5)
    lc23 = get_column_letter(COL_23_START + 5)
    ws.merge_cells(f"A{GRP}:{lc25}{GRP}")
    c = ws[f"A{GRP}"]
    c.value = "CyFun 2025"; c.fill = fill("navy")
    c.font = font(10, bold=True, color="white"); c.alignment = align("center", "center")
    c = ws.cell(row=GRP, column=COL_VERSIE, value="Versie")
    c.fill = fill("grey_dark"); c.font = font(10, bold=True, color="white")
    c.alignment = align("center", "center")
    ws.merge_cells(f"H{GRP}:{lc23}{GRP}")
    c = ws[f"H{GRP}"]
    c.value = "CyFun 2023"; c.fill = fill("blue_mid")
    c.font = font(10, bold=True, color="white"); c.alignment = align("center", "center")

    # Rij 3: kolomkoppen
    HDR = 3
    ws.row_dimensions[HDR].height = 36
    for i, h in enumerate(HDR_25):
        c = ws.cell(row=HDR, column=COL_25_START + i, value=h)
        c.fill = fill("navy"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")
    c = ws.cell(row=HDR, column=COL_VERSIE, value="Versie")
    c.fill = fill("grey_dark"); c.font = font(10, bold=True, color="white")
    c.alignment = align("center", "center"); c.border = border_all("grey_dark")
    for i, h in enumerate(HDR_23):
        c = ws.cell(row=HDR, column=COL_23_START + i, value=h)
        c.fill = fill("blue_mid"); c.font = font(10, bold=True, color="white")
        c.alignment = align("center", "center", wrap=True); c.border = border_all("navy")

    ws.auto_filter.ref = f"A{HDR}:{LAST_COL}{HDR}"
    ws.freeze_panes   = f"A{HDR + 1}"

    if not CYFUN_SRC.exists():
        ws[f"A{HDR+1}"].value = f"BESTAND NIET GEVONDEN: {CYFUN_SRC}"
        return

    # ── 2025 requirements ────────────────────────────────────────────────────
    src_wb   = load_workbook(str(CYFUN_SRC), read_only=True, data_only=True)
    data_row = HDR + 1

    for tab in CYFUN_TABS:
        if tab not in src_wb.sheetnames:
            continue
        src_ws  = src_wb[tab]
        cur_cat = ""
        cur_sub = ""
        for row in src_ws.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                cur_cat = str(row[0]).strip()
                cur_sub = ""
            if row[3] is not None:
                cur_sub = str(row[3]).strip()
            if row[5] is None:
                continue
            ctrl_linked = str(row[1]).strip() if (row[1] is not None and isinstance(row[1], str)) else ""
            key_meas    = "Key Measure" if (row[2] is not None and "Key" in str(row[2])) else ""
            assurance   = str(row[4]).strip() if row[4] is not None else ""
            req_text    = str(row[5]).strip().replace("\n", " ").replace("\t", "")
            req_id_norm = _norm_id(req_text.split(":")[0])

            if req_id_norm in both_map:
                status   = "Beide"
                id23     = both_map[req_id_norm]
                d23      = details_2023.get(id23, {})
                vals_23  = [d23.get("category", ""),
                            d23.get("ctrl_linked", ""),
                            d23.get("key_meas",    ""),
                            d23.get("subcategory", ""),
                            d23.get("assurance",   ""),
                            d23.get("requirement", "")]
                ass_23   = d23.get("assurance", "")
            else:
                status   = "Enkel 2025"
                vals_23  = None
                ass_23   = ""

            row_bg = "grey_light" if data_row % 2 == 0 else "white"
            vals_25 = [cur_cat, ctrl_linked, key_meas, cur_sub, assurance, req_text]
            _write_side(ws, data_row, COL_25_START, vals_25, assurance, row_bg, ASSURANCE_STYLE)
            _write_versie(ws, data_row, status)
            _write_side(ws, data_row, COL_23_START, vals_23, ass_23, row_bg, ASSURANCE_STYLE)
            data_row += 1

    src_wb.close()

    # ── Scheidingsregel ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{data_row}:{LAST_COL}{data_row}")
    c = ws.cell(row=data_row, column=1,
                value="CyFun 2023 — Controls niet opgenomen in versie 2025  (Enkel 2023)")
    c.fill = fill("orange"); c.font = font(11, bold=True, color="white")
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[data_row].height = 24
    data_row += 1

    # ── 2023-only requirements (A–F leeg, H–M gevuld) ────────────────────────
    for meta in only23_meta:
        rid    = meta["id"]
        d      = details_2023.get(rid, {})
        ass    = d.get("assurance",   "Essential")
        vals_23 = [d.get("category",    f"(2023) {rid}"),
                   meta["ctrl_linked"] or d.get("ctrl_linked", ""),
                   meta["key_meas"]    or d.get("key_meas",    ""),
                   d.get("subcategory", ""),
                   ass,
                   d.get("requirement", rid)]
        row_bg = "grey_light" if data_row % 2 == 0 else "white"
        _write_side(ws, data_row, COL_25_START, None,    "",  row_bg, ASSURANCE_STYLE)
        _write_versie(ws, data_row, "Enkel 2023")
        _write_side(ws, data_row, COL_23_START, vals_23, ass, row_bg, ASSURANCE_STYLE)
        data_row += 1

ws_cyfun = wb.create_sheet("CyFun Controls")
build_cyfun_controls(ws_cyfun)

ws_kwets = wb.create_sheet("Kwetsbaarheden")
afw_rows = build_kwetsbaarheden(ws_kwets)

ws_afw = wb.create_sheet("Afwijkingen")
build_afwijkingen(ws_afw, afw_rows)

# ── Nieuwe sheets: Risicobeheer | Controles | Acties ──────────────────────────
ws_risico = wb.create_sheet("Risicobeheer")
build_risicobeheer(ws_risico)

ws_controles = wb.create_sheet("Controles")
build_controles(ws_controles)

ws_acties = wb.create_sheet("Acties")
build_acties(ws_acties)

# ══════════════════════════════════════════════════════════════════════════════
# OPSLAAN ALS XLSX + VBA INJECTEREN VIA WIN32COM → XLSM
# ══════════════════════════════════════════════════════════════════════════════
# Stap 1: openpyxl kan geen .xlsm bestanden met macro's opslaan.
#         Sla eerst op als .xlsx (zonder macro's) als tussenstap.
# Stap 2: open het .xlsx-bestand via win32com (COM-automatisering met Excel),
#         injecteer alle VBA-code en sla op als .xlsm.
# Stap 3: verwijder het .xlsx-tussenstap.

OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
# _Lang werd als eerste aangemaakt (wb.active) — verschuif naar het einde van
# de tabbladen zodat het verborgen blad niet het eerste tabblad is na openen.
wb.move_sheet("_Lang", offset=len(wb.sheetnames) - 1)
wb.save(OUT_XLSX)
print(f"Tussenstap opgeslagen: {OUT_XLSX}")

# ── VBA-injectie via win32com ─────────────────────────────────────────────────
vba_ok = False
try:
    import win32com.client as win32
    import pythoncom

    pythoncom.CoInitialize()
    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    wb_com = xl.Workbooks.Open(str(OUT_XLSX.resolve()))

    # ── Voeg GRC_Macros standaardmodule toe ──────────────────────────────────
    # vbext_ct_StdModule = 1 (standaard codemodule)
    # Attribute VB_Name is VBE-bestandsmetadata en wordt afzonderlijk gezet
    # via mod.Name; het mag NIET in de AddFromString-aanroep staan
    # (veroorzaakt een VBE-compilatiefout).
    mod = wb_com.VBProject.VBComponents.Add(1)   # 1 = vbext_ct_StdModule
    mod.Name = "GRC_Macros"
    code_for_vba = "\n".join(
        line for line in VBA_CODE.splitlines()
        if not line.startswith("Attribute VB_Name")
    )
    mod.CodeModule.AddFromString(code_for_vba)

    # ── Actieknoppen op Import & Export-blad ─────────────────────────────────
    # De knoppen worden als msoShapeRectangle (1) toegevoegd; positie en afmetingen
    # zijn in punten (1 punt ≈ 0.035 cm). Elke knop krijgt een OnAction-macro.
    # De opmaakstijl (kleur, lettertype) bootst de knop-placeholder-cellen na
    # die in de openpyxl-fase zijn aangemaakt.
    ws_ie_com = wb_com.Sheets("Import & Export")
    def add_button(ws, left, top, w, h, caption, macro):
        """Voegt een rechthoekige klikknop toe aan het opgegeven werkblad."""
        btn = ws.Shapes.AddShape(1, left, top, w, h)   # 1 = msoShapeRectangle
        btn.Fill.ForeColor.RGB = int("2563EB", 16)
        btn.Line.Visible = False
        btn.TextFrame.Characters().Text = caption
        btn.TextFrame.Characters().Font.Bold = True
        btn.TextFrame.Characters().Font.Size = 11
        btn.TextFrame.Characters().Font.Color = int("FFFFFF", 16)
        btn.TextFrame.HorizontalAlignment = -4108   # xlHAlignCenter
        btn.TextFrame.VerticalAlignment = -4108     # xlVAlignCenter
        btn.OnAction = macro

    add_button(ws_ie_com, 10, 130, 200, 44, "▶  Importeer Alles",                    "GRC_Macros.ImportAlles")
    add_button(ws_ie_com, 10, 218, 200, 44, "▶  Importeer Informatieassets",          "GRC_Macros.ImportInformatieassets")
    add_button(ws_ie_com, 10, 306, 200, 44, "▶  Importeer Afhankelijke Assets",       "GRC_Macros.ImportAfhankelijkeAssets")
    add_button(ws_ie_com, 10, 394, 200, 44, "▶  Importeer Processen",                 "GRC_Macros.ImportProcessen")
    add_button(ws_ie_com, 10, 482, 200, 44, "▶  Importeer Kwetsbaarheden",            "GRC_Macros.ImportKwetsbaarheden")
    add_button(ws_ie_com, 10, 570, 200, 44, "▶  Importeer Links DA/Kwetsbaarheden",   "GRC_Macros.ImportLinksKwetsbaarheden")

    # ── UserForm AssetPicker ──────────────────────────────────────────────────
    # vbext_ct_MSForm = 3 (UserForm component)
    # Controls worden programmatisch toegevoegd via uf.Designer.Controls.Add().
    # De VBA-eventcode (USERFORM_CODE) wordt achteraf via CodeModule geïnjecteerd.
    uf = wb_com.VBProject.VBComponents.Add(3)   # 3 = vbext_ct_MSForm
    uf.Name = "AssetPicker"
    uf.Properties("Caption").Value = "Selecteer Informatieassets"
    uf.Properties("Width").Value  = 340
    uf.Properties("Height").Value = 360

    # Label bovenaan
    lbl = uf.Designer.Controls.Add("Forms.Label.1")
    lbl.Name    = "lblTitel"
    lbl.Caption = "Selecteer de gekoppelde assets:"
    lbl.Left    = 10; lbl.Top = 10; lbl.Width = 300; lbl.Height = 18
    lbl.Font.Bold = True

    # ListBox met checkboxen (MultiSelect = 1, ListStyle = 1)
    lb = uf.Designer.Controls.Add("Forms.ListBox.1")
    lb.Name        = "lstAssets"
    lb.Left        = 10; lb.Top = 34; lb.Width = 306; lb.Height = 252
    lb.MultiSelect = 1   # fmMultiSelectMulti: meerdere items tegelijk selecteerbaar
    lb.ListStyle   = 1   # fmListStyleOption: toont checkboxen naast elk item

    # OK-knop (groen)
    btnOK = uf.Designer.Controls.Add("Forms.CommandButton.1")
    btnOK.Name    = "btnOK"
    btnOK.Caption = "OK"
    btnOK.Left    = 10;  btnOK.Top = 296; btnOK.Width = 146; btnOK.Height = 28
    btnOK.BackColor = int("15803D", 16)
    btnOK.ForeColor = int("FFFFFF", 16)

    # Annuleer-knop (rood)
    btnAnn = uf.Designer.Controls.Add("Forms.CommandButton.1")
    btnAnn.Name    = "btnAnnuleer"
    btnAnn.Caption = "Annuleer"
    btnAnn.Left    = 170; btnAnn.Top = 296; btnAnn.Width = 146; btnAnn.Height = 28
    btnAnn.BackColor = int("B91C1C", 16)
    btnAnn.ForeColor = int("FFFFFF", 16)

    # Voeg UserForm code toe
    uf.CodeModule.AddFromString(USERFORM_CODE)

    # ── Sheet-event modules injecteren ───────────────────────────────────────
    # Elk werkblad heeft een eigen VBComponent (Sheet-module) dat bereikbaar is
    # via wb_com.Sheets(naam).CodeName (bv. "Sheet1", "Sheet2", …).
    # De event-handlers worden rechtstreeks in dat component geïnjecteerd.

    # Processes: SelectionChange → AssetPicker-popup voor kolom K/L
    proc_code_name = wb_com.Sheets("Processes").CodeName
    proc_mod = wb_com.VBProject.VBComponents(proc_code_name)
    proc_mod.CodeModule.AddFromString(PROC_SHEET_CODE)

    # Information Assets: Activate → herbereken "Gebruikt in processen" (col I)
    ia_code_name = wb_com.Sheets("Information Assets").CodeName
    ia_mod = wb_com.VBProject.VBComponents(ia_code_name)
    ia_mod.CodeModule.AddFromString(INFO_ASSET_SHEET_CODE)

    # Dependent Assets: Activate + Change + BeforeDoubleClick
    dep_code_name = wb_com.Sheets("Dependent Assets").CodeName
    dep_mod = wb_com.VBProject.VBComponents(dep_code_name)
    dep_mod.CodeModule.AddFromString(DEP_ASSET_SHEET_CODE)

    # Kwetsbaarheden: BeforeDoubleClick → ✔-toggle in matrix
    kwets_code_name = wb_com.Sheets("Kwetsbaarheden").CodeName
    kwets_mod = wb_com.VBProject.VBComponents(kwets_code_name)
    kwets_mod.CodeModule.AddFromString(KWETS_SHEET_CODE)

    # ── UserForm VulnPicker ───────────────────────────────────────────────────
    # Complexer formulier dan AssetPicker: bevat een scrollbaar Frame (frmVulns)
    # dat dynamisch in VBA wordt gevuld met CheckBox + ComboBox per kwetsbaarheid.
    # Afmetingen zijn in punten (+28 pt ≈ 1 cm extra voor marges).
    vp = wb_com.VBProject.VBComponents.Add(3)   # 3 = vbext_ct_MSForm
    vp.Name = "VulnPicker"
    vp.Properties("Caption").Value = "Kwetsbaarheden selecteren"
    vp.Properties("Width").Value  = 458   # +28 pt (≈1 cm) rechts
    vp.Properties("Height").Value = 514   # +28 pt (≈1 cm) onder + 16 pt rand onder knoppen
    vp_des = vp.Designer
    # Titel-label
    lbl_vp = vp_des.Controls.Add("Forms.Label.1")
    lbl_vp.Name = "lblTitel"; lbl_vp.Caption = "Kwetsbaarheden"
    lbl_vp.Left = 8; lbl_vp.Top = 8; lbl_vp.Width = 426; lbl_vp.Height = 20
    lbl_vp.Font.Bold = True
    # Kolomkopjes
    lbl_vuln_hdr = vp_des.Controls.Add("Forms.Label.1")
    lbl_vuln_hdr.Name = "lblVulnHdr"; lbl_vuln_hdr.Caption = "Kwetsbaarheid"
    lbl_vuln_hdr.Left = 8; lbl_vuln_hdr.Top = 30; lbl_vuln_hdr.Width = 216; lbl_vuln_hdr.Height = 14
    lbl_vuln_hdr.Font.Italic = True
    lbl_prob_hdr = vp_des.Controls.Add("Forms.Label.1")
    lbl_prob_hdr.Name = "lblProbHdr"; lbl_prob_hdr.Caption = "Kans"
    lbl_prob_hdr.Left = 228; lbl_prob_hdr.Top = 30; lbl_prob_hdr.Width = 198; lbl_prob_hdr.Height = 14
    lbl_prob_hdr.Font.Italic = True
    # Scrollable Frame voor CheckBox + ComboBox rijen (inhoud wordt dynamisch gevuld in VBA)
    frm_vp = vp_des.Controls.Add("Forms.Frame.1")
    frm_vp.Name = "frmVulns"
    frm_vp.Caption = ""
    frm_vp.Left = 8; frm_vp.Top = 46; frm_vp.Width = 426; frm_vp.Height = 388
    frm_vp.ScrollBars = 2          # fmScrollBarsVertical: enkel verticaal scrollen
    frm_vp.KeepScrollBarsVisible = 2  # toon scrollbalk alleen wanneer inhoud de frame overschrijdt
    # Annuleer-knop (links)
    btn_vp_ann = vp_des.Controls.Add("Forms.CommandButton.1")
    btn_vp_ann.Name = "btnAnnuleer"; btn_vp_ann.Caption = "Annuleer"
    btn_vp_ann.Left = 8; btn_vp_ann.Top = 446; btn_vp_ann.Width = 120; btn_vp_ann.Height = 30
    btn_vp_ann.BackColor = int("C0504D", 16)
    btn_vp_ann.ForeColor = int("FFFFFF", 16)
    # OK-knop (rechts)
    btn_vp_ok = vp_des.Controls.Add("Forms.CommandButton.1")
    btn_vp_ok.Name = "btnOK"; btn_vp_ok.Caption = "OK"
    btn_vp_ok.Left = 322; btn_vp_ok.Top = 446; btn_vp_ok.Width = 120; btn_vp_ok.Height = 30
    btn_vp_ok.BackColor  = int("70AD47", 16)
    btn_vp_ok.ForeColor  = int("FFFFFF", 16)
    vp.CodeModule.AddFromString(VULNPICKER_CODE)

    # RARM: Activate + BeforeDoubleClick + SelectionChange (VulnPicker)
    rarm_code_name = wb_com.Sheets("RARM").CodeName
    rarm_mod = wb_com.VBProject.VBComponents(rarm_code_name)
    rarm_mod.CodeModule.AddFromString(RARM_SHEET_CODE)

    # ── Opslaan als .xlsm en opruimen ────────────────────────────────────────
    # FileFormat=52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
    # Close(False) = niet nogmaals opslaan bij sluiten (voorkomen dubbele dialoog)
    wb_com.SaveAs(str(OUT.resolve()), FileFormat=52)
    wb_com.Close(False)
    xl.Quit()
    pythoncom.CoUninitialize()

    OUT_XLSX.unlink()   # verwijder het .xlsx-tussenstap — enkel het .xlsm blijft over
    vba_ok = True
    print(f"GRC Tool v0.3 (macro-enabled) opgeslagen: {OUT}")

except ImportError:
    print("INFO: pywin32 niet gevonden — bestand opgeslagen als .xlsx zonder macro's.")
    print(f"      Installeer pywin32 via: pip install pywin32")
except Exception as e:
    print(f"WAARSCHUWING: VBA-injectie mislukt ({e})")
    print(f"              Bestand opgeslagen als .xlsx: {OUT_XLSX}")

print(f"Gegenereerd door: {USERNAME} op {NOW_STR}")
